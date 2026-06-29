import re
from pathlib import Path
from typing import Any


HEADER_ALIASES = {
    "user_field_name": {"user field name", "source field name", "source_field_key", "field_key", "field key"},
    "acf_field_name": {"acf field name", "acf_field_name"},
    "shared_field_name": {"field name"},
    "ai_guidance": {"ai guidance", "guidance", "agent guidance", "field guidance", "prompt guidance", "guidance_de"},
    "description": {"description"},
    "group": {"group"},
    "section": {"section"},
    "required": {"required"},
    "source_type": {"source_type", "source type"},
    "value_type": {"value_type", "value type"},
    "format_or_enum": {"format_or_enum", "format or enum"},
    "output_target": {"output_target", "output target"},
    "example": {"example"},
    "validation_rule": {"validation_rule", "validation rule"},
    "enabled": {"enabled"},
    "min_words": {"min_words", "minimum words", "min words"},
    "max_words": {"max_words", "maximum words", "max words"},
}

PROCESS_TAB_NAME = "process + technical infos"
SHARED_GUIDANCE_TAB = "output_technical"
POST_TYPES_TAB = "post_types"
EVENT_FIELD_SCHEMA_TAB = "event_field_schema"
# Tabs treated as schema/mapping sources in addition to the main ACF tab.
# Each entry is the exact tab name or a normalized (lowercase, no spaces) alias.
SCHEMA_TAB_CANDIDATES = {
    "event_field_schema",
    "event_acf_mapping",
    "shared_field_schema",
    "location_field_schema",
    "location_acf_mapping",
    "cocktail_field_schema",
    "cocktail_acf_mapping",
    "bartender_field_schema",
    "bartender_acf_mapping",
}

POST_TYPE_SCHEMA_TABS = {
    "event": {"event_field_schema", "event_acf_mapping", "shared_field_schema"},
    "location": {"location_field_schema", "location_acf_mapping", "shared_field_schema"},
    "cocktail": {"cocktail_field_schema", "cocktail_acf_mapping", "shared_field_schema"},
    "bartender": {"bartender_field_schema", "bartender_acf_mapping", "shared_field_schema"},
}

POST_TYPES_REQUIRED_COLUMNS = {
    "post_type_key",
    "enabled",
    "blueprint_sheet",
    "examples_sheet",
    "user_selectable",
}

INTERNAL_LINKS_DATABASE_TAB = "internal_links_database"
INTERNAL_LINK_RULES_TAB = "internal_link_rules"


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def normalize_lookup(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def merged_cell_value(sheet: Any, row: int, column: int) -> Any:
    cell = sheet.cell(row=row, column=column)
    if cell.value not in (None, ""):
        return cell.value

    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def get_post_type_to_acf_tab_mapping(workbook: Any) -> dict[str, str]:
    """
    Read the 'process + technical infos' tab and extract mapping of
    post_type (e.g., 'Event') to ACF tab name (e.g., '01_output ACF_Event').
    Returns dict like {'Event': '01_output ACF_Event', 'Location': '01_output ACF_Location', ...}
    """
    mapping = {}
    if PROCESS_TAB_NAME not in workbook.sheetnames:
        return mapping
    
    sheet = workbook[PROCESS_TAB_NAME]
    
    # Find a row that contains both the post-type and ACF-tab headers.
    header_row = None
    post_type_col = None
    acf_tab_col = None

    for row_idx in range(1, min(sheet.max_row + 1, 60)):
        row_post_type_col = None
        row_acf_tab_col = None
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = normalize_key(str(merged_cell_value(sheet, row_idx, col_idx) or ""))
            if cell_value in {"post type", "type of post"} or "type of post" in cell_value or "post type" in cell_value:
                row_post_type_col = col_idx
            if "acf output tabname" in cell_value or "acf tabname" in cell_value or "acf tab" in cell_value or "acf output" in cell_value:
                row_acf_tab_col = col_idx

        if row_post_type_col and row_acf_tab_col:
            header_row = row_idx
            post_type_col = row_post_type_col
            acf_tab_col = row_acf_tab_col
            break
    
    if not (header_row and post_type_col and acf_tab_col):
        return mapping
    
    # Read mapping rows
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        post_type = str(merged_cell_value(sheet, row_idx, post_type_col) or "").strip()
        acf_tab = str(merged_cell_value(sheet, row_idx, acf_tab_col) or "").strip()
        if post_type and acf_tab:
            mapping[post_type] = acf_tab
    
    return mapping


def find_header_row(sheet: Any) -> tuple[int, dict[str, int]] | None:
    for row_index in range(1, min(sheet.max_row, 30) + 1):
        headers: dict[str, int] = {}
        for column_index in range(1, sheet.max_column + 1):
            value = merged_cell_value(sheet, row_index, column_index)
            normalized = normalize_key(str(value or ""))
            if normalized in HEADER_ALIASES["user_field_name"]:
                headers["user_field_name"] = column_index
            elif normalized in HEADER_ALIASES["acf_field_name"]:
                headers["acf_field_name"] = column_index
            elif normalized in HEADER_ALIASES["shared_field_name"]:
                headers.setdefault("user_field_name", column_index)
                headers.setdefault("acf_field_name", column_index)
            elif normalized in HEADER_ALIASES["ai_guidance"]:
                headers["ai_guidance"] = column_index
            elif normalized in HEADER_ALIASES["description"]:
                headers["description"] = column_index
            elif normalized in HEADER_ALIASES["group"]:
                headers["group"] = column_index
            elif normalized in HEADER_ALIASES["section"]:
                headers["section"] = column_index
            elif normalized in HEADER_ALIASES["required"]:
                headers["required"] = column_index
            elif normalized in HEADER_ALIASES["source_type"]:
                headers["source_type"] = column_index
            elif normalized in HEADER_ALIASES["value_type"]:
                headers["value_type"] = column_index
            elif normalized in HEADER_ALIASES["format_or_enum"]:
                headers["format_or_enum"] = column_index
            elif normalized in HEADER_ALIASES["output_target"]:
                headers["output_target"] = column_index
            elif normalized in HEADER_ALIASES["example"]:
                headers["example"] = column_index
            elif normalized in HEADER_ALIASES["validation_rule"]:
                headers["validation_rule"] = column_index
            elif normalized in HEADER_ALIASES["enabled"]:
                headers["enabled"] = column_index
            elif normalized in HEADER_ALIASES["min_words"]:
                headers["min_words"] = column_index
            elif normalized in HEADER_ALIASES["max_words"]:
                headers["max_words"] = column_index
        has_length_columns = "min_words" in headers or "max_words" in headers
        has_legacy = {"user_field_name", "acf_field_name"}.issubset(headers) and ("ai_guidance" in headers or has_length_columns)
        has_schema = "user_field_name" in headers and "group" in headers and (
            "description" in headers or "output_target" in headers
        )
        if has_legacy or has_schema:
            return row_index, headers
    return None


def parse_int_cell(value: Any) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"(\d+)", text)
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def parse_boolish(value: Any, default: bool = False) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return default
    if text in {"1", "true", "yes", "on", "y"}:
        return True
    if text in {"0", "false", "no", "off", "n"}:
        return False
    return default


def parse_word_count_range(value: Any) -> tuple[int | None, int | None]:
    """Parse min/max word count from ranges like '80-100', '80', or similar."""
    if isinstance(value, (int, float)):
        number = int(round(float(value)))
        return number, None

    text = str(value or "").strip()
    if not text:
        return None, None

    decimal_match = re.fullmatch(r"(\d+)(?:\.0+)?", text)
    if decimal_match:
        return int(decimal_match.group(1)), None

    numbers = [int(m) for m in re.findall(r"\d+", text)]
    if not numbers:
        return None, None
    if len(numbers) >= 2:
        return min(numbers[0], numbers[1]), max(numbers[0], numbers[1])
    return numbers[0], None


def parse_length_range_words(value: Any) -> tuple[int | None, int | None]:
    text = str(value or "").strip().lower()
    if not text:
        return None, None
    numbers = [int(match) for match in re.findall(r"(\d+)", text)]
    if not numbers:
        return None, None
    if len(numbers) >= 2:
        start, end = numbers[0], numbers[1]
        return min(start, end), max(start, end)
    return numbers[0], None


def sheet_score(sheet_name: str, preferred_sheet: str | None) -> int:
    normalized = normalize_key(sheet_name)
    if preferred_sheet and normalize_key(preferred_sheet) == normalized:
        return 100
    score = 0
    if "acf" in normalized:
        score += 20
    if "output" in normalized:
        score += 10
    if "post" in normalized:
        score += 5
    return score


def find_post_types_header_row(sheet: Any) -> tuple[int, dict[str, int]] | None:
    for row_index in range(1, min(sheet.max_row, 40) + 1):
        headers: dict[str, int] = {}
        for column_index in range(1, sheet.max_column + 1):
            value = merged_cell_value(sheet, row_index, column_index)
            normalized = normalize_key(str(value or "")).replace(" ", "_")
            if normalized:
                headers[normalized] = column_index
        if POST_TYPES_REQUIRED_COLUMNS.issubset(headers.keys()):
            return row_index, headers
    return None


def has_post_types_schema(workbook: Any) -> bool:
    if POST_TYPES_TAB not in workbook.sheetnames:
        return False
    ws = workbook[POST_TYPES_TAB]
    header_info = find_post_types_header_row(ws)
    if not header_info:
        return False
    header_row, headers = header_info
    enabled_col = headers["enabled"]
    post_type_col = headers["post_type_key"]
    for row_index in range(header_row + 1, ws.max_row + 1):
        post_type_key = str(merged_cell_value(ws, row_index, post_type_col) or "").strip()
        enabled_raw = str(merged_cell_value(ws, row_index, enabled_col) or "").strip().lower()
        if not post_type_key:
            continue
        if enabled_raw in {"true", "1", "yes", "active", "enabled"}:
            return True
    return False


def schema_tabs_for_post_type(post_type: str | None) -> set[str]:
    if not post_type:
        return {"shared_field_schema"}
    normalized = normalize_lookup(post_type)
    for key, tabs in POST_TYPE_SCHEMA_TABS.items():
        if normalize_lookup(key) == normalized:
            return set(tabs)
    return {"shared_field_schema"}


def load_workbook_guidance(
    workbook_path: str | Path | None,
    post_type: str | None = None,
    preferred_sheet: str | None = None,
) -> dict[str, Any]:
    """
    Load guidance from workbook sheets based on post_type and shared schema.
    
    - Reads 'process + technical infos' mapping to find post-type-specific ACF tab
    - Loads guidance from the type-specific ACF tab (e.g., '01_output ACF_Event')
    - Loads guidance from shared 'output_technical' tab
    - Skips hidden sheets
    """
    if not workbook_path:
        return {"source": None, "items": [], "sheets_loaded": []}

    path = Path(workbook_path)
    if not path.exists():
        return {"source": str(path), "error": "Workbook file not found.", "items": [], "sheets_loaded": []}

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return {"source": str(path), "error": "openpyxl is not installed.", "items": [], "sheets_loaded": []}

    workbook = load_workbook(path, read_only=False, data_only=True, keep_vba=True)
    schema_mode = "post_types_v3" if has_post_types_schema(workbook) else None
    
    # Determine which sheets to load
    sheets_to_load = []
    
    # Get post-type-specific ACF tab if post_type is provided
    if post_type:
        mapping = get_post_type_to_acf_tab_mapping(workbook)
        # Resolve post type in a case/spacing tolerant way.
        requested_post_type = normalize_lookup(post_type)
        mapped_tab = None
        for mapped_post_type, tab_name in mapping.items():
            if normalize_lookup(mapped_post_type) == requested_post_type:
                mapped_tab = tab_name
                break

        if mapped_tab:
            # Resolve tab name with tolerant matching as workbook tabs can contain minor spacing variants.
            direct_tab = mapped_tab.strip()
            if direct_tab in workbook.sheetnames:
                sheets_to_load.append(direct_tab)
            else:
                normalized_to_actual = {normalize_key(name): name for name in workbook.sheetnames}
                resolved = normalized_to_actual.get(normalize_key(direct_tab))
                if resolved:
                    sheets_to_load.append(resolved)
    
    # Always load shared guidance tab if it exists
    if SHARED_GUIDANCE_TAB in workbook.sheetnames:
        sheets_to_load.append(SHARED_GUIDANCE_TAB)

    # Include only the shared schema plus the active post-type schema/mapping tabs.
    allowed_schema_tabs = schema_tabs_for_post_type(post_type)
    for sheet_name in workbook.sheetnames:
        normalized_sheet = normalize_key(sheet_name).replace(" ", "_")
        if normalized_sheet in SCHEMA_TAB_CANDIDATES and normalized_sheet in allowed_schema_tabs:
            if sheet_name not in sheets_to_load:
                sheets_to_load.append(sheet_name)
    
    # If no sheets found yet, fall back to preferred_sheet or any sheet with guidance columns
    if not sheets_to_load:
        candidates = []
        for sheet_name in workbook.sheetnames:
            # Skip hidden sheets
            ws = workbook[sheet_name]
            if ws.sheet_state == "hidden":
                continue
            
            header = find_header_row(ws)
            if header:
                score = sheet_score(sheet_name, preferred_sheet)
                candidates.append((score, sheet_name))
        
        if candidates:
            candidates.sort(key=lambda x: (-x[0], x[1]))
            sheets_to_load = [c[1] for c in candidates[:1]]  # Take highest-scoring
    
    if not sheets_to_load:
        if schema_mode:
            return {
                "source": str(path),
                "sheets_loaded": [POST_TYPES_TAB],
                "items": [],
                "schema_mode": schema_mode,
            }
        return {"source": str(path), "error": "No sheet with guidance columns found.", "items": [], "sheets_loaded": []}
    
    # Load guidance from all selected sheets
    items = []
    sheets_loaded = []
    
    for sheet_name in dict.fromkeys(sheets_to_load):
        if sheet_name not in workbook.sheetnames:
            continue
        
        ws = workbook[sheet_name]
        # Skip hidden sheets
        if ws.sheet_state == "hidden":
            continue
        
        header = find_header_row(ws)
        if not header:
            continue
        
        header_row, headers = header
        sheets_loaded.append(sheet_name)
        
        for row_index in range(header_row + 1, ws.max_row + 1):
            user_field_name = str(merged_cell_value(ws, row_index, headers["user_field_name"]) or "").strip()
            group = str(merged_cell_value(ws, row_index, headers["group"]) or "").strip() if headers.get("group") else ""
            enabled = parse_boolish(merged_cell_value(ws, row_index, headers["enabled"]), default=True) if headers.get("enabled") else True
            if not enabled:
                continue

            acf_field_name = str(merged_cell_value(ws, row_index, headers["acf_field_name"]) or "").strip() if headers.get("acf_field_name") else ""
            description = str(merged_cell_value(ws, row_index, headers["description"]) or "").strip() if headers.get("description") else ""
            ai_guidance = str(merged_cell_value(ws, row_index, headers["ai_guidance"]) or "").strip() if headers.get("ai_guidance") else ""
            if not ai_guidance:
                ai_guidance = description

            group_lookup = normalize_lookup(group)
            if not acf_field_name and user_field_name and group_lookup in {"acf", "advancedcustomfields", "advancedcustomfield", "acr"}:
                acf_field_name = user_field_name
            min_words = None
            max_words = None
            
            # Try parsing min_words with range support (e.g., "80-100" or "80")
            if headers.get("min_words"):
                min_cell = merged_cell_value(ws, row_index, headers["min_words"])
                parsed_min, parsed_max = parse_word_count_range(min_cell)
                min_words = parsed_min
                if parsed_max is not None:
                    max_words = parsed_max
            
            # Try parsing max_words separately if not already set from range
            if max_words is None and headers.get("max_words"):
                max_cell = merged_cell_value(ws, row_index, headers["max_words"])
                max_words = parse_int_cell(max_cell)
            
            # Fallback to old "length" column format
            if (min_words is None and max_words is None) and headers.get("length"):
                parsed_min, parsed_max = parse_length_range_words(merged_cell_value(ws, row_index, headers["length"]))
                min_words = parsed_min if parsed_min is not None else min_words
                max_words = parsed_max if parsed_max is not None else max_words

            # Build constraint text from min/max values (prioritize schema values)
            constraint_text = ""
            if min_words is not None and max_words is not None:
                constraint_text = f"Etwa {int((min_words + max_words) / 2)} Wörter (mindestens {min_words}, höchstens {max_words})"
            elif min_words is not None:
                constraint_text = f"Mindestens {min_words} Wörter"
            elif max_words is not None:
                constraint_text = f"Höchstens {max_words} Wörter"
            
            # Merge constraint text with guidance, avoiding word-count duplication
            if constraint_text and ai_guidance:
                # Only append if guidance doesn't already mention word counts
                if "wort" not in ai_guidance.lower():
                    ai_guidance = f"{ai_guidance}. {constraint_text}"
            elif constraint_text and not ai_guidance:
                ai_guidance = constraint_text

            if not any((user_field_name, acf_field_name, ai_guidance)):
                continue
            items.append({
                "user_field_name": user_field_name,
                "acf_field_name": acf_field_name,
                "ai_guidance": ai_guidance,
                "min_words": min_words,
                "max_words": max_words,
                "group": group,
                "section": str(merged_cell_value(ws, row_index, headers["section"]) or "").strip() if headers.get("section") else "",
                "required": parse_boolish(merged_cell_value(ws, row_index, headers["required"]), default=False) if headers.get("required") else False,
                "source_type": str(merged_cell_value(ws, row_index, headers["source_type"]) or "").strip() if headers.get("source_type") else "",
                "value_type": str(merged_cell_value(ws, row_index, headers["value_type"]) or "").strip() if headers.get("value_type") else "",
                "format_or_enum": str(merged_cell_value(ws, row_index, headers["format_or_enum"]) or "").strip() if headers.get("format_or_enum") else "",
                "output_target": str(merged_cell_value(ws, row_index, headers["output_target"]) or "").strip() if headers.get("output_target") else "",
                "example": str(merged_cell_value(ws, row_index, headers["example"]) or "").strip() if headers.get("example") else "",
                "validation_rule": str(merged_cell_value(ws, row_index, headers["validation_rule"]) or "").strip() if headers.get("validation_rule") else "",
                "enabled": enabled,
                "source_sheet": sheet_name,
            })

    payload = {
        "source": str(path),
        "sheets_loaded": sheets_loaded,
        "items": items,
    }
    if schema_mode:
        payload["schema_mode"] = schema_mode
    return payload


def guidance_for_field(guidance_data: dict[str, Any], user_field_name: str, acf_field_name: str) -> list[str]:
    """
    Return all matching AI guidance items for a field from any sheet.
    Guidance can come from output_technical, ACF tabs, or others.
    """
    items = guidance_data.get("items", [])
    user_lookup = normalize_lookup(user_field_name)
    acf_lookup = normalize_lookup(acf_field_name)
    matches: list[str] = []
    for item in items:
        if field_matches(item.get("user_field_name", ""), user_lookup) or field_matches(item.get("acf_field_name", ""), acf_lookup):
            guidance_text = str(item.get("ai_guidance") or "").strip()
            if guidance_text:
                matches.append(guidance_text)
    return matches


def guidance_items_for_field(guidance_data: dict[str, Any], user_field_name: str, acf_field_name: str) -> list[dict[str, Any]]:
    items = guidance_data.get("items", [])
    user_lookup = normalize_lookup(user_field_name)
    acf_lookup = normalize_lookup(acf_field_name)
    matches: list[dict[str, Any]] = []
    for item in items:
        if field_matches(item.get("user_field_name", ""), user_lookup):
            matches.append(item)
        elif field_matches(item.get("acf_field_name", ""), acf_lookup):
            matches.append(item)
    return matches


def field_matches(candidate_value: str, target_lookup: str) -> bool:
    candidate_lookup = normalize_lookup(candidate_value)
    if not candidate_lookup or not target_lookup:
        return False
    return candidate_lookup == target_lookup


def load_internal_links_context(workbook_path: str | Path | None) -> dict[str, Any]:
    if not workbook_path:
        return {"source": None, "database": [], "rules": []}

    path = Path(workbook_path)
    if not path.exists():
        return {"source": str(path), "error": "Workbook file not found.", "database": [], "rules": []}

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return {"source": str(path), "error": "openpyxl is not installed.", "database": [], "rules": []}

    workbook = load_workbook(path, read_only=False, data_only=True, keep_vba=True)
    database: list[dict[str, Any]] = []
    rules: list[dict[str, Any]] = []

    if INTERNAL_LINKS_DATABASE_TAB in workbook.sheetnames:
        ws = workbook[INTERNAL_LINKS_DATABASE_TAB]
        header = [normalize_key(str(ws.cell(1, col).value or "")).replace(" ", "_") for col in range(1, ws.max_column + 1)]
        idx = {name: pos for pos, name in enumerate(header)}
        for row_index in range(2, ws.max_row + 1):
            raw = [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]
            if not any(str(value or "").strip() for value in raw):
                continue

            active_idx = idx.get("active")
            active_value = raw[active_idx] if active_idx is not None else True
            if not parse_boolish(active_value, default=True):
                continue

            def get(name: str) -> str:
                pos = idx.get(name)
                if pos is None:
                    return ""
                return str(raw[pos] or "").strip()

            item = {
                "link_id": get("link_id"),
                "keyword": get("keyword"),
                "anchor_text": get("anchor_text") or get("keyword"),
                "target_url": get("target_url"),
                "category": get("category"),
                "priority": get("priority"),
                "usage_context": get("usage_context"),
                "avoid_if_current_slug": get("avoid_if_current_slug"),
                "city": get("city"),
                "service_tags": get("service_tags"),
                "event_tags": get("event_tags"),
                "notes": get("notes"),
                "active": True,
            }
            target_url = item["target_url"].lower()
            if item["target_url"] and item["anchor_text"] and target_url.startswith(("http://", "https://")):
                database.append(item)

    if INTERNAL_LINK_RULES_TAB in workbook.sheetnames:
        ws = workbook[INTERNAL_LINK_RULES_TAB]
        header = [normalize_key(str(ws.cell(1, col).value or "")).replace(" ", "_") for col in range(1, ws.max_column + 1)]
        idx = {name: pos for pos, name in enumerate(header)}
        for row_index in range(2, ws.max_row + 1):
            raw = [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]
            if not any(str(value or "").strip() for value in raw):
                continue

            enabled_idx = idx.get("enabled")
            enabled_value = raw[enabled_idx] if enabled_idx is not None else True
            if not parse_boolish(enabled_value, default=True):
                continue

            def get(name: str) -> str:
                pos = idx.get(name)
                if pos is None:
                    return ""
                return str(raw[pos] or "").strip()

            applies_to = get("applies_to").lower()
            if applies_to and applies_to != "internal_links":
                continue

            rules.append(
                {
                    "rule_id": get("rule_id"),
                    "rule_type": get("rule_type"),
                    "operator": get("operator"),
                    "value": get("value"),
                    "value_type": get("value_type"),
                    "instruction": get("instruction"),
                    "priority": get("priority"),
                    "enabled": True,
                }
            )

    return {
        "source": str(path),
        "database": database,
        "rules": rules,
        "database_count": len(database),
        "rules_count": len(rules),
    }
