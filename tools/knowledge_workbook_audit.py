from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

from config import KNOWLEDGE_WORKBOOK_GCS_URI, KNOWLEDGE_WORKBOOK_PATH

try:
    from google.cloud import storage
except Exception:  # pragma: no cover - optional import failure path
    storage = None

from openpyxl import load_workbook


DEFAULT_LOCAL_WORKBOOK = APP_ROOT / "data/knowledge/FLAIRLAB_EventPost_Master_Knowledge.xlsm"
DEFAULT_OUTPUT_JSON = APP_ROOT / "data/audits/knowledge_workbook_audit.json"
DEFAULT_OUTPUT_MD = APP_ROOT / "data/audits/knowledge_workbook_audit.md"

SEO_TAB = "seo_rules"
SHARED_TAB = "shared_field_schema"


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def parse_gcs_uri(uri: str) -> tuple[str, str]:
    raw = str(uri or "").strip()
    if not raw.startswith("gs://"):
        raise ValueError("GCS URI must use gs://bucket/path format")
    bucket_name, _, blob_name = raw[5:].partition("/")
    if not bucket_name or not blob_name:
        raise ValueError("GCS URI must include both bucket and object path")
    return bucket_name, blob_name


def resolve_local_workbook_path() -> Path:
    if KNOWLEDGE_WORKBOOK_PATH:
        configured = Path(KNOWLEDGE_WORKBOOK_PATH)
        if not configured.is_absolute():
            configured = APP_ROOT / configured
        return configured
    return DEFAULT_LOCAL_WORKBOOK


def load_workbook_bytes_from_gcs(gcs_uri: str) -> tuple[bytes, dict[str, Any]]:
    if storage is None:
        raise RuntimeError("google-cloud-storage is required to parse workbook directly from GCS")
    bucket_name, blob_name = parse_gcs_uri(gcs_uri)
    client = storage.Client()
    blob = client.bucket(bucket_name).blob(blob_name)
    if not blob.exists():
        raise FileNotFoundError(f"Workbook object not found in GCS: {gcs_uri}")
    blob.reload()
    content = blob.download_as_bytes()
    metadata = {
        "storage_mode": "gcs",
        "gcs_uri": gcs_uri,
        "gcs_generation": str(blob.generation or ""),
        "gcs_updated": blob.updated.isoformat() if getattr(blob, "updated", None) else None,
        "size_bytes": len(content),
        "sha256": hashlib.sha256(content).hexdigest(),
    }
    return content, metadata


def load_workbook_bytes_from_local(path: Path) -> tuple[bytes, dict[str, Any]]:
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Local workbook file not found: {path}")
    content = path.read_bytes()
    stat = path.stat()
    metadata = {
        "storage_mode": "local_file",
        "local_path": str(path),
        "local_modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        "size_bytes": len(content),
        "sha256": hashlib.sha256(content).hexdigest(),
    }
    return content, metadata


def merged_cell_value(sheet: Any, row: int, column: int) -> Any:
    cell = sheet.cell(row=row, column=column)
    if cell.value not in (None, ""):
        return cell.value
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


GENERIC_FIELD_HEADER_ALIASES = {
    "parameter",
    "seo parameter",
    "seo_parameter",
    "field",
    "field name",
    "field_name",
    "key",
    "meta key",
    "meta_key",
    "source field name",
    "source_field_name",
    "source field key",
    "source_field_key",
    "user field name",
    "user_field_name",
    "shared field name",
    "acf field name",
    "acf_field_name",
    "column",
    "column name",
    "column_name",
}

TAB_FIELD_HEADER_ALIASES: dict[str, list[str]] = {
    SEO_TAB: ["applies_to", "field_key", "field name", "field_key"],
    SHARED_TAB: ["field_key", "field name", "shared field name"],
}


def header_positions(sheet: Any, max_scan_rows: int = 80) -> dict[str, tuple[int, int]]:
    positions: dict[str, tuple[int, int]] = {}
    for row_index in range(1, min(sheet.max_row, max_scan_rows) + 1):
        for col_index in range(1, sheet.max_column + 1):
            value = merged_cell_value(sheet, row_index, col_index)
            normalized = normalize_key(str(value or ""))
            if normalized:
                positions.setdefault(normalized, (row_index, col_index))
    return positions


def find_field_column(sheet: Any, tab_name: str) -> tuple[int, int] | None:
    positions = header_positions(sheet)
    for alias in TAB_FIELD_HEADER_ALIASES.get(tab_name, []):
        if alias in positions:
            return positions[alias]
    for alias in GENERIC_FIELD_HEADER_ALIASES:
        if alias in positions:
            return positions[alias]
    return None


def best_effort_field_column(sheet: Any) -> tuple[int, int] | None:
    best_col = None
    best_count = 0
    for col_index in range(1, sheet.max_column + 1):
        non_empty = 0
        for row_index in range(1, min(sheet.max_row, 200) + 1):
            value = str(merged_cell_value(sheet, row_index, col_index) or "").strip()
            if value:
                non_empty += 1
        if non_empty > best_count:
            best_col = col_index
            best_count = non_empty
    if best_col is None or best_count < 2:
        return None
    return 1, best_col


def extract_field_inventory(sheet: Any, tab_name: str) -> dict[str, Any]:
    header_info = find_field_column(sheet, tab_name)
    header_mode = "alias_match"
    if header_info is None:
        header_info = best_effort_field_column(sheet)
        header_mode = "best_effort"
    if header_info is None:
        return {
            "header_row": None,
            "field_column": None,
            "header_mode": "not_found",
            "rows_total": 0,
            "fields": [],
        }

    header_row, field_col = header_info
    positions = header_positions(sheet)
    enabled_pos = positions.get("enabled") if tab_name == SEO_TAB else None
    ordered_fields: list[str] = []
    seen: set[str] = set()
    rows: list[dict[str, Any]] = []

    for row_index in range(header_row + 1, sheet.max_row + 1):
        if enabled_pos is not None:
            enabled_value = str(merged_cell_value(sheet, row_index, enabled_pos[1]) or "").strip().lower()
            if enabled_value not in {"true", "1", "yes", "enabled", "active"}:
                continue

        raw_value = str(merged_cell_value(sheet, row_index, field_col) or "").strip()
        if not raw_value:
            continue
        normalized = normalize_key(raw_value)
        if normalized in {"-", "n/a", "na", "none", "x"}:
            continue
        rows.append({"row": row_index, "value": raw_value})
        if normalized not in seen:
            seen.add(normalized)
            ordered_fields.append(raw_value)

    return {
        "header_row": header_row,
        "field_column": field_col,
        "header_mode": header_mode,
        "rows_total": len(rows),
        "rows": rows,
        "fields": ordered_fields,
        "unique_fields_total": len(ordered_fields),
    }


def make_markdown_report(payload: dict[str, Any]) -> str:
    version = payload["workbook_version"]
    sheets = payload.get("sheet_names", [])
    seo = payload["tabs"].get(SEO_TAB, {})
    shared = payload["tabs"].get(SHARED_TAB, {})

    def lines_for_fields(values: list[str]) -> str:
        if not values:
            return "- (none)"
        return "\n".join(f"- {value}" for value in values)

    lines = [
        "# Knowledge Workbook Audit",
        "",
        f"- Generated at (UTC): {payload.get('generated_at')}",
        f"- Storage mode: {version.get('storage_mode')}",
        f"- Workbook size (bytes): {version.get('size_bytes')}",
        f"- Workbook sha256: {version.get('sha256')}",
        f"- GCS URI: {version.get('gcs_uri')}",
        f"- GCS generation: {version.get('gcs_generation')}",
        f"- GCS updated: {version.get('gcs_updated')}",
        f"- Local path: {version.get('local_path')}",
        f"- Local modified at: {version.get('local_modified_at')}",
        "",
        "## Sheet Names",
    ]
    if sheets:
        lines.extend([f"- {name}" for name in sheets])
    else:
        lines.append("- (none)")
    lines.extend(
        [
            "",
            f"## {SEO_TAB}",
            f"- Header mode: {seo.get('header_mode')}",
            f"- Header row: {seo.get('header_row')}",
            f"- Field column: {seo.get('field_column')}",
            f"- Data rows: {seo.get('rows_total', 0)}",
            f"- Unique fields: {seo.get('unique_fields_total', 0)}",
            lines_for_fields(seo.get("fields", [])),
            "",
            f"## {SHARED_TAB}",
            f"- Header mode: {shared.get('header_mode')}",
            f"- Header row: {shared.get('header_row')}",
            f"- Field column: {shared.get('field_column')}",
            f"- Data rows: {shared.get('rows_total', 0)}",
            f"- Unique fields: {shared.get('unique_fields_total', 0)}",
            lines_for_fields(shared.get("fields", [])),
            "",
        ]
    )
    return "\n".join(lines)


def run_audit(prefer_gcs: bool, strict_gcs: bool, output_json: Path, output_md: Path) -> dict[str, Any]:
    gcs_uri = str(KNOWLEDGE_WORKBOOK_GCS_URI or "").strip()
    local_path = resolve_local_workbook_path()

    if prefer_gcs and gcs_uri:
        try:
            workbook_bytes, workbook_version = load_workbook_bytes_from_gcs(gcs_uri)
        except Exception:
            if strict_gcs:
                raise
            workbook_bytes, workbook_version = load_workbook_bytes_from_local(local_path)
            workbook_version["gcs_uri"] = gcs_uri
            workbook_version["gcs_fallback_reason"] = "Direct GCS parse failed in current environment; used local workbook copy"
    elif local_path.exists():
        workbook_bytes, workbook_version = load_workbook_bytes_from_local(local_path)
        if gcs_uri:
            workbook_version["gcs_uri"] = gcs_uri
    elif gcs_uri:
        workbook_bytes, workbook_version = load_workbook_bytes_from_gcs(gcs_uri)
    else:
        raise FileNotFoundError(
            "No workbook source available. Configure KNOWLEDGE_WORKBOOK_GCS_URI or local knowledge workbook path."
        )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True, keep_vba=True, read_only=False)
    sheet_names = list(workbook.sheetnames)

    tabs: dict[str, Any] = {}
    for tab_name in (SEO_TAB, SHARED_TAB):
        if tab_name in workbook.sheetnames:
            tabs[tab_name] = extract_field_inventory(workbook[tab_name], tab_name)
        else:
            tabs[tab_name] = {
                "missing": True,
                "rows_total": 0,
                "fields": [],
                "unique_fields_total": 0,
            }

    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workbook_version": workbook_version,
        "sheet_names": sheet_names,
        "tabs": tabs,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    output_md.parent.mkdir(parents=True, exist_ok=True)
    output_md.write_text(make_markdown_report(payload), encoding="utf-8")
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit seo_rules/shared_field_schema field inventories from knowledge workbook")
    parser.add_argument(
        "--source",
        choices=["auto", "gcs", "local"],
        default="gcs",
        help="Workbook source preference",
    )
    parser.add_argument("--out-json", default=str(DEFAULT_OUTPUT_JSON), help="Output JSON path")
    parser.add_argument("--out-md", default=str(DEFAULT_OUTPUT_MD), help="Output Markdown path")
    args = parser.parse_args()

    prefer_gcs = args.source in {"auto", "gcs"}
    if args.source == "local":
        prefer_gcs = False

    payload = run_audit(
        prefer_gcs=prefer_gcs,
        strict_gcs=args.source == "gcs",
        output_json=Path(args.out_json),
        output_md=Path(args.out_md),
    )
    seo_count = payload["tabs"][SEO_TAB].get("unique_fields_total", 0)
    shared_count = payload["tabs"][SHARED_TAB].get("unique_fields_total", 0)
    print(f"Audit complete. seo_rules={seo_count} unique fields, shared_field_schema={shared_count} unique fields")
    print(f"JSON: {args.out_json}")
    print(f"MD: {args.out_md}")


if __name__ == "__main__":
    main()