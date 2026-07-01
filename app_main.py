import hashlib
import base64
import json
import os
import re
import shutil
import tempfile
import uuid
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import unquote
from urllib.request import urlopen

try:
  from google.cloud import storage as gcs_storage
except Exception:
  gcs_storage = None

from fastapi import BackgroundTasks, Depends, FastAPI, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import FileResponse, HTMLResponse, Response
from pydantic import BaseModel, Field as PydanticField

try:
  from PIL import Image, ImageEnhance, ImageFilter, ImageOps
except Exception:
  Image = None
  ImageEnhance = None
  ImageFilter = None
  ImageOps = None

try:
  from openai import OpenAI
except Exception:
  OpenAI = None

from action_api_event_import import (
    EventPostActionRequest,
    app as action_app,
    build_import_args,
    import_event_post_from_zip,
    post_response_from_output,
)
from app_draft_generator import (
  apply_refinement_answers,
  create_session_draft,
  default_faq_guidance,
  is_faq_field_name,
  merge_guidance_text,
  rebuild_session_package,
  revise_session_draft,
  suggest_image_metadata_for_state,
)
from app_knowledge_base import (
  field_matches,
  guidance_for_field,
  guidance_items_for_field,
  load_internal_links_context,
  load_workbook_guidance,
  normalize_lookup,
)
from app_transcription import DEFAULT_TRANSCRIPTION_MODEL, sanitize_transcript_text, transcribe_audio_file
from config import (
  OPENAI_API_KEY,
  KNOWLEDGE_SOURCE_POLICY,
  KNOWLEDGE_WORKBOOK_GCS_URI,
  KNOWLEDGE_WORKBOOK_PATH,
  KNOWLEDGE_WORKBOOK_SHEET,
  SESSION_STATE_GCS_PREFIX,
  CONTENT_PIPELINE_VERSION,
  get_client_config,
  set_active_client,
)
from run_event_import import run_import
from step_10_event_payload import ColumnSpec
from step_40_wordpress_api import preflight_wordpress_permissions, set_post_featured_media, update_media_metadata, upload_media
from app.v2.api.step_02_routes import create_router as create_v2_router, v2_error_handler
from app.v2.api.step_03_container import get_v2_service, v2_readiness
from app.v2.errors import V2Error
from app.v2.knowledge_base.step_02_loader import WorkbookLoader
from app.v2.knowledge_base.step_03_validator import WorkbookValidator
from app.v2.observability import V2RequestLoggingMiddleware


APP_ROOT = Path(__file__).resolve().parent
APP_SESSION_ROOT = APP_ROOT / "data/app_sessions"
DEFAULT_KNOWLEDGE_DIR = APP_ROOT / "data/knowledge"
DEFAULT_KNOWLEDGE_WORKBOOK_PATH = DEFAULT_KNOWLEDGE_DIR / "FLAIRLAB_EventPost_Master_Knowledge.xlsm"
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif"}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".m4v", ".webm"}
AUDIO_EXTENSIONS = {".mp3", ".m4a", ".ogg", ".opus", ".wav", ".webm", ".mp4"}
WORKBOOK_EXTENSIONS = {".xlsm", ".xlsx"}
WORKBOOK_METADATA_SUFFIX = ".gcs-meta.json"

_gcs_client: Any = None
_knowledge_gcs_synced = False


def knowledge_source_policy() -> str:
  configured = str(KNOWLEDGE_SOURCE_POLICY or "").strip().lower()
  if configured in {"gcs_required", "gcs_preferred", "local_only", "auto"}:
    return configured
  if configured:
    return "auto"
  if configured_knowledge_workbook_gcs_uri():
    return "gcs_required" if os.getenv("K_SERVICE") else "gcs_preferred"
  return "auto"

app = FastAPI(
    title="FLAIRLAB Post Generator",
    version="0.1.0",
    description="Mobile workflow for voice, media, AI draft review, and WordPress event post creation.",
)


def verify_api_key(x_api_key: str | None = Header(default=None)) -> None:
    from action_api_event_import import IMPORT_API_KEY as configured_key

    if configured_key and x_api_key != configured_key:
        raise HTTPException(status_code=401, detail="Invalid or missing API key.")


def verify_download_api_key(
  x_api_key: str | None = Header(default=None),
  api_key: str | None = Query(default=None),
) -> None:
  from action_api_event_import import IMPORT_API_KEY as configured_key

  if configured_key and x_api_key != configured_key and api_key != configured_key:
    raise HTTPException(status_code=401, detail="Invalid or missing API key.")


app.add_exception_handler(V2Error, v2_error_handler)
app.add_middleware(V2RequestLoggingMiddleware)
app.include_router(create_v2_router(get_v2_service, verify_api_key, v2_readiness))


@app.on_event("startup")
def validate_selected_pipeline_startup() -> None:
    if CONTENT_PIPELINE_VERSION == "v2":
        get_v2_service().knowledge.current()


class SessionCreateRequest(BaseModel):
    client_id: str = "flairlab"
    post_type: str = "Event"


class SessionCreateResponse(BaseModel):
    session_id: str
    client_id: str
    post_type: str
    status: str
    next_step: str


class SessionStateResponse(BaseModel):
    session_id: str
    client_id: str
    post_type: str
    status: str
    steps: dict[str, str]
    files: dict[str, Any] = PydanticField(default_factory=dict)
    transcript: dict[str, Any] = PydanticField(default_factory=dict)
    draft: dict[str, Any] = PydanticField(default_factory=dict)
    wordpress_post: dict[str, Any] = PydanticField(default_factory=dict)
    wordpress_media_library: dict[str, Any] = PydanticField(default_factory=dict)
    image_processing: dict[str, Any] = PydanticField(default_factory=dict)
    ui_cache: dict[str, Any] = PydanticField(default_factory=dict)
    vision: dict[str, Any] = PydanticField(default_factory=dict)
    media_recovery: dict[str, Any] = PydanticField(default_factory=dict)
    ai_usage: dict[str, Any] = PydanticField(default_factory=dict)


class TranscriptUpdateRequest(BaseModel):
    text: str


class DraftGenerateRequest(BaseModel):
    category: str = "auto event post"
    force_regenerate: bool = True


class DraftUpdateRequest(BaseModel):
    csv_text: str


class DraftChatRequest(BaseModel):
    message: str


class DraftChatTranscriptionResponse(BaseModel):
  text: str
  raw_text: str
  source: str
  model: str


class DraftRefinementRequest(BaseModel):
    answers: dict[str, str]  # {"field_name": "user provided text", ...}


class UiCacheUpdateRequest(BaseModel):
    cache: dict[str, Any] = PydanticField(default_factory=dict)


class CreateWordPressPostRequest(BaseModel):
	status: str = "draft"
	existing_post_mode: str = "update"
	existing_post_id: int | None = None
	update_existing_generated_post: bool = False


class FeaturedImageRequest(BaseModel):
    filename: str


class ImageMetadataUpdateRequest(BaseModel):
  alt_text: str = ""
  title: str = ""
  caption: str = ""
  description: str = ""
  use_suggestions: bool = False


class ImageOptimizationRequest(BaseModel):
  prompt: str = ""


class VisionSelectionRequest(BaseModel):
  filenames: list[str] = PydanticField(default_factory=list)


class SessionsDeleteRequest(BaseModel):
  session_ids: list[str] = PydanticField(default_factory=list)


def write_session_state(session_id: str, state: dict[str, Any]) -> None:
  session_dir = APP_SESSION_ROOT / session_id
  session_dir.mkdir(parents=True, exist_ok=True)
  state_path = session_dir / "state.json"
  state_path.write_text(
    json.dumps(state, ensure_ascii=False, indent=2),
    encoding="utf-8",
  )
  cache_dir = session_dir / "cache"
  cache_dir.mkdir(parents=True, exist_ok=True)
  (cache_dir / "session_state.json").write_text(
    json.dumps(state, ensure_ascii=False, indent=2),
    encoding="utf-8",
  )
  transcript_text = state.get("transcript", {}).get("text")
  if transcript_text is not None:
    (cache_dir / "transcript.txt").write_text(transcript_text, encoding="utf-8")
  draft_csv = state.get("draft", {}).get("csv_text")
  if draft_csv is not None:
    (cache_dir / "draft.csv").write_text(draft_csv, encoding="utf-8")
  wordpress_post = state.get("wordpress_post")
  if wordpress_post:
    (cache_dir / "wordpress_post.json").write_text(
      json.dumps(wordpress_post, ensure_ascii=False, indent=2),
      encoding="utf-8",
    )
  upload_session_state_to_gcs(session_id, state_path)


def prune_missing_session_media(state: dict[str, Any]) -> bool:
    files = state.get("files")
    if not isinstance(files, dict):
      return False

    changed = False

    def existing_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
      kept: list[dict[str, Any]] = []
      for item in items:
        path_value = str(item.get("path") or "").strip()
        if not path_value:
          continue
        path = Path(path_value)
        if path.exists() and path.is_file():
          kept.append(item)
      return kept

    images = existing_items(list(files.get("images") or []))
    if len(images) != len(list(files.get("images") or [])):
      files["images"] = images
      changed = True

    videos = existing_items(list(files.get("videos") or []))
    if len(videos) != len(list(files.get("videos") or [])):
      files["videos"] = videos
      changed = True

    voices = list(files.get("voices") or [])
    if not voices and isinstance(files.get("voice"), dict):
      voices = [files["voice"]]
    voices_kept = existing_items(voices)
    if len(voices_kept) != len(voices):
      files["voices"] = voices_kept
      changed = True

    normalized_voice = voices_kept[0] if voices_kept else None
    if files.get("voice") != normalized_voice:
      files["voice"] = normalized_voice
      changed = True

    featured = str(files.get("featured_image_filename") or "").strip()
    image_names = {str(item.get("filename") or "").strip() for item in images}
    if featured and featured not in image_names:
      files["featured_image_filename"] = images[0].get("filename") if images else None
      changed = True

    selected_video = str(files.get("selected_video_filename") or "").strip()
    video_names = {str(item.get("filename") or "").strip() for item in videos}
    if selected_video and selected_video not in video_names:
      files["selected_video_filename"] = videos[0].get("filename") if videos else None
      changed = True

    return changed


def iter_session_media_items(state: dict[str, Any]) -> list[tuple[str, dict[str, Any]]]:
    files = state.get("files") if isinstance(state, dict) else None
    if not isinstance(files, dict):
      return []

    items: list[tuple[str, dict[str, Any]]] = []
    for key in ("images", "videos", "voices"):
      for item in files.get(key) or []:
        if isinstance(item, dict):
          items.append((key, item))
    single_voice = files.get("voice")
    if isinstance(single_voice, dict):
      items.append(("voices", single_voice))
    return items


def session_media_counts(state: dict[str, Any]) -> dict[str, int]:
    files = state.get("files") if isinstance(state, dict) else None
    if not isinstance(files, dict):
      return {"images": 0, "videos": 0, "voices": 0}

    voice_items = list(files.get("voices") or [])
    if not voice_items and isinstance(files.get("voice"), dict):
      voice_items = [files.get("voice")]
    return {
      "images": len(list(files.get("images") or [])),
      "videos": len(list(files.get("videos") or [])),
      "voices": len([item for item in voice_items if isinstance(item, dict)]),
    }


def read_session_state(session_id: str) -> dict[str, Any]:
    path = APP_SESSION_ROOT / session_id / "state.json"
    if not path.exists():
        synced = sync_session_state_from_gcs(session_id)
        if not synced or not path.exists():
            raise HTTPException(status_code=404, detail="Session not found.")
    state = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(state, dict):
      counts_before = session_media_counts(state)
      state_changed = sync_session_media_from_gcs(session_id, state)
      if prune_missing_session_media(state):
        state_changed = True
      counts_after = session_media_counts(state)

      missing_counts = {
        kind: max(counts_before.get(kind, 0) - counts_after.get(kind, 0), 0)
        for kind in ("images", "videos", "voices")
      }
      missing_total = sum(missing_counts.values())

      current_notice = state.get("media_recovery") if isinstance(state.get("media_recovery"), dict) else {}
      if missing_total > 0:
        next_notice = {
          "missing_total": missing_total,
          "missing_images": missing_counts["images"],
          "missing_videos": missing_counts["videos"],
          "missing_voices": missing_counts["voices"],
          "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        if current_notice != next_notice:
          state["media_recovery"] = next_notice
          state_changed = True
      elif current_notice:
        state["media_recovery"] = {}
        state_changed = True

      if state_changed:
        write_session_state(session_id, state)
    return state


def safe_upload_name(filename: str | None, fallback: str) -> str:
    raw_name = Path(filename or fallback).name
    cleaned = "".join(char if char.isalnum() or char in "._- " else "_" for char in raw_name).strip()
    return cleaned or fallback


def filename_match_candidates(value: str | None) -> list[str]:
    text = str(value or "").strip()
    if not text:
      return []
    candidates: list[str] = [text]

    decoded = unquote(text)
    if decoded and decoded not in candidates:
      candidates.append(decoded)

    decoded_twice = unquote(decoded)
    if decoded_twice and decoded_twice not in candidates:
      candidates.append(decoded_twice)

    # Keep compatibility with older sessions that persisted encoded names.
    for candidate in list(candidates):
      if "%20" in candidate:
        plain = candidate.replace("%20", " ")
        if plain and plain not in candidates:
          candidates.append(plain)
      if " " in candidate:
        encoded = candidate.replace(" ", "%20")
        if encoded and encoded not in candidates:
          candidates.append(encoded)
    return candidates


def find_media_item_by_filename(items: list[dict[str, Any]], filename: str) -> dict[str, Any] | None:
    requested = {value.lower() for value in filename_match_candidates(filename) if value}
    if not requested:
      return None

    for item in items:
      if not isinstance(item, dict):
        continue
      names = [item.get("filename"), item.get("original_filename")]
      item_candidates: set[str] = set()
      for name in names:
        for candidate in filename_match_candidates(str(name or "")):
          item_candidates.add(candidate.lower())
      if requested.intersection(item_candidates):
        return item
    return None


def save_upload_file(upload: UploadFile, target_dir: Path, fallback: str) -> dict[str, Any]:
    target_dir.mkdir(parents=True, exist_ok=True)
    original_filename = upload.filename or fallback
    filename = safe_upload_name(upload.filename, fallback)
    destination = target_dir / filename
    counter = 2
    while destination.exists():
        destination = target_dir / f"{destination.stem}-{counter}{destination.suffix}"
        counter += 1

    with destination.open("wb") as output:
        while chunk := upload.file.read(1024 * 1024):
            output.write(chunk)

    return {
        "filename": destination.name,
        "original_filename": original_filename,
        "path": str(destination),
        "size": destination.stat().st_size,
        "content_type": upload.content_type,
    }


def validate_extension(file_info: dict[str, Any], allowed: set[str], label: str) -> None:
    suffix = Path(file_info["filename"]).suffix.lower()
    if suffix not in allowed:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported {label} file extension: {file_info['filename']}",
        )


def parse_first_int(value: str | None) -> int | None:
    text = str(value or "")
    match = re.search(r"(\d+)", text)
    if not match:
      return None
    try:
      return int(match.group(1))
    except ValueError:
      return None


def parse_aspect_ratio(value: str | None) -> float | None:
    text = str(value or "").strip().lower()
    if not text:
      return None

    # Supports "4:5", "16/9", "1200x630", "1.91:1".
    match = re.search(r"(\d+(?:\.\d+)?)\s*[:/x×]\s*(\d+(?:\.\d+)?)", text)
    if not match:
      return None
    try:
      left = float(match.group(1))
      right = float(match.group(2))
    except ValueError:
      return None
    if left <= 0 or right <= 0:
      return None
    ratio = left / right
    if ratio < 0.2 or ratio > 5.0:
      return None
    return ratio


def parse_boolish(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "on", "y"}


def load_image_pillow_rules() -> dict[str, Any]:
    rules: dict[str, Any] = {
      "target_kb": 50,
      "max_width": 1800,
      "max_height": 1800,
      "aspect_ratio": None,
      "crop_mode": "contain",
      "focal_x": 0.5,
      "focal_y": 0.5,
      "auto_orient": True,
      "strip_metadata": True,
      "noise_reduction": 0,
      "shadow_lift": 1.0,
      "quality_start": 90,
      "quality_min": 30,
      "quality_step": 8,
      "brightness": 1.06,
      "contrast": 1.08,
      "saturation": 1.08,
      "sharpness": 1.05,
      "autocontrast_cutoff": 1,
    }
    workbook_path = active_knowledge_workbook_path()
    try:
      from openpyxl import load_workbook
      workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True)
    except Exception:
      return rules

    sheet_name = ""
    if "image_rules_pillow" in workbook.sheetnames:
      sheet_name = "image_rules_pillow"
    elif "image_rules" in workbook.sheetnames:
      sheet_name = "image_rules"
    if not sheet_name:
      return rules

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
      return rules
    normalized_headers = {str(name or "").strip().lower(): idx for idx, name in enumerate(header)}
    field_idx = normalized_headers.get("field")
    rule_idx = normalized_headers.get("rule")
    example_idx = normalized_headers.get("example")
    rule_key_idx = normalized_headers.get("rule_key")
    enabled_idx = normalized_headers.get("enabled")
    value_idx = normalized_headers.get("value")

    # New workbook schema: rule_key, enabled, value
    if rule_key_idx is not None and value_idx is not None:
      for row in rows:
        if rule_key_idx >= len(row):
          continue
        key = str(row[rule_key_idx] or "").strip().lower()
        if not key:
          continue
        enabled = True
        if enabled_idx is not None and enabled_idx < len(row):
          enabled = parse_boolish(row[enabled_idx])
        if not enabled:
          continue

        value = row[value_idx] if value_idx < len(row) else None
        value_text = str(value or "").strip()
        if not value_text:
          continue

        if key in {"max_filesize", "target_filesize", "target_kb", "target_size_kb"}:
          parsed = parse_first_int(value_text)
          if parsed and parsed > 0:
            rules["target_kb"] = parsed
        elif key in {"max_width", "target_width"}:
          parsed = parse_first_int(value_text)
          if parsed and parsed >= 600:
            rules["max_width"] = parsed
        elif key in {"max_height", "target_height"}:
          parsed = parse_first_int(value_text)
          if parsed and parsed >= 600:
            rules["max_height"] = parsed
        elif key == "brightness":
          try:
            rules["brightness"] = float(value)
          except (TypeError, ValueError):
            pass
        elif key == "contrast":
          try:
            rules["contrast"] = float(value)
          except (TypeError, ValueError):
            pass
        elif key == "saturation":
          try:
            rules["saturation"] = float(value)
          except (TypeError, ValueError):
            pass
        elif key == "sharpness":
          try:
            rules["sharpness"] = float(value)
          except (TypeError, ValueError):
            pass
        elif key == "webp_quality_start":
          parsed = parse_first_int(value_text)
          if parsed:
            rules["quality_start"] = max(30, min(100, parsed))
        elif key == "webp_quality_min":
          parsed = parse_first_int(value_text)
          if parsed:
            rules["quality_min"] = max(1, min(95, parsed))
        elif key == "webp_quality_step":
          parsed = parse_first_int(value_text)
          if parsed:
            rules["quality_step"] = max(1, min(20, parsed))
        elif key == "crop_mode":
          rules["crop_mode"] = value_text.lower()
        elif key == "focal_x":
          try:
            rules["focal_x"] = max(0.0, min(1.0, float(value)))
          except (TypeError, ValueError):
            pass
        elif key == "focal_y":
          try:
            rules["focal_y"] = max(0.0, min(1.0, float(value)))
          except (TypeError, ValueError):
            pass
        elif key == "auto_orient":
          rules["auto_orient"] = parse_boolish(value)
        elif key == "strip_metadata":
          rules["strip_metadata"] = parse_boolish(value)
        elif key in {"noise_reduction", "median_size"}:
          parsed = parse_first_int(value_text)
          if parsed and parsed >= 3:
            # Median filter requires odd size >= 3.
            rules["noise_reduction"] = parsed if parsed % 2 == 1 else parsed + 1
        elif key in {"shadow_lift", "gamma"}:
          try:
            gamma = float(value)
            if 0.2 <= gamma <= 2.5:
              rules["shadow_lift"] = gamma
          except (TypeError, ValueError):
            pass
        elif key in {"target_ratio", "aspect_ratio", "output_ratio"}:
          parsed_ratio = parse_aspect_ratio(value_text)
          if parsed_ratio:
            rules["aspect_ratio"] = parsed_ratio

      if rules["quality_min"] > rules["quality_start"]:
        rules["quality_min"] = rules["quality_start"]
      return rules

    if field_idx is None:
      return rules

    for row in rows:
      if field_idx >= len(row):
        continue
      field_name = str(row[field_idx] or "").strip().lower()
      if not field_name:
        continue
      rule_text = str(row[rule_idx] if rule_idx is not None and rule_idx < len(row) else "")
      example_text = str(row[example_idx] if example_idx is not None and example_idx < len(row) else "")
      payload = f"{rule_text} {example_text}".strip()

      if field_name in {"target_filesize", "target_kb", "target_size_kb"}:
        parsed = parse_first_int(payload)
        if parsed and parsed > 0:
          rules["target_kb"] = parsed
      elif field_name in {"resolution_rule", "max_width", "target_width"}:
        parsed = parse_first_int(payload)
        if parsed and parsed >= 600:
          rules["max_width"] = parsed
      if field_name in {"crop_ratio", "aspect_ratio", "output_ratio", "resolution_rule"}:
        parsed_ratio = parse_aspect_ratio(payload)
        if parsed_ratio:
          rules["aspect_ratio"] = parsed_ratio
    return rules


def _best_window_start(energies: list[int], window: int) -> int:
    if window <= 0 or not energies:
      return 0
    if window >= len(energies):
      return 0
    current = sum(energies[:window])
    best = current
    best_start = 0
    for idx in range(window, len(energies)):
      current += energies[idx] - energies[idx - window]
      start = idx - window + 1
      if current > best:
        best = current
        best_start = start
    return best_start


def smart_crop_to_aspect_ratio(
  image: Any,
  aspect_ratio: float | None,
  focal_x: float | None = None,
  focal_y: float | None = None,
) -> Any:
    if Image is None or ImageFilter is None:
      return image
    if not aspect_ratio or image.width <= 1 or image.height <= 1:
      return image

    current_ratio = image.width / float(image.height)
    if abs(current_ratio - aspect_ratio) < 0.01:
      return image

    if current_ratio > aspect_ratio:
      crop_width = max(1, int(image.height * aspect_ratio))
      crop_height = image.height
      axis = "x"
    else:
      crop_width = image.width
      crop_height = max(1, int(image.width / aspect_ratio))
      axis = "y"

    # Use edge density as a lightweight saliency proxy for better composition than center-crop.
    sample_w = max(40, min(320, image.width))
    sample_h = max(40, min(320, image.height))
    sample = image.convert("L").resize((sample_w, sample_h), Image.BILINEAR).filter(ImageFilter.FIND_EDGES)
    pixels = list(sample.getdata())

    if axis == "x":
      if focal_x is not None:
        center_x = int(max(0.0, min(1.0, focal_x)) * image.width)
        left = max(0, min(center_x - (crop_width // 2), image.width - crop_width))
        top = 0
        return image.crop((left, top, left + crop_width, top + crop_height))
      col_energy = [0] * sample_w
      for y in range(sample_h):
        base = y * sample_w
        for x in range(sample_w):
          col_energy[x] += pixels[base + x]
      window = max(1, int(crop_width * sample_w / float(image.width)))
      start = _best_window_start(col_energy, window)
      left = int(start * image.width / float(sample_w))
      left = max(0, min(left, image.width - crop_width))
      top = 0
    else:
      if focal_y is not None:
        center_y = int(max(0.0, min(1.0, focal_y)) * image.height)
        top = max(0, min(center_y - (crop_height // 2), image.height - crop_height))
        left = 0
        return image.crop((left, top, left + crop_width, top + crop_height))
      row_energy = [0] * sample_h
      for y in range(sample_h):
        base = y * sample_w
        row_energy[y] = sum(pixels[base:base + sample_w])
      window = max(1, int(crop_height * sample_h / float(image.height)))
      start = _best_window_start(row_energy, window)
      top = int(start * image.height / float(sample_h))
      top = max(0, min(top, image.height - crop_height))
      left = 0

    return image.crop((left, top, left + crop_width, top + crop_height))


def save_webp_under_target(input_path: Path, output_path: Path, rules: dict[str, Any]) -> dict[str, Any]:
    """Process image with Pillow rules and return applied operations metadata."""
    if Image is None or ImageEnhance is None or ImageOps is None:
        raise RuntimeError("Pillow is not available.")

    applied: list[str] = []
    image = Image.open(input_path)
    orig_size = image.size

    if bool(rules.get("auto_orient", True)):
        image = ImageOps.exif_transpose(image)
        applied.append("auto_orient")

    if image.mode not in ("RGB", "L"):
        image = image.convert("RGB")
        applied.append(f"color_convert_{image.mode}")
    elif image.mode == "L":
        image = image.convert("RGB")
        applied.append("grayscale_to_rgb")

    aspect_ratio = rules.get("aspect_ratio")
    crop_mode = str(rules.get("crop_mode") or "contain").strip().lower()
    # If a target ratio is configured, apply ratio crop unless mode explicitly disables cropping.
    # This prevents silent no-op when users set target_ratio but leave crop_mode on "contain".
    if isinstance(aspect_ratio, (int, float)) and crop_mode not in {"none", "off", "disabled"}:
        image = smart_crop_to_aspect_ratio(
            image,
            float(aspect_ratio),
            focal_x=rules.get("focal_x"),
            focal_y=rules.get("focal_y"),
        )
        applied.append(f"crop_to_{aspect_ratio:.2f}")

    max_width = int(rules.get("max_width") or 1800)
    max_height = int(rules.get("max_height") or 1800)
    scale = min(
        1.0,
        max_width / float(image.width) if image.width > 0 else 1.0,
        max_height / float(image.height) if image.height > 0 else 1.0,
    )
    if scale < 1.0:
        image = image.resize(
            (max(1, int(image.width * scale)), max(1, int(image.height * scale))),
            Image.LANCZOS,
        )
        applied.append(f"resize_to_{image.width}x{image.height}")

    noise_size = int(rules.get("noise_reduction") or 0)
    if noise_size >= 3 and ImageFilter is not None:
        image = image.filter(ImageFilter.MedianFilter(size=noise_size))
        applied.append(f"denoise_median{noise_size}")

    image = ImageOps.autocontrast(image, cutoff=int(rules.get("autocontrast_cutoff") or 1))
    applied.append("autocontrast")

    brightness = float(rules.get("brightness") or 1.06)
    if abs(brightness - 1.0) > 0.01:
        image = ImageEnhance.Brightness(image).enhance(brightness)
        applied.append(f"brightness_{brightness:.2f}x")

    contrast = float(rules.get("contrast") or 1.08)
    if abs(contrast - 1.0) > 0.01:
        image = ImageEnhance.Contrast(image).enhance(contrast)
        applied.append(f"contrast_{contrast:.2f}x")

    saturation = float(rules.get("saturation") or 1.08)
    if abs(saturation - 1.0) > 0.01:
        image = ImageEnhance.Color(image).enhance(saturation)
        applied.append(f"saturation_{saturation:.2f}x")

    sharpness = float(rules.get("sharpness") or 1.05)
    if abs(sharpness - 1.0) > 0.01:
        image = ImageEnhance.Sharpness(image).enhance(sharpness)
        applied.append(f"sharpen_{sharpness:.2f}x")

    gamma = float(rules.get("shadow_lift") or 1.0)
    gamma = float(rules.get("shadow_lift") or 1.0)
    if abs(gamma - 1.0) > 1e-3:
        # Gamma < 1.0 brightens shadows; >1.0 darkens.
        lut = [max(0, min(255, int((i / 255.0) ** gamma * 255.0))) for i in range(256)]
        image = image.point(lut * 3)
        applied.append(f"gamma_{gamma:.2f}")

    target_bytes = int(rules.get("target_kb") or 50) * 1024
    quality_start = int(rules.get("quality_start") or 90)
    quality_min = int(rules.get("quality_min") or 30)
    quality_step = int(rules.get("quality_step") or 8)
    working = image
    output_path.parent.mkdir(parents=True, exist_ok=True)

    final_quality = quality_start
    compression_iterations = 0
    while True:
        for quality in range(quality_start, quality_min - 1, -quality_step):
            final_quality = quality
            save_kwargs: dict[str, Any] = {
                "format": "WEBP",
                "quality": quality,
                "method": 6,
            }
            if bool(rules.get("strip_metadata", True)):
                save_kwargs["exif"] = b""
                save_kwargs["icc_profile"] = None
            working.save(output_path, **save_kwargs)
            if output_path.exists() and output_path.stat().st_size <= target_bytes:
                applied.append(f"compress_webp_q{final_quality}")
                if bool(rules.get("strip_metadata", True)):
                    applied.append("strip_metadata")
                return {"applied_operations": applied, "original_size": orig_size, "final_size": working.size}

        if working.width <= 700:
            applied.append(f"compress_webp_q{final_quality}")
            if bool(rules.get("strip_metadata", True)):
                applied.append("strip_metadata")
            return {"applied_operations": applied, "original_size": orig_size, "final_size": working.size}
        compression_iterations += 1
        resized_width = max(700, int(working.width * 0.88))
        resized_height = max(1, int(working.height * resized_width / float(working.width)))
        working = working.resize((resized_width, resized_height), Image.LANCZOS)
        applied.append(f"compress_iteration_{compression_iterations}_resize_to_{resized_width}x{resized_height}")


def process_session_images_with_pillow(session_id: str) -> None:
    try:
      state = read_session_state(session_id)
    except Exception as exc:
      print(f"Warning: cannot process images for missing session {session_id}: {exc}")
      return

    files = state.get("files") if isinstance(state, dict) else None
    if not isinstance(files, dict):
      return
    images = list(files.get("images") or [])
    if not images:
      return

    rules = load_image_pillow_rules()
    processed_dir = APP_SESSION_ROOT / session_id / "images_processed"
    processed_dir.mkdir(parents=True, exist_ok=True)

    changed = False
    errors: list[str] = []
    processed_count = 0
    total_count = len([item for item in images if isinstance(item, dict)])
    renamed: dict[str, str] = {}
    for item in images:
      if not isinstance(item, dict):
        continue
      # Always prefer the immutable original upload as processing source on reruns.
      original_path_value = str(item.get("original_path") or "").strip()
      source_path_value = original_path_value or str(item.get("path") or "").strip()
      source_path = Path(source_path_value)
      if original_path_value and (not source_path.exists() or not source_path.is_file()):
        source_path = Path(str(item.get("path") or "").strip())
      if not source_path.exists() or not source_path.is_file():
        filename = str(item.get("filename") or "").strip() or "unknown"
        errors.append(f"missing_source:{filename}")
        continue

      old_filename = str(item.get("filename") or source_path.name).strip()
      output_filename = f"{Path(old_filename).stem}.webp"
      output_path = processed_dir / output_filename
      image_rules = dict(rules)
      vision_analysis = item.get("vision_analysis") if isinstance(item.get("vision_analysis"), dict) else {}
      vision_focus = parse_vision_crop_focus((vision_analysis or {}).get("crop_focus"))
      if vision_focus:
        image_rules["focal_x"] = vision_focus["x"]
        image_rules["focal_y"] = vision_focus["y"]
        item["vision_crop_focus"] = vision_focus
      try:
        result = save_webp_under_target(source_path, output_path, image_rules)
      except Exception as exc:
        item["processing_error"] = str(exc)
        errors.append(f"processing_error:{old_filename}:{exc}")
        changed = True
        continue

      if not output_path.exists():
        message = f"output_missing_after_processing:{old_filename}"
        item["processing_error"] = message
        errors.append(message)
        changed = True
        continue

      if not original_path_value:
        item["original_path"] = str(source_path)
      item["path"] = str(output_path)
      item["filename"] = output_filename
      item["content_type"] = "image/webp"
      item["size"] = output_path.stat().st_size
      item["processed_at"] = datetime.now(timezone.utc).isoformat()
      if isinstance(result, dict):
        item["applied_operations"] = result.get("applied_operations", [])
        item["original_dimensions"] = result.get("original_size")
        item["final_dimensions"] = result.get("final_size")
      processed_count += 1
      if old_filename and old_filename != output_filename:
        renamed[old_filename] = output_filename
      changed = True

      # Save progress after each processed image so UI can unlock compare per file.
      state["files"]["images"] = images
      state["image_processing"] = {
        "status": "processing",
        "processed_count": processed_count,
        "total_count": total_count,
        "warnings_count": len(errors),
        "warnings": errors[:20],
        "rules": rules,
        "updated_at": datetime.now(timezone.utc).isoformat(),
      }
      upload_session_media_to_gcs(session_id, state)
      write_session_state(session_id, state)

    if renamed:
      featured_filename = str(files.get("featured_image_filename") or "").strip()
      if featured_filename in renamed:
        files["featured_image_filename"] = renamed[featured_filename]
      selected_video = str(files.get("selected_video_filename") or "").strip()
      if selected_video in renamed:
        files["selected_video_filename"] = renamed[selected_video]

    state["files"]["images"] = images
    state.setdefault("image_processing", {})
    state["image_processing"] = {
      "status": "complete",
      "processed_count": processed_count,
      "total_count": total_count,
      "warnings_count": len(errors),
      "warnings": errors[:20],
      "rules": rules,
      "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    upload_session_media_to_gcs(session_id, state)
    write_session_state(session_id, state)


def state_with_vision_context(state: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(state, dict):
      return state
    vision_payload = state.get("vision") if isinstance(state.get("vision"), dict) else {}
    summaries = vision_payload.get("summaries") if isinstance(vision_payload.get("summaries"), list) else []
    if not summaries:
      return state

    lines: list[str] = []
    for item in summaries:
      if not isinstance(item, dict):
        continue
      filename = str(item.get("filename") or "").strip()
      summary = str(item.get("summary") or "").strip()
      if not filename or not summary:
        continue
      lines.append(f"- {filename}: {summary}")
    if not lines:
      return state

    cloned = json.loads(json.dumps(state, ensure_ascii=False))
    transcript = cloned.setdefault("transcript", {})
    base_text = str(transcript.get("text") or "").strip()
    vision_block = "Bildanalyse (selektierte Fotos):\n" + "\n".join(lines)
    transcript["text"] = f"{base_text}\n\n{vision_block}" if base_text else vision_block
    return cloned


def selected_vision_filenames(state: dict[str, Any]) -> list[str]:
    files = state.get("files") if isinstance(state, dict) else None
    if not isinstance(files, dict):
      return []
    selected = files.get("vision_selected_filenames")
    if not isinstance(selected, list):
      return []
    return [str(value).strip() for value in selected if str(value).strip()]


def detect_image_mime(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
      return "image/jpeg"
    if suffix == ".png":
      return "image/png"
    if suffix == ".webp":
      return "image/webp"
    if suffix == ".gif":
      return "image/gif"
    return "application/octet-stream"


def parse_vision_crop_focus(value: Any) -> dict[str, float] | None:
    if isinstance(value, dict):
      x_value = value.get("x")
      y_value = value.get("y")
    elif isinstance(value, (list, tuple)) and len(value) >= 2:
      x_value, y_value = value[0], value[1]
    else:
      return None

    try:
      x = float(x_value)
      y = float(y_value)
    except (TypeError, ValueError):
      return None

    if not (0.0 <= x <= 1.0 and 0.0 <= y <= 1.0):
      return None
    return {"x": x, "y": y}


AI_MODEL_PRICING_PER_MILLION: dict[str, dict[str, float]] = {
    "gpt-4o-mini": {"input": 0.15, "output": 0.60},
}


def estimate_ai_cost_usd(model: str, prompt_tokens: int, completion_tokens: int) -> float | None:
    pricing = AI_MODEL_PRICING_PER_MILLION.get(model)
    if not pricing:
      return None
    return (prompt_tokens * pricing["input"] + completion_tokens * pricing["output"]) / 1_000_000


def track_ai_usage(
    state: dict[str, Any],
    *,
    call_name: str,
    model: str,
  service: str = "openai_text",
    prompt_tokens: int = 0,
    completion_tokens: int = 0,
    total_tokens: int = 0,
    estimated_cost_usd: float | None = None,
    unknown_usage: bool = False,
)-> dict[str, Any]:
    usage = state.setdefault("ai_usage", {})
    usage.setdefault("call_count", 0)
    usage.setdefault("prompt_tokens", 0)
    usage.setdefault("completion_tokens", 0)
    usage.setdefault("total_tokens", 0)
    usage.setdefault("estimated_cost_usd", 0.0)
    usage.setdefault("unknown_usage_calls", 0)
    usage.setdefault("services", {})

    usage["call_count"] = int(usage.get("call_count") or 0) + 1
    usage["prompt_tokens"] = int(usage.get("prompt_tokens") or 0) + int(prompt_tokens or 0)
    usage["completion_tokens"] = int(usage.get("completion_tokens") or 0) + int(completion_tokens or 0)
    resolved_total = int(total_tokens or (int(prompt_tokens or 0) + int(completion_tokens or 0)))
    usage["total_tokens"] = int(usage.get("total_tokens") or 0) + resolved_total
    if estimated_cost_usd is not None:
      usage["estimated_cost_usd"] = float(usage.get("estimated_cost_usd") or 0.0) + float(estimated_cost_usd)
    if unknown_usage:
      usage["unknown_usage_calls"] = int(usage.get("unknown_usage_calls") or 0) + 1

    service_key = re.sub(r"[^a-z0-9_]+", "_", str(service or "openai_text").lower()).strip("_") or "openai_text"
    services = usage.setdefault("services", {})
    service_stats = services.setdefault(
      service_key,
      {
        "service": service_key,
        "call_count": 0,
        "prompt_tokens": 0,
        "completion_tokens": 0,
        "total_tokens": 0,
        "estimated_cost_usd": 0.0,
        "unknown_usage_calls": 0,
      },
    )
    service_stats["call_count"] = int(service_stats.get("call_count") or 0) + 1
    service_stats["prompt_tokens"] = int(service_stats.get("prompt_tokens") or 0) + int(prompt_tokens or 0)
    service_stats["completion_tokens"] = int(service_stats.get("completion_tokens") or 0) + int(completion_tokens or 0)
    service_stats["total_tokens"] = int(service_stats.get("total_tokens") or 0) + resolved_total
    if estimated_cost_usd is not None:
      service_stats["estimated_cost_usd"] = float(service_stats.get("estimated_cost_usd") or 0.0) + float(estimated_cost_usd)
    if unknown_usage:
      service_stats["unknown_usage_calls"] = int(service_stats.get("unknown_usage_calls") or 0) + 1
    service_stats["updated_at"] = datetime.now(timezone.utc).isoformat()
    service_stats["estimated_cost_usd"] = round(float(service_stats.get("estimated_cost_usd") or 0.0), 8)

    usage["last_call"] = {
      "call": call_name,
      "model": model,
      "service": service_key,
      "prompt_tokens": int(prompt_tokens or 0),
      "completion_tokens": int(completion_tokens or 0),
      "total_tokens": resolved_total,
      "estimated_cost_usd": round(float(estimated_cost_usd), 8) if estimated_cost_usd is not None else None,
      "unknown_usage": bool(unknown_usage),
      "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    usage["updated_at"] = usage["last_call"]["updated_at"]
    usage["estimated_cost_usd"] = round(float(usage.get("estimated_cost_usd") or 0.0), 8)
    return usage


def track_ai_usage_from_response(
    state: dict[str, Any],
    *,
    call_name: str,
    model: str,
    service: str,
    response: Any,
) -> dict[str, Any]:
    usage = getattr(response, "usage", None)
    if usage is None:
      return track_ai_usage(state, call_name=call_name, model=model, service=service, unknown_usage=True)
    prompt_tokens = int(getattr(usage, "prompt_tokens", 0) or 0)
    completion_tokens = int(getattr(usage, "completion_tokens", 0) or 0)
    total_tokens = int(getattr(usage, "total_tokens", prompt_tokens + completion_tokens) or (prompt_tokens + completion_tokens))
    estimated_cost_usd = estimate_ai_cost_usd(model, prompt_tokens, completion_tokens)
    return track_ai_usage(
      state,
      call_name=call_name,
      model=model,
      service=service,
      prompt_tokens=prompt_tokens,
      completion_tokens=completion_tokens,
      total_tokens=total_tokens,
      estimated_cost_usd=estimated_cost_usd,
      unknown_usage=estimated_cost_usd is None,
    )


def track_ai_usage_events(
    state: dict[str, Any],
    usage_events: list[dict[str, Any]] | None,
    *,
    default_service: str,
  ) -> dict[str, Any] | None:
    if not usage_events:
      return None
    snapshot = None
    for event in usage_events:
      if not isinstance(event, dict):
        continue
      model = str(event.get("model") or "gpt-4o-mini")
      snapshot = track_ai_usage(
        state,
        call_name=str(event.get("call") or "draft_ai_call"),
        model=model,
        service=str(event.get("service") or default_service),
        prompt_tokens=int(event.get("prompt_tokens") or 0),
        completion_tokens=int(event.get("completion_tokens") or 0),
        total_tokens=int(event.get("total_tokens") or 0),
        estimated_cost_usd=(float(event["estimated_cost_usd"]) if event.get("estimated_cost_usd") is not None else None),
        unknown_usage=event.get("estimated_cost_usd") is None,
      )
    return snapshot


def analyze_selected_images_with_vision(session_id: str) -> dict[str, Any]:
    if not OPENAI_API_KEY or OpenAI is None:
      return {"enabled": False, "error": "Vision analysis unavailable because OPENAI_API_KEY or openai package is missing."}

    state = read_session_state(session_id)
    files = state.get("files") if isinstance(state.get("files"), dict) else {}
    selected = selected_vision_filenames(state)
    images = [item for item in list(files.get("images") or []) if isinstance(item, dict)]
    if not selected:
      return {"enabled": False, "error": "No images selected for vision analysis."}

    selected_items = [item for item in images if str(item.get("filename") or "").strip() in set(selected)]
    if not selected_items:
      return {"enabled": False, "error": "Selected images are not available in this session."}

    client = OpenAI(api_key=OPENAI_API_KEY)
    summaries: list[dict[str, Any]] = []

    for item in selected_items:
      filename = str(item.get("filename") or "").strip()
      path = Path(str(item.get("path") or "").strip())
      if not path.exists() or not path.is_file():
        summaries.append({"filename": filename, "error": "File not found in session storage."})
        continue
      image_bytes = path.read_bytes()
      mime = detect_image_mime(path)
      data_uri = f"data:{mime};base64,{base64.b64encode(image_bytes).decode('ascii')}"
      try:
        response = client.chat.completions.create(
          model="gpt-4o-mini",
          temperature=0.2,
          messages=[
            {
              "role": "system",
              "content": (
                "You analyze event photos for a German event-post workflow. "
                "Return concise factual notes in German: venue clues, service setup, drinks, mood, crowd, highlights. "
                "No marketing fluff."
              ),
            },
            {
              "role": "user",
              "content": [
                {"type": "text", "text": "Beschreibe dieses Bild in 4-6 sachlichen Stichpunkten für die spätere Post-Erstellung."},
                {"type": "image_url", "image_url": {"url": data_uri}},
              ],
            },
          ],
        )
        summary = str(response.choices[0].message.content or "").strip()
        summaries.append({"filename": filename, "summary": summary})
        track_ai_usage_from_response(
          state,
          call_name="vision_batch_summary",
          model="gpt-4o-mini",
          service="openai_vision",
          response=response,
        )
      except Exception as exc:
        summaries.append({"filename": filename, "error": str(exc)})

    vision_result = {
      "enabled": True,
      "model": "gpt-4o-mini",
      "selected_filenames": selected,
      "summaries": summaries,
      "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    state["vision"] = vision_result
    write_session_state(session_id, state)
    return vision_result


def apply_session_image_metadata_to_wordpress_assets(
  state: dict[str, Any],
  post_response: dict[str, Any],
  import_logs: dict[str, Any],
) -> dict[str, Any]:
  files = state.get("files") if isinstance(state.get("files"), dict) else {}
  images = [item for item in list(files.get("images") or []) if isinstance(item, dict)]
  if not images:
    return {"applied": 0, "featured_set": False, "reason": "no_session_images"}

  plan = import_logs.get("media_upload_plan.json") if isinstance(import_logs, dict) else None
  media_plan = [item for item in list(plan or []) if isinstance(item, dict)]
  if not media_plan:
    return {"applied": 0, "featured_set": False, "reason": "no_media_upload_plan"}

  applied = 0
  for uploaded in media_plan:
    media_id = uploaded.get("media_id")
    if not media_id:
      continue
    source = find_media_item_by_filename(images, str(uploaded.get("filename") or ""))
    if not source:
      continue
    metadata = source.get("wp_metadata") if isinstance(source.get("wp_metadata"), dict) else {}
    alt_text = str(metadata.get("alt_text") or "").strip()
    title = str(metadata.get("title") or "").strip()
    caption = str(metadata.get("caption") or "").strip()
    description = str(metadata.get("description") or "").strip()
    if not any((alt_text, title, caption, description)):
      continue
    update_media_metadata(
      media_id=int(media_id),
      alt_text=alt_text or None,
      title=title or None,
      caption=caption or None,
      description=description or None,
    )
    applied += 1

  featured_filename = str(files.get("featured_image_filename") or "").strip()
  featured_set = False
  post_id = int(post_response.get("post_id") or 0)
  if featured_filename and post_id:
    featured_upload = find_media_item_by_filename(media_plan, featured_filename)
    featured_media_id = int((featured_upload or {}).get("media_id") or 0)
    if featured_media_id:
      set_post_featured_media(post_id=post_id, media_id=featured_media_id)
      featured_set = True

  return {
    "applied": applied,
    "featured_set": featured_set,
    "featured_filename": featured_filename,
  }


def analyze_image_metadata_with_vision(state: dict[str, Any], filename: str) -> dict[str, Any]:
    if not OPENAI_API_KEY or OpenAI is None:
        raise RuntimeError("Vision analysis unavailable because OPENAI_API_KEY or openai package is missing.")

    files = state.get("files") if isinstance(state.get("files"), dict) else {}
    images = [item for item in list(files.get("images") or []) if isinstance(item, dict)]
    match = find_media_item_by_filename(images, filename)
    if not match:
        raise FileNotFoundError("Image not found in this session.")

    path = Path(str(match.get("path") or "").strip())
    if not path.exists() or not path.is_file():
        raise FileNotFoundError("Image file is missing for vision analysis.")

    image_bytes = path.read_bytes()
    mime = detect_image_mime(path)
    data_uri = f"data:{mime};base64,{base64.b64encode(image_bytes).decode('ascii')}"

    featured_filename = str(files.get("featured_image_filename") or "").strip()
    role = "featured" if str(match.get("filename") or "").strip() == featured_filename else "gallery"
    draft_payload = state.get("draft") if isinstance(state.get("draft"), dict) else {}
    draft_row = draft_payload.get("row") if isinstance(draft_payload.get("row"), dict) else {}
    seed_metadata = match.get("wp_metadata") if isinstance(match.get("wp_metadata"), dict) else {}
    if not seed_metadata:
      seed_metadata = suggest_image_metadata_for_state(state, str(match.get("filename") or filename))

    seo_context = {
      "purpose": "SEO + Accessibility metadata for WordPress media library",
      "role": role,
      "filename": str(match.get("original_filename") or match.get("filename") or filename),
      "post_title": str(draft_row.get("post_title") or draft_row.get("title") or "").strip(),
      "category": str(draft_payload.get("category") or "").strip(),
      "tags": str(draft_row.get("tags") or state.get("post_type") or "").strip(),
      "seed_metadata": {
        "alt_text": str(seed_metadata.get("alt_text") or "").strip(),
        "title": str(seed_metadata.get("title") or "").strip(),
        "caption": str(seed_metadata.get("caption") or "").strip(),
        "description": str(seed_metadata.get("description") or "").strip(),
      },
    }

    client = OpenAI(api_key=OPENAI_API_KEY)
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        temperature=0.2,
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "system",
                "content": (
                "You are a strict SEO + accessibility metadata assistant for event photos in WordPress. "
                "Return concise German output as JSON only. "
                "Prioritize factual accuracy, search intent relevance, and readability. "
                "No keyword stuffing, no hype language, no invented facts."
                ),
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                    "Analysiere dieses Eventbild für WordPress-Mediathek-SEO. "
                    "Nutze den folgenden Kontext:")
                },
                {
                  "type": "text",
                  "text": json.dumps(seo_context, ensure_ascii=False),
                },
                {
                  "type": "text",
                  "text": (
                    "Gib JSON mit exakt diesen Keys zurück: alt_text, title, caption, description, issues, advice. "
                    "Zusätzlich darf crop_focus als Objekt mit x und y zwischen 0 und 1 zurückgegeben werden, "
                    "wenn der Bildaufbau einen klaren Mittelpunkt hat. "
                    "Regeln: "
                    "1) alt_text: 8-18 Wörter, konkrete sichtbare Szene, keine Füllwörter, kein 'Bild von'. "
                    "2) title: 35-70 Zeichen, prägnant, eventbezogen. "
                    "3) caption: 70-160 Zeichen, informativer Kontextsatz. "
                    "4) description: 140-280 Zeichen, SEO-tauglich, natürlich lesbar, ohne Keyword-Stuffing. "
                    "5) Nutze relevante Begriffe aus post_title/category/tags nur natürlich und höchstens einmal pro Feld. "
                    "6) issues: kurze Stichpunkte zu Bildproblemen (Schärfe, Belichtung, störende Elemente, Branding-Risiken, Datenschutz). "
                    "7) advice: kurze, umsetzbare Verbesserungsvorschläge für SEO und Bildqualität. "
                    "8) crop_focus: wenn ein Motiv klar erkennbar ist, nenne dessen Mittelpunkt als x/y im Bereich 0..1. "
                    "Wenn der Ausschnitt bereits gut sitzt oder kein sicherer Mittelpunkt erkennbar ist, lass crop_focus leer oder null. "
                    "9) Prüfe ausdrücklich auf schlechten Zuschnitt (z.B. angeschnittene Köpfe/Produkte, abgeschnittenes Hauptmotiv). "
                    "Wenn ein solcher Zuschnitt vorliegt, nenne es in issues und gib nach Möglichkeit crop_focus für einen besseren Fokus an. "
                    "Wenn kein Problem sichtbar ist, gib issues als leeres Array zurück."
                        ),
                    },
                    {"type": "image_url", "image_url": {"url": data_uri}},
                ],
            },
        ],
    )
    usage_snapshot = track_ai_usage_from_response(
      state,
      call_name="vision_image_metadata",
      model="gpt-4o-mini",
      service="openai_vision",
      response=response,
    )
    raw_content = str(response.choices[0].message.content or "{}").strip()
    try:
        parsed = json.loads(raw_content)
    except Exception:
        parsed = {}

    def as_text(value: Any) -> str:
        return str(value or "").strip()

    def as_list(value: Any) -> list[str]:
        if isinstance(value, list):
            return [str(item).strip() for item in value if str(item).strip()]
        if isinstance(value, str) and value.strip():
            return [value.strip()]
        return []

    crop_focus = parse_vision_crop_focus(parsed.get("crop_focus"))

    metadata = {
        "alt_text": as_text(parsed.get("alt_text")),
        "title": as_text(parsed.get("title")),
        "caption": as_text(parsed.get("caption")),
        "description": as_text(parsed.get("description")),
    }
    issues = as_list(parsed.get("issues"))
    advice = as_list(parsed.get("advice"))

    if not any(metadata.values()):
        metadata = suggest_image_metadata_for_state(state, str(match.get("filename") or filename))
        if not advice:
            advice = ["Vision-Antwort war unvollständig; ohne Vision-Fallback wurde verwendet."]

    return {
        "filename": str(match.get("filename") or filename),
        "metadata": metadata,
        "issues": issues,
        "advice": advice,
        "crop_focus": crop_focus,
        "model": "gpt-4o-mini",
      "usage": usage_snapshot.get("last_call", {}),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


def build_session_media_library_items(state: dict[str, Any]) -> list[dict[str, Any]]:
    files = state.get("files") if isinstance(state.get("files"), dict) else {}
    images = [item for item in list(files.get("images") or []) if isinstance(item, dict)]
    featured_filename = str(files.get("featured_image_filename") or "").strip()
    items: list[dict[str, Any]] = []
    for image in images:
      filename = str(image.get("filename") or "").strip()
      path_value = str(image.get("path") or "").strip()
      if not filename or not path_value:
        continue
      path = Path(path_value)
      if not path.exists() or not path.is_file():
        continue
      role = "featured" if filename == featured_filename else "gallery"
      items.append({
        "filename": filename,
        "path": str(path),
        "role": role,
        "alt_text": str(image.get("alt_text") or "").strip() or filename,
        "title": str(image.get("title") or image.get("original_filename") or path.stem).strip(),
        "caption": str(image.get("caption") or "").strip(),
        "description": str(image.get("description") or "").strip(),
      })
    return items


def upload_session_images_to_wordpress_media_library(session_id: str) -> dict[str, Any]:
    state = read_session_state(session_id)
    client_id = str(state.get("client_id") or "flairlab")
    set_active_client(client_id)
    plan = build_session_media_library_items(state)
    if not plan:
      raise HTTPException(status_code=400, detail="No images available to upload to WordPress media library.")

    uploaded: list[dict[str, Any]] = []
    for item in plan:
      media_id, media_url = upload_media(item["path"])
      try:
        update_media_metadata(
          media_id=media_id,
          alt_text=item.get("alt_text"),
          title=item.get("title"),
          caption=item.get("caption"),
          description=item.get("description"),
        )
      except Exception:
        pass
      uploaded.append({
        **item,
        "media_id": media_id,
        "source_url": media_url,
      })

    state["wordpress_media_library"] = {
      "uploaded_items": uploaded,
      "uploaded_count": len(uploaded),
      "uploaded_at": datetime.now(timezone.utc).isoformat(),
      "client_id": client_id,
    }
    write_session_state(session_id, state)
    return state["wordpress_media_library"]


def load_workbook_specs(post_type: str | None = None) -> list[ColumnSpec]:
    workbook_guidance = load_workbook_guidance(
        active_knowledge_workbook_path(),
        post_type=post_type,
        preferred_sheet=KNOWLEDGE_WORKBOOK_SHEET,
    )

    specs: list[ColumnSpec] = []
    by_source: dict[str, int] = {}
    post_type_lookup = normalize_lookup(str(post_type or "event"))
    mapping_sheet_lookup = normalize_lookup(f"{post_type_lookup}_acf_mapping")
    schema_sheet_lookup = normalize_lookup(f"{post_type_lookup}_field_schema")
    allowed_spec_sheet_lookups = {"sharedfieldschema", mapping_sheet_lookup, schema_sheet_lookup}

    for item in workbook_guidance.get("items", []):
      source_sheet_lookup = normalize_lookup(str(item.get("source_sheet") or "").strip())
      # Keep generated CSV/draft fields constrained to the shared schema and
      # the active post-type mapping tab. Other sheets still enrich guidance.
      if source_sheet_lookup and source_sheet_lookup not in allowed_spec_sheet_lookups:
        continue

      source_name = str(item.get("user_field_name") or "").strip()
      acf_name = str(item.get("acf_field_name") or "").strip()
      group = normalize_lookup(str(item.get("group") or "").strip())
      output_target = normalize_lookup(str(item.get("output_target") or "").strip())
      is_acf_group = group in {"acf", "advancedcustomfields", "advancedcustomfield", "acr"}
      is_acf_target = output_target in {"acf", "acfpayload"}

      if not acf_name and source_name and (is_acf_group or is_acf_target):
        acf_name = source_name
      if not source_name and acf_name:
        source_name = acf_name
      if not source_name:
        continue

      guidance_text = str(item.get("ai_guidance") or "").strip()
      min_words = item.get("min_words")
      max_words = item.get("max_words")
      if is_faq_field_name(source_name) or is_faq_field_name(acf_name):
        guidance_text = merge_guidance_text(guidance_text, default_faq_guidance(source_name))
      marker = "ACF" if (is_acf_group or is_acf_target or acf_name) else ""

      source_key = normalize_lookup(source_name)
      existing_index = by_source.get(source_key)
      if existing_index is None:
        spec = ColumnSpec(
          index=len(specs),
          marker=marker,
          acf_name=acf_name,
          source_name=source_name,
          guidance=guidance_text,
          min_words=min_words,
          max_words=max_words,
          display_name=source_name or acf_name,
        )
        by_source[source_key] = len(specs)
        specs.append(spec)
        continue

      # Merge duplicate source rows from multiple workbook tabs.
      existing = specs[existing_index]
      if acf_name:
        existing.acf_name = acf_name
      if marker:
        existing.marker = marker
      if guidance_text:
        combined = [item for item in [existing.guidance, guidance_text] if item]
        existing.guidance = "\n".join(dict.fromkeys(combined))
      if existing.min_words is None and min_words is not None:
        existing.min_words = min_words
      if existing.max_words is None and max_words is not None:
        existing.max_words = max_words
      if not existing.display_name:
        existing.display_name = existing.source_name or existing.acf_name

    for spec in specs:
        guidance_items = guidance_for_field(workbook_guidance, spec.source_name, spec.acf_name)
        if guidance_items:
            combined = [item for item in [spec.guidance, *guidance_items] if item]
            spec.guidance = "\n".join(dict.fromkeys(combined))

        if spec.min_words is None or spec.max_words is None:
            for item in workbook_guidance.get("items", []):
                if field_matches(item.get("user_field_name", ""), normalize_lookup(spec.source_name)) or \
                   field_matches(item.get("acf_field_name", ""), normalize_lookup(spec.acf_name)):
                    if spec.min_words is None:
                        spec.min_words = item.get("min_words")
                    if spec.max_words is None:
                        spec.max_words = item.get("max_words")
                    break

        if not spec.display_name:
            spec.display_name = spec.source_name or spec.acf_name

    return specs


def resolve_app_path(path_value: str | Path) -> Path:
    path = Path(path_value)
    return path if path.is_absolute() else APP_ROOT / path


def configured_knowledge_workbook_path() -> Path:
    if KNOWLEDGE_WORKBOOK_PATH:
        return resolve_app_path(KNOWLEDGE_WORKBOOK_PATH)
    return DEFAULT_KNOWLEDGE_WORKBOOK_PATH


def configured_knowledge_workbook_gcs_uri() -> str:
  return str(KNOWLEDGE_WORKBOOK_GCS_URI or "").strip()


def configured_session_state_gcs_prefix() -> str:
  return str(SESSION_STATE_GCS_PREFIX or "").strip()


def parse_gcs_uri(uri: str) -> tuple[str, str]:
  raw = str(uri or "").strip()
  if not raw.startswith("gs://"):
    raise ValueError("KNOWLEDGE_WORKBOOK_GCS_URI must use gs://bucket/path format.")
  bucket_name, _, blob_name = raw[5:].partition("/")
  if not bucket_name or not blob_name:
    raise ValueError("KNOWLEDGE_WORKBOOK_GCS_URI must include both bucket and object path.")
  return bucket_name, blob_name


def parse_gcs_prefix(uri: str) -> tuple[str, str]:
  raw = str(uri or "").strip()
  if not raw.startswith("gs://"):
    raise ValueError("SESSION_STATE_GCS_PREFIX must use gs://bucket[/prefix] format.")
  path = raw[5:]
  bucket_name, _, prefix = path.partition("/")
  if not bucket_name:
    raise ValueError("SESSION_STATE_GCS_PREFIX must include a bucket name.")
  return bucket_name, prefix.strip("/")


def gcs_client() -> Any:
  global _gcs_client
  if _gcs_client is not None:
    return _gcs_client
  if gcs_storage is None:
    raise RuntimeError("google-cloud-storage is not installed.")
  _gcs_client = gcs_storage.Client()
  return _gcs_client


def gcs_blob_from_uri(uri: str):
  bucket_name, blob_name = parse_gcs_uri(uri)
  client = gcs_client()
  bucket = client.bucket(bucket_name)
  return bucket.blob(blob_name)


def gcs_blob_for_session_state(session_id: str):
  prefix_uri = configured_session_state_gcs_prefix()
  if not prefix_uri:
    return None
  bucket_name, prefix = parse_gcs_prefix(prefix_uri)
  blob_name = f"{session_id}/state.json" if not prefix else f"{prefix}/{session_id}/state.json"
  client = gcs_client()
  bucket = client.bucket(bucket_name)
  return bucket.blob(blob_name)


def gcs_blob_for_session_media(session_id: str, media_kind: str, filename: str):
  prefix_uri = configured_session_state_gcs_prefix()
  if not prefix_uri:
    return None
  bucket_name, prefix = parse_gcs_prefix(prefix_uri)
  safe_filename = Path(filename).name
  blob_name = f"{session_id}/media/{media_kind}/{safe_filename}" if not prefix else f"{prefix}/{session_id}/media/{media_kind}/{safe_filename}"
  client = gcs_client()
  bucket = client.bucket(bucket_name)
  return bucket.blob(blob_name)


def upload_session_media_to_gcs(session_id: str, state: dict[str, Any]) -> None:
  if not configured_session_state_gcs_prefix():
    return
  for media_kind, item in iter_session_media_items(state):
    filename = str(item.get("filename") or "").strip()
    path_value = str(item.get("path") or "").strip()
    if not filename or not path_value:
      continue
    path = Path(path_value)
    if not path.exists() or not path.is_file():
      continue
    blob = gcs_blob_for_session_media(session_id, media_kind, filename)
    if blob is None:
      continue
    try:
      blob.upload_from_filename(str(path))
    except Exception as exc:
      print(f"Warning: failed to upload session media to GCS for {session_id}/{media_kind}/{filename}: {exc}")

    if media_kind == "images":
      original_path_value = str(item.get("original_path") or "").strip()
      if not original_path_value:
        continue
      original_path = Path(original_path_value)
      if not original_path.exists() or not original_path.is_file():
        continue
      original_name = Path(original_path_value).name
      original_blob = gcs_blob_for_session_media(session_id, "images_original", original_name)
      if original_blob is None:
        continue
      try:
        original_blob.upload_from_filename(str(original_path))
      except Exception as exc:
        print(f"Warning: failed to upload original image to GCS for {session_id}/images_original/{original_name}: {exc}")


def sync_session_media_from_gcs(session_id: str, state: dict[str, Any]) -> bool:
  if not configured_session_state_gcs_prefix():
    return False
  changed = False
  for media_kind, item in iter_session_media_items(state):
    filename = str(item.get("filename") or "").strip()
    if not filename:
      continue
    path_value = str(item.get("path") or "").strip()
    local_path = Path(path_value) if path_value else (APP_SESSION_ROOT / session_id / media_kind / Path(filename).name)
    if local_path.exists() and local_path.is_file():
      continue
    blob = gcs_blob_for_session_media(session_id, media_kind, filename)
    if blob is None:
      continue
    try:
      if not blob.exists():
        continue
      local_path.parent.mkdir(parents=True, exist_ok=True)
      blob.download_to_filename(str(local_path))
      item["path"] = str(local_path)
      changed = True
    except Exception as exc:
      print(f"Warning: failed to sync session media from GCS for {session_id}/{media_kind}/{filename}: {exc}")

    if media_kind == "images":
      original_path_value = str(item.get("original_path") or "").strip()
      original_name = Path(original_path_value).name if original_path_value else ""
      if original_name:
        original_local = APP_SESSION_ROOT / session_id / "images_original" / original_name
        if not (original_local.exists() and original_local.is_file()):
          original_blob = gcs_blob_for_session_media(session_id, "images_original", original_name)
          if original_blob is not None:
            try:
              if original_blob.exists():
                original_local.parent.mkdir(parents=True, exist_ok=True)
                original_blob.download_to_filename(str(original_local))
                item["original_path"] = str(original_local)
                changed = True
            except Exception as exc:
              print(f"Warning: failed to sync original image from GCS for {session_id}/images_original/{original_name}: {exc}")
  return changed


def upload_session_state_to_gcs(session_id: str, local_state_path: Path) -> None:
  blob = gcs_blob_for_session_state(session_id)
  if blob is None:
    return
  try:
    blob.upload_from_filename(str(local_state_path), content_type="application/json")
  except Exception as exc:
    print(f"Warning: failed to upload session state to GCS for {session_id}: {exc}")


def sync_session_state_from_gcs(session_id: str) -> bool:
  blob = gcs_blob_for_session_state(session_id)
  if blob is None:
    return False
  try:
    if not blob.exists():
      return False
    local_path = APP_SESSION_ROOT / session_id / "state.json"
    local_path.parent.mkdir(parents=True, exist_ok=True)
    blob.download_to_filename(str(local_path))
    return True
  except Exception as exc:
    print(f"Warning: failed to sync session state from GCS for {session_id}: {exc}")
    return False


def list_gcs_session_states() -> list[dict[str, Any]]:
  prefix_uri = configured_session_state_gcs_prefix()
  if not prefix_uri:
    return []
  try:
    bucket_name, prefix = parse_gcs_prefix(prefix_uri)
    client = gcs_client()
    blob_prefix = f"{prefix}/" if prefix else ""
    items: list[dict[str, Any]] = []
    for blob in client.list_blobs(bucket_name, prefix=blob_prefix):
      if not blob.name.endswith("/state.json"):
        continue
      payload = blob.download_as_text(encoding="utf-8")
      state = json.loads(payload)
      if not isinstance(state, dict):
        continue
      if not state.get("session_id"):
        state["session_id"] = Path(blob.name).parent.name
      updated = blob.updated.isoformat() if getattr(blob, "updated", None) else None
      items.append({
        "state": state,
        "storage": "gcs",
        "storage_updated_at": updated,
      })
    return items
  except Exception as exc:
    print(f"Warning: failed to list session states from GCS: {exc}")
    return []


def list_local_session_states() -> list[dict[str, Any]]:
  if not APP_SESSION_ROOT.exists() or not APP_SESSION_ROOT.is_dir():
    return []
  items: list[dict[str, Any]] = []
  for session_dir in APP_SESSION_ROOT.iterdir():
    if not session_dir.is_dir():
      continue
    state_path = session_dir / "state.json"
    state = read_json_if_exists(state_path)
    if not isinstance(state, dict):
      continue
    if not state.get("session_id"):
      state["session_id"] = session_dir.name
    updated = datetime.fromtimestamp(state_path.stat().st_mtime, tz=timezone.utc).isoformat()
    items.append({
      "state": state,
      "storage": "local",
      "storage_updated_at": updated,
    })
  return items


def delete_local_session_state(session_id: str) -> bool:
  session_dir = APP_SESSION_ROOT / str(session_id or "").strip()
  if not session_dir.exists() or not session_dir.is_dir():
    return False
  shutil.rmtree(session_dir)
  return True


def delete_gcs_session_state(session_id: str) -> bool:
  prefix_uri = configured_session_state_gcs_prefix()
  if not prefix_uri:
    return False
  bucket_name, prefix = parse_gcs_prefix(prefix_uri)
  safe_session_id = str(session_id or "").strip().strip("/")
  if not safe_session_id:
    return False
  session_prefix = f"{prefix}/{safe_session_id}/" if prefix else f"{safe_session_id}/"
  client = gcs_client()
  deleted = False
  for blob in client.list_blobs(bucket_name, prefix=session_prefix):
    blob.delete()
    deleted = True
  return deleted


def session_last_updated_at(state: dict[str, Any], storage_updated_at: str | None = None) -> str:
  candidates = [
    state.get("updated_at"),
    state.get("created_at"),
    state.get("ui_cache", {}).get("updated_at"),
    state.get("transcript", {}).get("updated_at"),
    state.get("draft", {}).get("updated_at"),
    state.get("wordpress_post", {}).get("updated_at"),
    storage_updated_at,
  ]
  values = [str(value).strip() for value in candidates if value]
  return values[0] if values else ""


def session_sort_value(state: dict[str, Any], storage_updated_at: str | None = None) -> datetime:
  for value in (
    session_last_updated_at(state, storage_updated_at),
    str(state.get("created_at") or "").strip(),
    str(storage_updated_at or "").strip(),
  ):
    if not value:
      continue
    try:
      return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
      continue
  return datetime.fromtimestamp(0, tz=timezone.utc)


def workbook_metadata_path(workbook_path: Path) -> Path:
  return workbook_path.with_suffix(workbook_path.suffix + WORKBOOK_METADATA_SUFFIX)


def read_workbook_metadata(workbook_path: Path) -> dict[str, Any]:
  meta_path = workbook_metadata_path(workbook_path)
  if not meta_path.exists() or not meta_path.is_file():
    return {}
  try:
    payload = json.loads(meta_path.read_text(encoding="utf-8"))
    return payload if isinstance(payload, dict) else {}
  except json.JSONDecodeError:
    return {}


def write_workbook_metadata(workbook_path: Path, payload: dict[str, Any]) -> None:
  meta_path = workbook_metadata_path(workbook_path)
  meta_path.parent.mkdir(parents=True, exist_ok=True)
  meta_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def local_workbook_sha256(path: Path) -> str | None:
  if not path.exists() or not path.is_file():
    return None
  digest = hashlib.sha256()
  with path.open("rb") as handle:
    while chunk := handle.read(1024 * 1024):
      digest.update(chunk)
  return digest.hexdigest()


def workbook_version_info(workbook_path: Path) -> dict[str, Any]:
  info: dict[str, Any] = {
    "storage_mode": "gcs" if configured_knowledge_workbook_gcs_uri() else "local_file",
    "sha256": local_workbook_sha256(workbook_path),
  }
  if workbook_path.exists() and workbook_path.is_file():
    stat = workbook_path.stat()
    info["size_bytes"] = stat.st_size
    info["local_modified_at"] = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
  meta = read_workbook_metadata(workbook_path)
  if meta:
    info.update(
      {
        "gcs_uri": meta.get("gcs_uri"),
        "gcs_generation": meta.get("gcs_generation"),
        "gcs_updated": meta.get("gcs_updated"),
        "last_synced_at": meta.get("synced_at"),
      }
    )
  return info


def sync_knowledge_workbook_from_gcs(configured_path: Path) -> Path:
  gcs_uri = configured_knowledge_workbook_gcs_uri()
  if not gcs_uri:
    return configured_path

  blob = gcs_blob_from_uri(gcs_uri)
  if not blob.exists():
    raise RuntimeError(f"Workbook object does not exist in GCS: {gcs_uri}")
  blob.reload()
  configured_path.parent.mkdir(parents=True, exist_ok=True)
  blob.download_to_filename(str(configured_path))
  write_workbook_metadata(
    configured_path,
    {
      "gcs_uri": gcs_uri,
      "gcs_generation": str(blob.generation or ""),
      "gcs_updated": blob.updated.isoformat() if getattr(blob, "updated", None) else None,
      "synced_at": datetime.now(timezone.utc).isoformat(),
    },
  )
  return configured_path


def upload_knowledge_workbook_to_gcs(workbook_path: Path) -> dict[str, Any]:
  gcs_uri = configured_knowledge_workbook_gcs_uri()
  if not gcs_uri:
    return {}
  blob = gcs_blob_from_uri(gcs_uri)
  content_type = (
    "application/vnd.ms-excel.sheet.macroEnabled.12"
    if workbook_path.suffix.lower() == ".xlsm"
    else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
  )
  blob.upload_from_filename(str(workbook_path), content_type=content_type)
  blob.reload()
  return {
    "gcs_uri": gcs_uri,
    "gcs_generation": str(blob.generation or ""),
    "gcs_updated": blob.updated.isoformat() if getattr(blob, "updated", None) else None,
  }


def discover_knowledge_workbook_path(configured_path: Path) -> Path:
    if configured_path.exists() and configured_path.is_file():
        return configured_path

    knowledge_dir = configured_path.parent
    if not knowledge_dir.exists() or not knowledge_dir.is_dir():
        return configured_path

    candidates = [
        path
        for suffix in ("*.xlsm", "*.xlsx")
        for path in knowledge_dir.glob(suffix)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        return configured_path

    candidates.sort(
        key=lambda path: (
            path.suffix.lower() != ".xlsm",
            -path.stat().st_mtime,
            path.name.lower(),
        )
    )
    return candidates[0]


def ensure_knowledge_workbook_available() -> Path:
    global _knowledge_gcs_synced
    configured_path = configured_knowledge_workbook_path()
    gcs_uri = configured_knowledge_workbook_gcs_uri()
    policy = knowledge_source_policy()

    if policy == "local_only":
        if configured_path.exists() and configured_path.is_file():
            return configured_path

        discovered_path = discover_knowledge_workbook_path(configured_path)
        if discovered_path.exists() and discovered_path.is_file() and discovered_path != configured_path:
            configured_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(discovered_path, configured_path)
            return configured_path
        return discovered_path

    if policy in {"gcs_required", "gcs_preferred"}:
        if not gcs_uri:
            if policy == "gcs_required":
                raise RuntimeError(
                    "KNOWLEDGE_SOURCE_POLICY=gcs_required but KNOWLEDGE_WORKBOOK_GCS_URI is not configured."
                )
        else:
            try:
                # In GCS mode, GCS object is canonical source of truth.
                sync_knowledge_workbook_from_gcs(configured_path)
                _knowledge_gcs_synced = True
                return configured_path
            except Exception:
                if policy == "gcs_required":
                    raise

    if configured_path.exists() and configured_path.is_file():
        return configured_path

    discovered_path = discover_knowledge_workbook_path(configured_path)
    if discovered_path.exists() and discovered_path.is_file() and discovered_path != configured_path:
        configured_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(discovered_path, configured_path)
        return configured_path

    return discovered_path


def active_knowledge_workbook_path() -> Path:
    return ensure_knowledge_workbook_available()


def get_workbook_version_hash(workbook_path: Path) -> str:
  """Get hash of workbook file contents to detect version changes."""
  if not workbook_path.exists():
    return "missing"
  try:
    return hashlib.sha256(workbook_path.read_bytes()).hexdigest()[:16]
  except Exception:
    return "unknown"


def get_session_guidance_cached(
    state: dict[str, Any],
    workbook_path: Path | None,
    post_type: str | None,
    preferred_sheet: str | None,
) -> dict[str, Any]:
    """
    Get guidance data with session-level caching.
    Reloads if workbook version changes (detected via mtime hash).
    """
    if not workbook_path:
        return {"source": None, "items": [], "sheets_loaded": []}

    current_version = get_workbook_version_hash(workbook_path)
    cached_data = state.get("_guidance_cache")
    cached_version = state.get("_guidance_version")

    # Return cached if version matches
    if cached_data and cached_version == current_version:
        return cached_data

    # Load fresh and cache
    guidance_data = load_workbook_guidance(workbook_path, post_type, preferred_sheet)
    state["_guidance_cache"] = guidance_data
    state["_guidance_version"] = current_version
    return guidance_data


def get_session_internal_links_cached(
    state: dict[str, Any],
    workbook_path: Path | None,
) -> dict[str, Any]:
    if not workbook_path:
        return {"source": None, "database": [], "rules": []}

    current_version = get_workbook_version_hash(workbook_path)
    cached_data = state.get("_internal_links_cache")
    cached_version = state.get("_internal_links_version")

    if cached_data and cached_version == current_version:
        return cached_data

    context = load_internal_links_context(workbook_path)
    state["_internal_links_cache"] = context
    state["_internal_links_version"] = current_version
    return context


@app.on_event("startup")
def ensure_knowledge_on_startup() -> None:
  # On Cloud Run, protect against accidental local fallback by default.
  if os.getenv("K_SERVICE") and knowledge_source_policy() == "auto" and not configured_knowledge_workbook_gcs_uri():
    raise RuntimeError(
      "Cloud runtime detected but KNOWLEDGE_WORKBOOK_GCS_URI is missing. "
      "Set KNOWLEDGE_SOURCE_POLICY=local_only only if this is intentional."
    )
    ensure_knowledge_workbook_available()


def read_json_if_exists(path: Path) -> Any:
    if not path.exists() or not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"error": "Could not parse JSON file.", "path": str(path)}

def collect_wordpress_import_logs(output_dir: Path | None) -> dict[str, Any]:
    if not output_dir or not output_dir.exists():
        return {}
    logs: dict[str, Any] = {}
    for filename in (
        "technical_log.json",
        "warnings.json",
        "created_post.json",
        "wordpress_create_payload.json",
        "wordpress_payload_preview.json",
        "acf_payload.json",
        "advanced_custom_fields_update_response.json",
        "media_upload_plan.json",
        "media_upload_cache.json",
    ):
        value = read_json_if_exists(output_dir / filename)
        if value is not None:
            logs[filename] = value
    return logs


def knowledge_status_payload(post_type: str | None = None) -> dict[str, Any]:
  configured_path = configured_knowledge_workbook_path()
  path = active_knowledge_workbook_path()
  guidance = load_workbook_guidance(
    path,
    post_type=post_type,
    preferred_sheet=KNOWLEDGE_WORKBOOK_SHEET,
  )
  download_filename = path.name
  if path.suffix.lower() == ".xlsx":
    download_filename = f"{path.stem}.xlsm"

  acf_guidance_list: list[dict[str, Any]] = []
  try:
    specs = load_workbook_specs(post_type=post_type)
    for spec in specs:
      matched_items = guidance_items_for_field(guidance, spec.source_name, spec.acf_name)
      acf_guidance_list.append(
        {
          "user_field": spec.source_name,
          "acf_field": spec.acf_name,
          "guidance": spec.guidance,
          "source_sheets": list(
            dict.fromkeys(
              item.get("source_sheet", "")
              for item in matched_items
              if item.get("source_sheet")
            )
          ),
        }
      )
  except Exception:
    pass

  return {
    "path": str(path),
    "configured_path": str(configured_path),
    "filename": path.name,
    "exists": path.exists(),
    "storage_mode": "gcs" if configured_knowledge_workbook_gcs_uri() else "local_file",
    "knowledge_source_policy": knowledge_source_policy(),
    "gcs_uri": configured_knowledge_workbook_gcs_uri() or None,
    "workbook_version": workbook_version_info(path),
    "fallback_used": path != configured_path,
    "post_type": post_type,
    "configured_sheet": KNOWLEDGE_WORKBOOK_SHEET,
    "loaded_sheets": guidance.get("sheets_loaded", []),
    "guidance_items": len(guidance.get("items", [])),
    "guidance_preview": guidance.get("items", [])[:5],
    "acf_guidance_list": acf_guidance_list,
    "error": guidance.get("error"),
    "download_url": "/app/knowledge/workbook",
    "download_filename": download_filename,
  }


APP_HTML = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>FLAIRLAB Post Generator</title>
  <style>
    :root { color-scheme: light; --ink:#1f2933; --muted:#627386; --line:#d8e0e8; --brand:#0f766e; --brand-strong:#0b5f59; --soft:#f5f7fa; --paper:#ffffff; --danger:#b42318; }
    * { box-sizing: border-box; }
    body { margin:0; font-family: Inter, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; color:var(--ink); background:linear-gradient(180deg, #eef7f5 0, #f8fafc 260px, #ffffff 100%); }
    main { width:min(920px, 100%); margin:0 auto; padding:22px 16px 56px; }
    @media (min-width: 1260px) {
      main { margin-left:max(16px, calc((100vw - 1180px) / 2)); margin-right:380px; }
    }
    header { padding:16px 0 22px; margin-bottom:18px; }
    h1 { font-size:clamp(24px, 6vw, 38px); line-height:1.05; margin:0 0 8px; letter-spacing:0; }
    h2 { font-size:18px; margin:0 0 14px; }
    p { color:var(--muted); line-height:1.45; margin:0; }
    section { padding:18px 0; border-bottom:1px solid var(--line); }
    details.panel { background:rgba(255,255,255,.92); border:1px solid var(--line); border-radius:8px; padding:0; margin:0 0 12px; box-shadow:0 14px 34px rgba(31,41,51,.06); overflow:hidden; }
    details.panel { --panel-accent:var(--brand); --panel-soft:#eef7f5; }
    details.panel > summary { border-left:6px solid var(--panel-accent); background:linear-gradient(90deg, var(--panel-soft), #fff 78%); }
    #panelAccess { --panel-accent:#314352; --panel-soft:#eef2f6; }
    #panelKnowledge { --panel-accent:#8b5cf6; --panel-soft:#f4f0ff; }
    #panelSession { --panel-accent:#0f766e; --panel-soft:#eef7f5; }
    #panelArchive { --panel-accent:#475569; --panel-soft:#f8fafc; }
    #panelUpload { --panel-accent:#2563eb; --panel-soft:#eef4ff; }
    #panelTranscript { --panel-accent:#d97706; --panel-soft:#fff7ed; }
    #panelDraft { --panel-accent:#be123c; --panel-soft:#fff1f2; }
    #panelWordPress { --panel-accent:#15803d; --panel-soft:#f0fdf4; }
    #panelStatus { --panel-accent:#64748b; --panel-soft:#f8fafc; }
    details.panel > summary { list-style:none; display:flex; align-items:center; justify-content:space-between; gap:12px; padding:16px 18px; cursor:pointer; font-weight:800; font-size:18px; }
    details.panel > summary::-webkit-details-marker { display:none; }
    details.panel > summary::after { content:"+"; flex:0 0 auto; width:28px; height:28px; border-radius:8px; display:grid; place-items:center; background:var(--soft); border:1px solid var(--line); color:var(--muted); font-size:18px; line-height:1; }
    details.panel[open] > summary::after { content:"-"; }
    .panel-body { padding:0 18px 18px; }
    label { display:block; font-weight:650; font-size:14px; margin:14px 0 7px; }
    input, select, textarea, button { width:100%; font:inherit; border-radius:8px; }
    input, select, textarea { border:1px solid var(--line); padding:11px 12px; background:#fff; transition:border-color .15s ease, box-shadow .15s ease; }
    input:focus, select:focus, textarea:focus { outline:0; border-color:var(--brand); box-shadow:0 0 0 3px rgba(15,118,110,.12); }
    input[type="file"] { padding:10px; background:var(--soft); }
    textarea { min-height:220px; resize:vertical; line-height:1.45; }
    button { border:0; padding:12px 14px; background:var(--brand); color:white; font-weight:750; cursor:pointer; margin-top:14px; transition:transform .12s ease, background .15s ease, box-shadow .15s ease; box-shadow:0 10px 22px rgba(15,118,110,.18); }
    button:hover:not(:disabled) { transform:translateY(-1px); background:var(--brand-strong); }
    button.secondary { background:#314352; box-shadow:0 10px 22px rgba(49,67,82,.14); }
    button.secondary:hover:not(:disabled) { background:#253644; }
    button:disabled { opacity:.55; cursor:not-allowed; }
    .grid { display:grid; gap:12px; grid-template-columns:repeat(auto-fit, minmax(210px, 1fr)); }
    .status { background:var(--soft); border:1px solid var(--line); padding:12px; border-radius:8px; white-space:pre-wrap; font-family:ui-monospace, SFMono-Regular, Menlo, monospace; font-size:12px; overflow:auto; }
    .status.is-error { border-color:#f3b4ad; background:#fff4f2; color:var(--danger); }
    .status-rail { position:fixed; right:16px; top:16px; width:340px; max-height:calc(100vh - 32px); overflow:auto; z-index:12; background:#fff; border:1px solid var(--line); border-radius:8px; padding:14px; box-shadow:0 18px 46px rgba(31,41,51,.14); }
    .status-rail h2 { margin-bottom:10px; }
    .status-rail-output { max-height:48vh; }
    .status-rail-state { display:flex; align-items:center; gap:10px; margin-top:12px; color:var(--muted); font-size:14px; }
    .status-rail-state.is-error { color:var(--danger); }
    .status-rail-usage-wrap { margin-top:10px; }
    .status-rail-usage-wrap summary { cursor:pointer; }
    .status-rail-usage { margin-top:8px; background:#f7fafc; border:1px solid #d8e0e8; border-radius:8px; padding:10px 12px; color:#334155; font-size:13px; line-height:1.45; }
    .status-rail-usage-list { margin:0; padding-left:18px; }
    .mini-spinner { width:18px; height:18px; border-radius:999px; border:3px solid #d8e0e8; border-top-color:var(--brand); display:none; animation:spin .8s linear infinite; }
    .is-busy .mini-spinner { display:block; }
    .loading-backdrop { position:fixed; inset:0; display:none; align-items:center; justify-content:center; padding:18px; background:rgba(255,255,255,.72); backdrop-filter:blur(2px); z-index:30; }
    .loading-backdrop.open { display:flex; }
    .loading-box { width:min(360px, 100%); background:#fff; border:1px solid var(--line); border-radius:8px; padding:20px; box-shadow:0 22px 60px rgba(31,41,51,.22); text-align:center; }
    .spinner { width:42px; height:42px; margin:0 auto 14px; border-radius:999px; border:5px solid #d8e0e8; border-top-color:var(--brand); animation:spin .8s linear infinite; }
    @keyframes spin { to { transform:rotate(360deg); } }
    @media (max-width: 1259px) {
      .status-rail { left:12px; right:12px; top:auto; bottom:12px; width:auto; max-height:30vh; }
      main { padding-bottom:220px; }
    }
    .summary { background:var(--soft); border:1px solid var(--line); padding:12px; border-radius:8px; color:var(--muted); font-size:14px; line-height:1.45; margin-top:14px; }
    #imageCompareSummary strong { color:var(--fg); font-weight:600; }
    #imageCompareSummary div { margin-top:8px; }
    .session-recovery-banner { margin-top:14px; border:1px solid #86efac; background:#f0fdf4; color:#166534; border-radius:8px; padding:10px 12px; font-size:13px; line-height:1.4; }
    .session-recovery-banner.hidden { display:none; }
    .media-recovery-notice { margin-top:14px; border:1px solid #f6c081; background:#fff7ed; color:#9a3412; border-radius:8px; padding:10px 12px; font-size:13px; line-height:1.4; }
    .media-recovery-notice.hidden { display:none; }
    .acf-fields-container { background:var(--soft); border:1px solid var(--line); border-radius:8px; padding:12px; margin-top:14px; max-height:400px; overflow-y:auto; }
    .acf-field-item { padding:10px; border-left:4px solid var(--brand); background:#fff; margin-bottom:8px; border-radius:4px; }
    .acf-field-item strong { display:block; color:var(--ink); font-size:13px; margin-bottom:4px; }
    .field-label { font-weight:normal; color:var(--muted); font-size:12px; }
    .field-source { color:var(--muted); font-size:12px; margin-top:4px; }
    .guidance-text { background:var(--soft); border:1px solid #d8e0e8; border-radius:4px; padding:8px; margin-top:6px; font-size:12px; line-height:1.4; font-family:ui-monospace, SFMono-Regular, Menlo, monospace; white-space:pre-wrap; word-break:break-word; margin-bottom:0; }
    #imageCompareDescription { min-height:72px; height:72px; resize:vertical; }
    #imageOptimizePrompt { min-height:120px; height:120px; resize:vertical; }
    .links { display:grid; gap:10px; grid-template-columns:repeat(auto-fit, minmax(180px, 1fr)); margin-top:14px; }
    .links a { display:block; text-align:center; text-decoration:none; border-radius:8px; padding:12px 14px; background:#314352; color:#fff; font-weight:750; }
    .summary-actions { display:grid; gap:8px; grid-template-columns:repeat(auto-fit, minmax(150px, 1fr)); margin-top:10px; }
    .summary-actions a, .summary-actions button { display:block; text-align:center; text-decoration:none; border-radius:8px; padding:10px 12px; background:#314352; color:#fff; font-weight:750; margin:0; }
    .draft-table-wrap { border:1px solid var(--line); border-radius:8px; overflow-y:auto; overflow-x:hidden; max-height:72vh; background:#fff; }
    .draft-table { width:100%; border-collapse:collapse; table-layout:fixed; }
    .draft-table th, .draft-table td { border-bottom:1px solid #eef2f6; padding:9px; vertical-align:top; text-align:left; }
    .draft-table th { position:sticky; top:0; background:var(--soft); z-index:1; font-size:13px; }
    .draft-table td:first-child, .draft-table th:first-child { width:var(--draft-col1-width, 18ch); color:var(--muted); font-weight:650; overflow-wrap:anywhere; word-break:break-word; }
    .draft-table td:nth-child(2), .draft-table th:nth-child(2) { width:var(--draft-col2-width, 20ch); color:var(--muted); font-family:ui-monospace, SFMono-Regular, Menlo, monospace; font-size:12px; overflow-wrap:anywhere; word-break:break-word; }
    .draft-table td:nth-child(3), .draft-table th:nth-child(3) { width:auto; }
    .draft-table textarea { width:100%; min-height:86px; border:0; padding:0; border-radius:0; resize:vertical; line-height:1.4; }
    .draft-field-label { display:flex; align-items:flex-start; gap:6px; }
    .prompt-trace-button { border:1px solid var(--line); background:#fff; color:var(--brand); border-radius:999px; width:24px; height:24px; padding:0; margin:0; display:inline-flex; align-items:center; justify-content:center; font-weight:800; line-height:1; flex:0 0 auto; }
    .prompt-trace-button:hover { background:var(--soft); }
    .prompt-trace-modal .modal { width:min(860px, 100%); }
    .prompt-trace-summary { white-space:pre-wrap; font-size:13px; }
    #promptTraceText { min-height:360px; font-family:ui-monospace, SFMono-Regular, Menlo, monospace; font-size:12px; line-height:1.45; }
    .draft-value-input { height:96px; min-height:96px; max-height:96px; resize:none; overflow-y:scroll; }
    .raw-csv { display:none; }
    .chat-log { display:grid; gap:10px; max-height:320px; overflow:auto; }
    .chat-message { padding:10px 12px; border-radius:8px; border:1px solid var(--line); white-space:pre-wrap; line-height:1.45; }
    .chat-message.user { background:#eef7f5; }
    .chat-message.assistant { background:var(--soft); }
    .modal-backdrop { position:fixed; inset:0; display:none; align-items:center; justify-content:center; padding:18px; background:rgba(31,41,51,.46); z-index:20; }
    .modal-backdrop.open { display:flex; }
    #imageCompareModal { z-index:80; }
    #imageCompareModal.open { display:flex; align-items:center; justify-content:center; }
    .modal { width:min(620px, 100%); max-height:90vh; overflow:auto; background:#fff; border-radius:8px; padding:18px; box-shadow:0 22px 60px rgba(31,41,51,.28); }
    .modal h2 { margin-bottom:8px; }
    .modal-actions { display:grid; gap:10px; grid-template-columns:repeat(auto-fit, minmax(150px, 1fr)); margin-top:16px; }
    .modal-actions a, .modal-actions button { display:block; text-align:center; text-decoration:none; border-radius:8px; padding:12px 14px; background:var(--brand); color:#fff; font-weight:750; margin:0; }
    .modal-actions .secondary { background:#314352; }
    .step-actions { display:grid; gap:10px; grid-template-columns:repeat(auto-fit, minmax(190px, 1fr)); align-items:end; margin-top:14px; }
    .step-actions button { margin-top:0; }
    .image-choice { display:flex; gap:8px; align-items:center; padding:8px 0; border-bottom:1px solid #eef2f6; }
    .image-choice input { width:auto; }
    .image-choice span { overflow-wrap:anywhere; }
    .image-preview-grid { display:grid; grid-template-columns:repeat(auto-fill, minmax(120px, 1fr)); gap:10px; margin-top:12px; }
    .image-preview { border:1px solid var(--line); border-radius:8px; overflow:hidden; background:#fff; box-shadow:0 8px 18px rgba(31,41,51,.06); position:relative; }
    .image-preview button { box-shadow:none; }
    .image-preview img { width:100%; aspect-ratio:1; object-fit:cover; display:block; }
    .image-preview img.compare-trigger { cursor:pointer !important; }
    .image-preview video { width:100%; aspect-ratio:16/10; object-fit:cover; display:block; background:#111827; }
    .image-preview span { display:block; padding:7px; font-size:12px; color:var(--muted); overflow-wrap:anywhere; }
    .image-preview-label { border-top:1px solid #eef2f6; min-height:50px; max-height:50px; overflow:hidden; line-height:1.25; display:-webkit-box; -webkit-line-clamp:3; -webkit-box-orient:vertical; }
    .image-card-controls { display:grid; gap:0; border-top:1px solid #eef2f6; }
    .image-preview-featured { display:flex; align-items:center; gap:7px; padding:8px; font-size:13px; font-weight:750; color:var(--ink); cursor:pointer; margin:0; }
    .image-card-controls .image-preview-featured + .image-preview-featured { border-top:1px solid #eef2f6; }
    .image-preview-featured input { width:auto; margin:0; accent-color:var(--brand); }
    .pillow-status-btn { width:100%; margin:0; border-radius:0; border-top:1px solid #eef2f6; background:#0f766e; color:#fff; font-size:13px; font-weight:800; display:flex; align-items:center; justify-content:center; gap:8px; min-height:38px; white-space:nowrap; }
    .pillow-status-btn.loading { background:#64748b; cursor:wait; opacity:1; color:#fff; }
    .pillow-status-btn.loading span { opacity:1; color:#fff; }
    .pillow-spinner { width:14px; height:14px; border:2px solid rgba(255,255,255,.42); border-top-color:#fff; border-radius:50%; display:inline-block; animation:spin 0.8s linear infinite; }
    .image-preview.is-featured { border-color:var(--brand); box-shadow:0 10px 24px rgba(15,118,110,.16); }
    .media-remove { position:absolute; right:6px; top:6px; width:30px; height:30px; margin:0; padding:0; border-radius:999px; background:rgba(31,41,51,.82); color:#fff; font-size:18px; line-height:1; display:grid; place-items:center; }
    .media-compare { position:absolute; left:6px; bottom:6px; width:auto; min-width:84px; margin:0; padding:6px 10px; border-radius:999px; background:rgba(15,118,110,.92); color:#fff; font-size:12px; line-height:1; box-shadow:none; }
    .media-compare-hint { position:absolute; left:6px; bottom:6px; background:rgba(15,118,110,.92); color:#fff; border-radius:999px; padding:5px 8px; font-size:11px; font-weight:700; pointer-events:none; }
    .media-select-label { position:absolute; left:6px; top:6px; background:rgba(15,118,110,.92); color:#fff; border-radius:999px; padding:5px 8px; font-size:12px; font-weight:800; display:none; }
    .image-preview.is-featured .media-select-label, .image-preview.is-selected .media-select-label { display:block; }
    .recording-controls { display:grid; gap:10px; grid-template-columns:repeat(auto-fit, minmax(150px, 1fr)); }
    .icon-button { display:flex; align-items:center; justify-content:center; gap:10px; min-height:48px; }
    .icon-symbol { width:22px; height:22px; border-radius:999px; display:inline-grid; place-items:center; background:rgba(255,255,255,.18); flex:0 0 auto; position:relative; }
    .icon-record::before { content:""; width:10px; height:10px; border-radius:999px; background:#ffebe7; box-shadow:0 0 0 4px rgba(255,235,231,.18); }
    .icon-stop::before { content:""; width:10px; height:10px; border-radius:2px; background:#fff; }
    .recording-indicator { color:var(--danger); font-weight:750; margin-top:10px; display:none; align-items:center; gap:8px; }
    .recording-indicator::before { content:""; width:10px; height:10px; border-radius:999px; background:var(--danger); animation:pulse 1s ease-in-out infinite; }
    .is-recording .recording-indicator { display:flex; }
    .is-draft-chat-recording #draftChatRecordingIndicator { display:inline; }
    .recent-sessions-list { display:grid; gap:8px; margin-top:12px; }
    .recent-session-item { border:1px solid var(--line); border-radius:8px; background:#fff; padding:10px; }
    .recent-session-head { display:flex; align-items:flex-start; gap:8px; }
    .recent-session-head input { width:auto; margin:2px 0 0; }
    .recent-session-meta { color:var(--muted); font-size:12px; line-height:1.4; margin-top:5px; overflow-wrap:anywhere; }
    .recent-session-warning { margin-top:8px; border:1px solid #f6c081; background:#fff7ed; color:#9a3412; border-radius:8px; padding:8px 10px; font-size:12px; line-height:1.35; }
    .recent-session-actions { display:grid; gap:8px; grid-template-columns:repeat(auto-fit, minmax(150px, 1fr)); margin-top:8px; }
    .recent-session-actions button { margin-top:0; }
    .recent-session-select-all { display:flex; align-items:center; gap:8px; margin-top:10px; font-size:13px; color:var(--muted); }
    .recent-session-select-all input { width:auto; margin:0; }
    .image-compare-grid { display:grid; gap:12px; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); margin-top:12px; }
    .image-compare-card { border:1px solid var(--line); border-radius:8px; background:#fff; overflow:hidden; }
    .image-compare-card strong { display:block; padding:8px 10px; font-size:13px; border-bottom:1px solid #eef2f6; }
    .image-compare-card img { width:100%; display:block; max-height:48vh; object-fit:contain; background:#f8fafc; }
    .image-meta-editor { margin-top:14px; padding:14px; border:1px solid var(--line); border-radius:10px; background:#fff; }
    .image-meta-editor-grid { display:grid; gap:12px; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); }
    .image-meta-editor-grid .full-width { grid-column:1 / -1; }
    .image-meta-role { display:inline-flex; align-items:center; gap:8px; margin-top:6px; color:var(--muted); font-size:13px; }
    .image-vision-feedback { margin-top:12px; }
    .image-vision-feedback details { margin-top:8px; }
    .image-compare-loading { display:flex; align-items:center; gap:10px; margin-top:12px; }
    .inline-spinner-small { width:18px; height:18px; border:3px solid #d8e0e8; border-top-color:var(--brand); border-radius:50%; animation:spin .8s linear infinite; flex:0 0 auto; }
    @keyframes pulse { 0%, 100% { transform:scale(1); opacity:.55; } 50% { transform:scale(1.45); opacity:1; } }
    @keyframes spin { from { transform:rotate(0deg); } to { transform:rotate(360deg); } }
  </style>
</head>
<body>
<main>
  <header>
    <h1>FLAIRLAB Post Generator</h1>
    <p>Eventdaten per Sprache und Medien erfassen, KI-Entwurf prüfen und den WordPress-Beitrag erstellen.</p>
  </header>

  <details id="panelAccess" class="panel">
    <summary>Zugang</summary>
    <div class="panel-body">
    <label for="apiKey">API-Schlüssel</label>
    <input id="apiKey" type="password" autocomplete="off" placeholder="X-API-Key">
    <button class="secondary" onclick="run(saveKeyAndMaybeCreateSession)">Schlüssel speichern</button>
    </div>
  </details>

  <details id="panelKnowledge" class="panel">
    <summary>Database Datei</summary>
    <div class="panel-body">
    <label for="knowledgeWorkbook">Template-/Datenbank-Datei</label>
    <input id="knowledgeWorkbook" type="file" accept=".xlsm,.xlsx">
    <button id="uploadKnowledgeButton" class="secondary" onclick="run(uploadKnowledgeWorkbook)">Database Datei aktualisieren</button>
    <div id="knowledgeSummary" class="summary">Der Status der Database Datei erscheint hier.</div>
    <div id="knowledgeActions" class="summary-actions"></div>
    </div>
  </details>

  <details id="panelSession" class="panel">
    <summary>1. Session</summary>
    <div class="panel-body">
    <div class="grid">
      <div>
        <label for="clientId">Kunde</label>
        <input id="clientId" value="flairlab">
      </div>
      <div>
        <label for="postType">Beitragstyp</label>
        <select id="postType">
          <option>Event</option>
          <option>Location</option>
          <option>Cocktail</option>
        </select>
      </div>
      <div>
        <label for="category">Kategorie</label>
        <input id="category" value="auto event post">
      </div>
    </div>
    <div class="step-actions">
      <button onclick="run(createSession)">Weiter: Session erstellen</button>
      <button class="secondary" type="button" onclick="run(createNewPostSession)">Neue Beitrag erstellen</button>
    </div>
    <div id="sessionRecoveryBanner" class="session-recovery-banner hidden"></div>
    <div id="mediaRecoveryNotice" class="media-recovery-notice hidden"></div>
    <div id="sessionSummary" class="summary">Keine aktive Session.</div>
    </div>
  </details>

  <details id="panelArchive" class="panel">
    <summary>Sessions Archive</summary>
    <div class="panel-body">
    <div class="grid">
      <div>
        <label for="recentClientId">Client Filter</label>
        <input id="recentClientId" placeholder="z.B. flairlab">
      </div>
      <div>
        <label for="recentPostType">Post Type Filter</label>
        <input id="recentPostType" placeholder="z.B. Event">
      </div>
      <div>
        <label for="recentStatus">Status Filter</label>
        <input id="recentStatus" placeholder="z.B. wordpress_post_created">
      </div>
      <div>
        <label for="recentLimit">Limit</label>
        <input id="recentLimit" type="number" min="1" max="200" value="20">
      </div>
    </div>
    <div class="step-actions">
      <button class="secondary" onclick="run(loadRecentSessions)">Sessions laden</button>
      <button class="secondary" type="button" onclick="run(deleteSelectedRecentSessions)">Ausgewählte löschen</button>
    </div>
    <label class="recent-session-select-all"><input id="recentSessionsSelectAll" type="checkbox" onchange="toggleAllRecentSessions(this.checked)">Alle aktuell angezeigten Sessions auswählen</label>
    <div id="recentSessionsList" class="summary">Lade Sessions, um Logs anderer Runs direkt aufzurufen.</div>
    </div>
  </details>

  <details id="panelUpload" class="panel">
    <summary>2. Upload</summary>
    <div class="panel-body">
    <label for="images">Bilder</label>
    <input id="images" type="file" accept="image/*" multiple onchange="renderImageChoices()">
    <label class="recent-session-select-all" style="margin-top:8px;">
      <input id="v2UseVisionOnUpload" type="checkbox" checked>
      Metadata mit Vision nach dem ersten Entwurf generieren lassen
    </label>
    <div id="featuredChoices" style="display:none"></div>
    <div id="imagePreviews" class="image-preview-grid"></div>
    <div class="summary">Bilder werden direkt nach Auswahl automatisch hochgeladen und per Pillow optimiert. Finale Bildmetadaten werden mit dem Kontext des ersten Entwurfs erzeugt.</div>
    <label for="videos">Video</label>
    <input id="videos" type="file" accept="video/*" onchange="renderVideoChoices()">
    <div id="videoPreviews" class="image-preview-grid"></div>
    <label for="voice">Sprachnachrichten</label>
    <div class="summary">
      <strong>Empfohlene Struktur fuer Sprachnachrichten</strong><br>
      1. <strong>Kontext in 1 Satz:</strong> Was fuer ein Event war es und fuer welchen Anlass?<br>
      2. <strong>Faktenblock:</strong> Datum, Ort, Eventtyp, Gaestezahl, Dauer, gebuchte Leistung.<br>
      3. <strong>Ablauf:</strong> Kurz erklaeren, was vor Ort passiert ist (Setup, Service, besondere Momente).<br>
      4. <strong>Highlights & Wirkung:</strong> Reaktionen der Gaeste, besondere Drinks, Teamleistung, Ergebnis.<br>
      5. <strong>Herausforderung + Loesung:</strong> Was war anspruchsvoll und wie wurde es geloest?<br>
      6. <strong>Ton:</strong> Konkret, sachlich, mit echten Details statt allgemeinen Werbesaetzen.
    </div>
    <div class="recording-controls">
      <button id="startRecordingButton" class="secondary icon-button" onclick="run(startRecording)" type="button" title="Testaufnahme: Wir waren mit FLAIRLAB am 24.06.2026 beim Sommerfest von Beispielkunde im Spreespeicher Berlin im Einsatz. Es handelte sich um eine Firmenveranstaltung mit rund 180 Gästen. Gebucht waren eine hochwertige mobile Cocktailbar, Cocktailcatering und Show-Bartending. Vor Ort waren Barkeeper Max und Barkeeperin Sofia im Einsatz. Serviert wurden drei Signature Drinks: Basil Smash, Espresso Martini und ein alkoholfreier Berry Mule. Zusätzlich gab es eine alkoholfreie Auswahl für Gäste, die keinen Alkohol trinken wollten. Das besondere Highlight des Events war die Kombination aus schneller Premium-Bar, hochwertiger Präsentation und einer sichtbaren Flair-Show direkt an der Bar. Die Gäste wurden aktiv in das Bar-Erlebnis einbezogen, und viele von ihnen haben die Showeinlage gefilmt. Der Fokus lag auf kurzen Wartezeiten, einem professionellen und hochwertigen Auftritt sowie einem reibungslosen Ablauf trotz des hohen Gästeaufkommens. Gleichzeitig sollte die Bar nicht nur Getränke ausgeben, sondern als sichtbares Erlebnis und Treffpunkt der Veranstaltung funktionieren. Die größte Herausforderung war der enge Aufbau im Spreespeicher. Hinter der Bar stand nur wenig Arbeitsfläche zur Verfügung, das Zeitfenster für den Aufbau war knapp, und direkt nach der Begrüßung kamen sehr viele Gäste gleichzeitig zur Bar. Dadurch bestand das Risiko langer Wartezeiten und unübersichtlicher Arbeitswege. Die Lösung war eine klare Aufteilung in zwei Arbeitsstationen. Eine Station war hauptsächlich für die schnelle Ausgabe der vorbereiteten Signature Drinks zuständig, während die zweite Station individuellere Bestellungen und den Show-Moment übernommen hat. Das Mise en Place wurde vollständig vorbereitet, die Zutaten wurden nach Arbeitsablauf sortiert und die Laufwege hinter der Bar wurden klar festgelegt. Während der stärksten Stoßzeit wurde zusätzlich ein kurzer Flair-Moment eingesetzt, um die wartenden Gäste zu unterhalten. Die Atmosphäre war sommerlich, lebendig, hochwertig und sehr positiv. Die Gäste waren neugierig, haben viele Fragen zu den Drinks gestellt und die Flair-Einlage aufmerksam verfolgt. Besonders positiv wurde auch der alkoholfreie Berry Mule aufgenommen. Mehrere Gäste haben ausdrücklich gelobt, dass die alkoholfreie Option genauso hochwertig präsentiert wurde wie die alkoholischen Cocktails. Als zusätzliche Leistungen wurden die mobile Bar, das komplette Bar-Equipment, Gläser, Eis, Zutaten, Dekoration und die Showeinlage von FLAIRLAB gestellt. Der Ablauf war insgesamt schnell, professionell und gut organisiert. Bitte erstelle aus diesen Informationen einen sachlichen, hochwertigen und konkreten Eventbeitrag. Erfinde keine zusätzlichen Fakten. Verwende alle bestätigten Informationen passend in den vorgesehenen Feldern, insbesondere für Herausforderung, Lösung, Highlight, Atmosphäre und Fokus."><span class="icon-symbol icon-record" aria-hidden="true"></span><span>Aufnahme starten</span></button>
      <button id="stopRecordingButton" class="secondary icon-button" onclick="stopRecording()" disabled type="button"><span class="icon-symbol icon-stop" aria-hidden="true"></span><span>Stoppen und hochladen</span></button>
    </div>
    <div id="recordingIndicator" class="recording-indicator">Aufnahme läuft</div>
    <audio id="recordingPlayback" controls style="display:none;width:100%;margin-top:12px;"></audio>
    <input id="voice" type="file" accept="audio/*,video/mp4,video/webm" multiple onchange="renderVoiceChoices()">
    <div id="voiceUploads" class="summary">Noch keine Sprachnachricht gespeichert.</div>
    <div class="step-actions">
      <button id="uploadButton" onclick="run(uploadFiles)" disabled>Auswahl speichern</button>
      <button id="uploadWpMediaButton" class="secondary" type="button" onclick="run(uploadWordPressMediaLibrary)" disabled>Finale Bilder in WP Mediathek hochladen</button>
    </div>
    </div>
  </details>

  <details id="panelTranscript" class="panel">
    <summary>3. Transkript</summary>
    <div class="panel-body">
    <button id="transcribeButton" onclick="run(transcribe)" disabled>Sprache transkribieren</button>
    <label for="transcript">Bearbeitbares Transkript</label>
    <textarea id="transcript" placeholder="Das Transkript erscheint hier. Du kannst es jederzeit bearbeiten."></textarea>
    <div class="step-actions">
      <button id="saveTranscriptButton" class="secondary" onclick="run(saveTranscript)" disabled>Transkript speichern</button>
      <button id="generateFactsButton" class="secondary" onclick="run(generateFactsFromTranscript)" disabled>Fakten aus Transkript generieren</button>
      <button id="transcriptNextButton" onclick="run(generateDraft)" disabled>Weiter: Entwurf erstellen</button>
    </div>
    </div>
  </details>

  <details id="panelDraft" class="panel">
    <summary>4. Entwurf</summary>
    <div class="panel-body">
    <button id="generateDraftButton" onclick="run(generateDraft)" disabled>CSV-Entwurf erstellen</button>
    <label>Bearbeitbare Felder</label>
    <div id="draftTable" class="draft-table-wrap summary">Erstelle einen Entwurf, um die Felder hier zu prüfen.</div>
    <textarea id="draftCsv" class="raw-csv" aria-hidden="true"></textarea>
    <button id="saveDraftButton" class="secondary" onclick="run(saveDraft)" disabled>Entwurf speichern</button>
    <label for="draftChatInput">Entwurf mit dem Agenten verbessern</label>
    <div id="draftChatLog" class="chat-log summary">Erstelle zuerst einen Entwurf. Danach kannst du hier Änderungen anfragen.</div>
    <div class="recording-controls" style="margin-bottom: 12px;">
      <button id="startDraftChatRecordingButton" class="secondary icon-button" onclick="run(startDraftChatRecording)" type="button"><span class="icon-symbol icon-record" aria-hidden="true"></span><span>Sprache aufnehmen</span></button>
      <button id="stopDraftChatRecordingButton" class="secondary icon-button" onclick="stopDraftChatRecording()" disabled type="button"><span class="icon-symbol icon-stop" aria-hidden="true"></span><span>Stoppen</span></button>
      <span id="draftChatRecordingIndicator" class="recording-indicator" style="display:none;">Aufnahme läuft</span>
    </div>
    <audio id="draftChatRecordingPlayback" controls style="display:none;width:100%;margin-bottom:12px;"></audio>
    <textarea id="draftChatInput" placeholder="Beispiel: Bitte FAQ aus den Fakten bauen, CTA stärker machen und Hero-Felder spezifischer formulieren."></textarea>
    <button id="sendDraftChatButton" onclick="run(sendDraftChat)" disabled>An Agenten senden</button>
    </div>
  </details>

  <details id="panelWordPress" class="panel">
    <summary>5. WordPress</summary>
    <div class="panel-body">
    <label for="postStatus">Beitragsstatus</label>
    <select id="postStatus">
      <option value="draft">draft</option>
      <option value="publish">publish</option>
      <option value="pending">pending</option>
      <option value="private">private</option>
    </select>
    <div class="step-actions">
      <button id="createPostButton" onclick="run(createWordPressPost)" disabled>Weiter: WordPress-Beitrag erstellen</button>
      <button id="updatePostButton" class="secondary" onclick="run(updateExistingWordPressPost)" disabled>Bereits erstellten Beitrag aktualisieren</button>
    </div>
    <div id="updateTargetInfo" class="summary">Update-Ziel: noch kein zuvor erstellter Beitrag in dieser Session.</div>
    <div id="postLinks" class="links"></div>
    </div>
  </details>

  <details id="panelStatus" class="panel">
    <summary>Status</summary>
    <div class="panel-body">
    <div id="status" class="summary">Bereit.</div>
    <details class="summary">
      <summary>Entwicklerdetails</summary>
      <div id="statusDeveloper" class="status">Noch keine technischen Details.</div>
    </details>
    <details id="statusAiUsageWrap" class="summary status-rail-usage-wrap">
      <summary id="statusAiUsageSummary">AI Nutzung für Session: 0 Calls · 0 Tokens · 0.0000 USD</summary>
      <div id="statusAiUsage" class="status-rail-usage">Noch keine AI-Aufrufe in dieser Session.</div>
    </details>
    </div>
  </details>
</main>

<aside id="statusRail" class="status-rail" aria-live="polite">
  <h2>Status</h2>
  <div id="statusRailContent" class="summary">Bereit.</div>
  <details class="summary">
    <summary>Entwicklerdetails</summary>
    <div id="statusRailDeveloper" class="status status-rail-output">Noch keine technischen Details.</div>
  </details>
  <details id="statusRailAiUsageWrap" class="summary status-rail-usage-wrap">
    <summary id="statusRailAiUsageSummary">AI Nutzung pro Beitrag: 0 Calls · 0 Tokens · 0.0000 USD</summary>
    <div id="statusRailAiUsage" class="status-rail-usage">Noch keine AI-Aufrufe in dieser Session.</div>
  </details>
  <div id="statusRailState" class="status-rail-state">
    <span class="mini-spinner" aria-hidden="true"></span>
    <span id="statusRailStateText">Bereit</span>
  </div>
</aside>

<div id="loadingOverlay" class="loading-backdrop" role="status" aria-live="polite" aria-label="Läuft">
  <div class="loading-box">
    <div class="spinner" aria-hidden="true"></div>
    <strong id="loadingText">Bitte warten...</strong>
  </div>
</div>

<div id="resultModal" class="modal-backdrop" role="dialog" aria-modal="true" aria-labelledby="resultModalTitle">
  <div class="modal">
    <h2 id="resultModalTitle">Beitrag erstellt</h2>
    <div id="resultModalSummary" class="summary"></div>
    <div id="resultModalLinks" class="modal-actions"></div>
  </div>
</div>

<div id="apiKeyModal" class="modal-backdrop" role="dialog" aria-modal="true" aria-labelledby="apiKeyModalTitle">
  <div class="modal">
    <h2 id="apiKeyModalTitle">API-Schlüssel erforderlich</h2>
    <p>Gib den API-Schlüssel ein, um zu starten.</p>
    <label for="apiKeyModalInput">API-Schlüssel</label>
    <input id="apiKeyModalInput" type="password" autocomplete="off" placeholder="X-API-Key">
    <div class="modal-actions">
      <button type="button" onclick="run(saveKeyFromModal)">Weiter</button>
    </div>
  </div>
</div>

<div id="errorModal" class="modal-backdrop" role="alertdialog" aria-modal="true" aria-labelledby="errorModalTitle">
  <div class="modal">
    <h2 id="errorModalTitle">Fehler</h2>
    <div id="errorModalMessage" class="summary"></div>
    <details class="summary">
      <summary>Entwicklerdetails</summary>
      <div id="errorModalDetails" class="status">Keine technischen Details.</div>
    </details>
    <div class="modal-actions">
      <button class="secondary" type="button" onclick="closeErrorModal()">Schließen</button>
    </div>
  </div>
</div>

<div id="promptTraceModal" class="modal-backdrop prompt-trace-modal" role="dialog" aria-modal="true" aria-labelledby="promptTraceTitle">
  <div class="modal">
    <h2 id="promptTraceTitle">Prompt-Regeln</h2>
    <div id="promptTraceSummary" class="summary prompt-trace-summary"></div>
    <label for="promptTraceText">Kopierbarer Kontext</label>
    <textarea id="promptTraceText" readonly></textarea>
    <div class="modal-actions">
      <button type="button" onclick="copyPromptTrace()">Kontext kopieren</button>
      <button class="secondary" type="button" onclick="closePromptTraceModal()">Schließen</button>
    </div>
  </div>
</div>

<div id="imageCompareModal" class="modal-backdrop" role="dialog" aria-modal="true" aria-labelledby="imageCompareTitle">
  <div class="modal">
    <h2 id="imageCompareTitle">Pillow Vorher/Nachher</h2>
    <div id="imageCompareSummary" class="summary"></div>
    <div class="image-compare-grid">
      <div class="image-compare-card">
        <strong>Vorher</strong>
        <img id="imageCompareBefore" alt="Originalbild">
      </div>
      <div class="image-compare-card">
        <strong>Nachher</strong>
        <img id="imageCompareAfter" alt="Optimiertes Bild">
      </div>
    </div>
    <div class="image-meta-editor">
      <strong>WordPress-Metadaten für dieses Bild</strong>
      <div id="imageCompareRole" class="image-meta-role"></div>
      <div class="image-meta-editor-grid">
        <div>
          <label for="imageCompareAlt">Alt-Text</label>
          <input id="imageCompareAlt" type="text" placeholder="Beschreibender Alt-Text">
        </div>
        <div>
          <label for="imageCompareMetaTitle">Titel</label>
          <input id="imageCompareMetaTitle" type="text" placeholder="Titel in WordPress">
        </div>
        <div>
          <label for="imageCompareCaption">Caption</label>
          <input id="imageCompareCaption" type="text" placeholder="Kurzbeschreibung unter dem Bild">
        </div>
        <div class="full-width">
          <label for="imageCompareDescription">Beschreibung</label>
          <textarea id="imageCompareDescription" placeholder="Beschreibung für die Mediathek"></textarea>
        </div>
      </div>
      <div id="imageCompareVisionFeedback" class="summary image-vision-feedback" style="display:none;"></div>
      <div id="imageOptimizePromptWrap" class="summary" style="display:none;">
        <label for="imageOptimizePrompt">Prompt für OpenAI Bildoptimierung</label>
        <textarea id="imageOptimizePrompt" placeholder="Beispiel: Halte den Kupferbecher komplett im Frame, erhöhe die Details am Rand, reduziere Hintergrundablenkung und bewahre realistische Farben."></textarea>
        <div class="modal-actions" style="margin-top:10px;">
          <button type="button" onclick="run(sendImageOptimizationPrompt)">Prompt senden</button>
          <button class="secondary" type="button" onclick="closeImageOptimizationPrompt()">Abbrechen</button>
        </div>
      </div>
      <div id="imageCompareOptimizeLoading" class="summary image-compare-loading" style="display:none;">
        <span class="inline-spinner-small" aria-hidden="true"></span>
        <span id="imageCompareOptimizeLoadingText">OpenAI Bildoptimierung läuft...</span>
      </div>
    </div>
    <div class="modal-actions">
      <button type="button" onclick="run(applyVisionMetadataFromCompare)">AI Vision: Metadaten & Hinweise</button>
      <button class="secondary" type="button" onclick="run(openImageOptimizationPrompt)">OpenAI Bildoptimierung</button>
      <button class="secondary" type="button" onclick="run(restoreComparedImageToOriginal)">Bearbeitung verwerfen (Original)</button>
      <button type="button" onclick="run(saveImageMetadata)">Metadaten speichern</button>
      <button class="secondary" type="button" onclick="closeImageCompareModal()">Schließen</button>
    </div>
  </div>
</div>

<script>
let sessionId = sessionStorage.getItem("flairlab_session_id") || "";
let mediaRecorder = null;
let recordedVoiceBlob = null;
let recordedVoiceName = "";
let recordedChunks = [];
let imagePreviewUrls = [];
let imageCompareUrls = [];
let videoPreviewUrls = [];
let currentSessionData = null;
let currentCompareImageFilename = "";
let imageAutoUploadTimer = null;
let imageAutoUploadInFlight = false;
let imageAutoUploadQueued = false;
let pillowStatusPollTimer = null;
let lastPillowCompletedAt = "";
let pillowRecoveryTriggered = false;
let lastPillowProgressKey = "";
let recentSelectedSessionIds = new Set();
let recentSessionIdsInView = [];
document.getElementById("apiKey").value = sessionStorage.getItem("flairlab_api_key") || "";
function key(){ return document.getElementById("apiKey").value.trim(); }
function headers(json=true){ const h = {"X-API-Key": key()}; if(json) h["Content-Type"]="application/json"; return h; }
function userMessage(obj){
  if(typeof obj === "string") return obj.replace(/^Error:/, "Fehler:");
  obj = obj || {};
  const state = obj.status || obj.message || "";
  const messages = {
    created:"Session wurde erstellt.",
    files_uploaded:"Dateien wurden hochgeladen.",
    transcribed:"Transkript wurde erstellt.",
    transcript_updated:"Transkript wurde gespeichert.",
    draft_generated:"Entwurf wurde erstellt.",
    draft_updated:"Entwurf wurde gespeichert.",
    draft_revised:"Entwurf wurde aktualisiert.",
    wordpress_post_created:"WordPress-Beitrag wurde erstellt."
  };
  if(obj.success && obj.message) return obj.message;
  if(messages[state]) return messages[state];
  if(obj.session_id) return "Session ist bereit.";
  return "Aktion abgeschlossen.";
}
function status(obj){
  const userText = userMessage(obj);
  const technicalText = typeof obj === "string" ? obj : JSON.stringify(obj, null, 2);
  const isError = /^\\s*(Fehler|Error):/i.test(userText);
  document.getElementById("status").textContent = userText;
  document.getElementById("status").classList.toggle("is-error", isError);
  document.getElementById("statusRailContent").textContent = userText;
  document.getElementById("statusRailContent").classList.toggle("is-error", isError);
  document.getElementById("statusDeveloper").textContent = technicalText;
  document.getElementById("statusRailDeveloper").textContent = technicalText;
  document.getElementById("statusRailState").classList.toggle("is-error", isError);
}
function formatElapsedDuration(ms){
  const totalSeconds = Math.max(0, Math.round(Number(ms || 0) / 1000));
  const minutes = Math.floor(totalSeconds / 60);
  const seconds = totalSeconds % 60;
  if(minutes > 0) return `${minutes}m ${String(seconds).padStart(2, "0")}s`;
  return `${seconds}s`;
}
function appendElapsedDuration(text, startedAt){
  const elapsed = formatElapsedDuration(Date.now() - startedAt);
  const clean = String(text || "Aktion abgeschlossen.").replace(/\\s+\\(Dauer:.*?\\)\\s*$/, "");
  return `${clean} (Dauer: ${elapsed})`;
}
function readableError(error){
  const raw = (error && error.message) || String(error || "Unbekannter Fehler");
  try {
    const parsed = JSON.parse(raw);
    if(parsed && parsed.detail) return typeof parsed.detail === "string" ? parsed.detail : JSON.stringify(parsed.detail);
  } catch {}
  return raw.replace(/^Error:\\s*/, "").replace(/^Fehler:\\s*/, "");
}
function isMissingSessionError(error){
  const message = readableError(error).toLowerCase();
  return message.includes("session not found");
}
function showErrorModal(error){
  const message = readableError(error);
  document.getElementById("errorModalMessage").textContent = message;
  document.getElementById("errorModalDetails").textContent = (error && error.message) || String(error || "");
  document.getElementById("errorModal").classList.add("open");
}
function clearSessionRecoveryBanner(){
  const banner = document.getElementById("sessionRecoveryBanner");
  if(!banner) return;
  banner.textContent = "";
  banner.classList.add("hidden");
}
function showSessionRecoveryBanner(previousSessionId, currentSessionId){
  const banner = document.getElementById("sessionRecoveryBanner");
  if(!banner) return;
  const oldId = previousSessionId || "unbekannt";
  const newId = currentSessionId || "unbekannt";
  banner.textContent = `Session wiederhergestellt: ${oldId} -> ${newId}`;
  banner.classList.remove("hidden");
}
function clearMediaRecoveryNotice(){
  const banner = document.getElementById("mediaRecoveryNotice");
  if(!banner) return;
  banner.textContent = "";
  banner.classList.add("hidden");
}
function renderMediaRecoveryNotice(mediaRecovery){
  const banner = document.getElementById("mediaRecoveryNotice");
  if(!banner) return;
  const payload = mediaRecovery && typeof mediaRecovery === "object" ? mediaRecovery : {};
  const missingTotal = Number(payload.missing_total || 0);
  if(!missingTotal){
    clearMediaRecoveryNotice();
    return;
  }
  const parts = [];
  if(Number(payload.missing_images || 0)) parts.push(`${payload.missing_images} Bilder`);
  if(Number(payload.missing_videos || 0)) parts.push(`${payload.missing_videos} Videos`);
  if(Number(payload.missing_voices || 0)) parts.push(`${payload.missing_voices} Sprachnachrichten`);
  const suffix = parts.length ? ` (${parts.join(", ")})` : "";
  banner.textContent = `Hinweis: ${missingTotal} Medien-Datei(en) dieser Session konnten nicht wiederhergestellt werden${suffix}. Bitte erneut hochladen, falls benötigt.`;
  banner.classList.remove("hidden");
}
function closeErrorModal(){
  document.getElementById("errorModal").classList.remove("open");
}
function closeImageCompareModal(){
  currentCompareImageFilename = "";
  closeImageOptimizationPrompt();
  setImageCompareOptimizeLoading(false);
  revokeImageCompareUrls();
  const feedback = document.getElementById("imageCompareVisionFeedback");
  if(feedback){
    feedback.style.display = "none";
    feedback.innerHTML = "";
  }
  document.getElementById("imageCompareModal").classList.remove("open");
}
function openImageOptimizationPrompt(){
  if(!currentCompareImageFilename) throw new Error("Kein Bild im Vergleichsfenster ausgewählt.");
  const wrap = document.getElementById("imageOptimizePromptWrap");
  const input = document.getElementById("imageOptimizePrompt");
  if(!wrap || !input) return;
  wrap.style.display = "block";
  if(!input.value.trim()){
    input.value = "Halte das Hauptmotiv vollständig im Bild, verbessere die Schärfe am Motiv, reduziere Hintergrundablenkung und behalte natürliche Farben.";
  }
  setTimeout(() => input.focus(), 0);
}
function closeImageOptimizationPrompt(){
  const wrap = document.getElementById("imageOptimizePromptWrap");
  if(!wrap) return;
  wrap.style.display = "none";
}
function setImageCompareOptimizeLoading(isLoading, message="OpenAI Bildoptimierung läuft..."){
  const wrap = document.getElementById("imageCompareOptimizeLoading");
  const text = document.getElementById("imageCompareOptimizeLoadingText");
  if(text) text.textContent = message;
  if(wrap) wrap.style.display = isLoading ? "flex" : "none";
}
function waitMs(ms){
  return new Promise(resolve => setTimeout(resolve, ms));
}
async function fetchOriginalImageBlobWithRetry(filename){
  if(!sessionId) throw new Error("No active session.");
  let lastResponse = null;
  for(let attempt = 1; attempt <= 2; attempt += 1){
    const response = await fetch(`/app/sessions/${sessionId}/images/${encodeURIComponent(filename)}/original`, {headers:headers(false)});
    if(response.ok) return await response.blob();
    lastResponse = response;
    if(response.status !== 404) break;
    if(attempt === 1){
      try {
        await api(`/app/sessions/${sessionId}/images/process`, {method:"POST", headers:headers(false)});
      } catch (error) {
        console.warn("Pillow process retry trigger failed", error);
      }
    }
    await waitMs(250 * attempt);
  }
  throw new Error(lastResponse && lastResponse.status === 404
    ? "Before/after is still finalizing. Please click 'Edits anzeigen' again in a moment."
    : "Could not load before/after preview right now.");
}
function esc(value){
  return String(value == null ? "" : value).replace(/[&<>"']/g, char => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[char]));
}
function formatInt(value){
  return Number(value || 0).toLocaleString("de-DE");
}
function formatCost(value){
  return Number(value || 0).toFixed(4);
}
function aiUsageSummaryText(aiUsage){
  const usage = aiUsage && typeof aiUsage === "object" ? aiUsage : {};
  const calls = Number(usage.call_count || 0);
  const tokens = Number(usage.total_tokens || 0);
  const cost = Number(usage.estimated_cost_usd || 0);
  const unknownCalls = Number(usage.unknown_usage_calls || 0);
  let text = `AI Nutzung pro Beitrag: ${formatInt(calls)} Calls · ${formatInt(tokens)} Tokens · ${formatCost(cost)} USD`;
  if(unknownCalls > 0){
    text += ` · ${formatInt(unknownCalls)} Call(s) ohne Tokendaten`;
  }
  return text;
}
function serviceLabel(service){
  const map = {
    openai_text: "OpenAI Text",
    openai_vision: "OpenAI Vision",
    openai_transcription: "OpenAI Transcription",
    openai_images: "OpenAI Images"
  };
  return map[String(service || "")] || String(service || "Unbekannt");
}
function aiUsageDetailsHtml(aiUsage){
  const usage = aiUsage && typeof aiUsage === "object" ? aiUsage : {};
  const services = (usage.services && typeof usage.services === "object") ? usage.services : {};
  const rows = Object.values(services)
    .filter(item => item && typeof item === "object")
    .sort((a, b) => {
      const costDiff = Number(b.estimated_cost_usd || 0) - Number(a.estimated_cost_usd || 0);
      if(costDiff !== 0) return costDiff;
      return Number(b.call_count || 0) - Number(a.call_count || 0);
    });
  if(!rows.length) return "Noch keine AI-Aufrufe in dieser Session.";
  const items = rows.map(item => {
    const calls = formatInt(item.call_count || 0);
    const tokens = formatInt(item.total_tokens || 0);
    const cost = formatCost(item.estimated_cost_usd || 0);
    const unknown = Number(item.unknown_usage_calls || 0);
    const unknownText = unknown > 0 ? ` · ${formatInt(unknown)} ohne Tokendaten` : "";
    return `<li><strong>${esc(serviceLabel(item.service))}</strong>: ${calls} Calls · ${tokens} Tokens · ${cost} USD${unknownText}</li>`;
  }).join("");
  return `<ul class="status-rail-usage-list">${items}</ul>`;
}
function renderAiUsage(aiUsage){
  const summaryText = aiUsageSummaryText(aiUsage);
  const detailsHtml = aiUsageDetailsHtml(aiUsage);
  const summaryTargets = [
    document.getElementById("statusRailAiUsageSummary"),
    document.getElementById("statusAiUsageSummary"),
  ];
  const detailsTargets = [
    document.getElementById("statusRailAiUsage"),
    document.getElementById("statusAiUsage"),
  ];
  summaryTargets.forEach(target => {
    if(target) target.textContent = summaryText;
  });
  detailsTargets.forEach(target => {
    if(target) target.innerHTML = detailsHtml;
  });
}
function actionLabel(fn){
  return ({
    createSession:"Session wird erstellt...",
    uploadFiles:"Dateien werden hochgeladen...",
    transcribe:"Sprache wird transkribiert...",
    saveTranscript:"Transkript wird gespeichert...",
    generateFactsFromTranscript:"Fakten werden aus dem Transkript generiert...",
    generateDraft:"Entwurf wird erstellt...",
    saveDraft:"Entwurf wird gespeichert...",
    sendDraftChat:"Entwurf wird aktualisiert...",
    goToWordPressStep:"WordPress-Schritt wird geöffnet...",
    createWordPressPost:"WordPress-Beitrag wird erstellt...",
    openSessionLogs:"Logs werden geöffnet...",
    loadRecentSessions:"Session-Archiv wird geladen...",
    uploadWordPressMediaLibrary:"Bilder werden in die WP Mediathek hochgeladen...",
    transcribe:"Transkription läuft...",
    uploadKnowledgeWorkbook:"Database Datei wird aktualisiert...",
    downloadKnowledgeWorkbook:"Database Datei wird geladen...",
    startRecording:"Mikrofon wird geöffnet...",
    openImageOptimizationPrompt:"Prompt für Bildoptimierung wird geöffnet...",
    sendImageOptimizationPrompt:"OpenAI Bildoptimierung läuft...",
    saveKeyFromModal:"Startet..."
  })[(fn && fn.name) || ""] || "Bitte warten...";
}
let busyCount = 0;
function setBusy(isBusy, label="Bitte warten..."){
  busyCount = Math.max(0, busyCount + (isBusy ? 1 : -1));
  const active = busyCount > 0;
  document.body.classList.toggle("is-busy", active);
  document.getElementById("loadingOverlay").classList.toggle("open", active);
  document.getElementById("loadingText").textContent = label;
  document.getElementById("statusRailStateText").textContent = active ? label : "Bereit";
}
function openPanel(id, scroll=false){
  const panel = document.getElementById(id);
  if(!panel) return;
  panel.open = true;
  if(scroll) panel.scrollIntoView({behavior:"smooth", block:"start"});
}
function closePanel(id){
  const panel = document.getElementById(id);
  if(panel) panel.open = false;
}
function openWorkflowPanels(data){
  if(!data) return;
  if(data.session_id) openPanel("panelUpload");
  const files = data.files || {};
  const draft = data.draft || {};
  if(files.voice || (files.voices && files.voices.length)) openPanel("panelTranscript");
  if(draft.csv_text) openPanel("panelDraft");
  if(draft.csv_text) openPanel("panelWordPress");
}
function openPanelsFromCurrentInputs(){
  if(sessionId) openPanel("panelUpload");
  if(document.getElementById("draftCsv").value.trim()) openPanel("panelWordPress");
}
function uiCacheKey(){
  return sessionId ? `flairlab_ui_cache_${sessionId}` : "";
}
function collectUiCache(){
  const tableCsv = syncDraftCsvFromTable();
  return {
    client_id: document.getElementById("clientId").value,
    post_type: document.getElementById("postType").value,
    category: document.getElementById("category").value,
    post_status: document.getElementById("postStatus").value,
    transcript_text: document.getElementById("transcript").value,
    draft_csv_text: tableCsv || document.getElementById("draftCsv").value,
    draft_chat_input: document.getElementById("draftChatInput").value,
    knowledge_status: JSON.parse(sessionStorage.getItem("flairlab_knowledge_status") || "null"),
    saved_at: new Date().toISOString()
  };
}
function saveUiCacheLocal(){
  const key = uiCacheKey();
  if(!key) return;
  const cache = collectUiCache();
  if(!cache.transcript_text && !cache.draft_csv_text) return;
  sessionStorage.setItem(key, JSON.stringify(cache));
}
function applyUiCache(cache){
  if(!cache) return;
  if(cache.client_id) document.getElementById("clientId").value = cache.client_id;
  if(cache.post_type) document.getElementById("postType").value = cache.post_type;
  if(cache.category) document.getElementById("category").value = cache.category;
  if(cache.post_status) document.getElementById("postStatus").value = cache.post_status;
  if(cache.transcript_text !== undefined) document.getElementById("transcript").value = cache.transcript_text;
  if(cache.draft_csv_text !== undefined){
    document.getElementById("draftCsv").value = cache.draft_csv_text;
    renderDraftTable(cache.draft_csv_text);
  }
  if(cache.draft_chat_input !== undefined) document.getElementById("draftChatInput").value = cache.draft_chat_input;
  if(cache.knowledge_status) renderKnowledgeStatus(cache.knowledge_status);
}
let uiCacheTimer = null;
async function saveUiCache(){
  if(!sessionId) return;
  saveUiCacheLocal();
  if(!key()) return;
  const cache = collectUiCache();
  try {
    await api(`/app/sessions/${sessionId}/ui-cache`, {
      method:"PUT",
      headers:headers(),
      body:JSON.stringify({cache})
    });
  } catch (error) {
    if(!isMissingSessionError(error)) throw error;
    console.warn("UI cache save skipped because server session is temporarily unavailable.");
  }
}
function scheduleUiCacheSave(){
  saveUiCacheLocal();
  clearTimeout(uiCacheTimer);
  uiCacheTimer = setTimeout(() => {
    saveUiCache().catch(error => console.warn("UI cache save failed", error));
  }, 900);
}
function parseCsv(text){
  const rows = [];
  let row = [], field = "", quoted = false;
  for(let index = 0; index < text.length; index++){
    const char = text[index];
    const next = text[index + 1];
    if(quoted){
      if(char === '"' && next === '"'){ field += '"'; index++; }
      else if(char === '"'){ quoted = false; }
      else { field += char; }
    } else if(char === '"'){
      quoted = true;
    } else if(char === ","){
      row.push(field); field = "";
    } else if(char === "\\n"){
      row.push(field); rows.push(row); row = []; field = "";
    } else if(char !== "\\r"){
      field += char;
    }
  }
  if(field || row.length) { row.push(field); rows.push(row); }
  return rows;
}
function csvEscape(value){
  const text = String(value == null ? "" : value);
  return /[",\\n\\r]/.test(text) ? `"${text.replaceAll('"', '""')}"` : text;
}
function csvFromRows(rows){
  return rows.map(row => row.map(csvEscape).join(",")).join("\\n") + "\\n";
}
const draftFieldAliasLookup = {
  bartendertitle: "bartendertile",
  bartendertile: "bartendertitle",
};
function normalizeDraftFieldKey(field){
  return String(field || "").trim().toLowerCase().replace(/[^a-z0-9]+/g, "");
}
function draftFieldCandidates(field){
  const raw = String(field || "").trim();
  if(!raw) return [];
  const normalized = normalizeDraftFieldKey(raw);
  const alias = draftFieldAliasLookup[normalized] || "";
  return Array.from(new Set([raw, normalized, alias].filter(Boolean)));
}
function displayDraftFieldName(field){
  return String(field || "").trim();
}
function valueByDraftField(mapLike, field){
  if(!mapLike || typeof mapLike !== "object") return "";
  for(const candidate of draftFieldCandidates(field)){
    if(candidate in mapLike) return mapLike[candidate];
  }
  return "";
}
function fieldMaps(){
  const statusData = JSON.parse(sessionStorage.getItem("flairlab_knowledge_status") || "null") || {};
  const pairs = statusData.acf_guidance_list || [];
  const map = {};
  const guidanceMap = {};
  pairs.forEach(item => {
    const source = (item.user_field || "").trim();
    const acf = (item.acf_field || "").trim();
    if(source && acf){
      map[source] = acf;
      const normalizedSource = normalizeDraftFieldKey(source);
      if(normalizedSource) map[normalizedSource] = acf;
      const alias = draftFieldAliasLookup[normalizedSource] || "";
      if(alias) map[alias] = acf;
    }
    const guidance = String(item.guidance || "").trim();
    if(source && guidance){
      if(!(source in guidanceMap)) guidanceMap[source] = guidance;
      const normalizedSource = normalizeDraftFieldKey(source);
      if(normalizedSource && !(normalizedSource in guidanceMap)) guidanceMap[normalizedSource] = guidance;
      const alias = draftFieldAliasLookup[normalizedSource] || "";
      if(alias && !(alias in guidanceMap)) guidanceMap[alias] = guidance;
    }
  });
  return {acfMap: map, guidanceMap};
}
function draftGenerationTrace(){
  const draft = (currentSessionData && currentSessionData.draft) || {};
  const trace = draft.generation_trace || {};
  return trace && typeof trace === "object" ? trace : {};
}
function traceForDraftField(field){
  const trace = draftGenerationTrace();
  for(const candidate of draftFieldCandidates(field)){
    if(candidate in trace) return trace[candidate];
  }
  return null;
}
function promptTraceRuleLines(title, rules){
  if(!Array.isArray(rules) || !rules.length) return [];
  const lines = [`\\n${title}:`];
  rules.forEach(rule => {
    if(!rule || typeof rule !== "object") return;
    const shared = rule.shared ? " [shared rule]" : "";
    const parts = [
      rule.rule_id || rule.key || rule.title || rule.field_key || "Regel",
      rule.scope ? `scope=${rule.scope}` : "",
      rule.section ? `section=${rule.section}` : "",
      rule.group ? `group=${rule.group}` : "",
    ].filter(Boolean);
    lines.push(`- ${parts.join(" · ")}${shared}`);
    const text = rule.guidance || rule.instruction || rule.rule || rule.description || rule.text || "";
    if(text) lines.push(`  ${String(text)}`);
  });
  return lines;
}
function formatPromptTrace(field, trace){
  if(!trace || typeof trace !== "object") return `Keine Prompt-Regeln für ${field} gespeichert.`;
  const lines = [
    `Feld: ${field}`,
    trace.acf_field_name ? `ACF Feld: ${trace.acf_field_name}` : "",
    trace.field_role ? `Rolle: ${trace.field_role}` : "",
    trace.section ? `Sektion: ${trace.section}` : "",
    trace.group ? `Gruppe: ${trace.group}` : "",
    trace.value_type ? `Werttyp: ${trace.value_type}` : "",
    trace.required_for_output !== undefined ? `Pflichtfeld: ${trace.required_for_output ? "ja" : "nein"}` : "",
    trace.min_words ? `Min. Wörter: ${trace.min_words}` : "",
    trace.max_words ? `Max. Wörter: ${trace.max_words}` : "",
  ].filter(Boolean);
  if(trace.description_de) lines.push(`\\nBeschreibung:\\n${trace.description_de}`);
  if(trace.guidance_de) lines.push(`\\nFeld-Guidance:\\n${trace.guidance_de}`);
  if(trace.example) lines.push(`\\nBeispiel:\\n${trace.example}`);
  lines.push(...promptTraceRuleLines("SEO-Regeln", trace.seo_rules));
  lines.push(...promptTraceRuleLines("Style-Regeln", trace.style_rules));
  lines.push(...promptTraceRuleLines("Agent-Instruktionen", trace.agent_instructions));
  lines.push(...promptTraceRuleLines("Story Patterns", trace.story_patterns));
  if(Array.isArray(trace.source_fact_keys) && trace.source_fact_keys.length){
    lines.push(`\\nQuell-Fakten:\\n- ${trace.source_fact_keys.join("\\n- ")}`);
  }
  return lines.join("\\n");
}
function openPromptTrace(field){
  const trace = traceForDraftField(field);
  const modal = document.getElementById("promptTraceModal");
  const summary = document.getElementById("promptTraceSummary");
  const text = document.getElementById("promptTraceText");
  if(summary) summary.textContent = trace ? `Prompt-Kontext für "${field}"` : `Für "${field}" ist in dieser Session kein Prompt-Kontext gespeichert.`;
  if(text) text.value = formatPromptTrace(field, trace);
  if(modal) modal.classList.add("open");
}
function closePromptTraceModal(){
  document.getElementById("promptTraceModal").classList.remove("open");
}
async function copyPromptTrace(){
  const text = document.getElementById("promptTraceText");
  const value = text ? text.value : "";
  if(!value) return;
  if(navigator.clipboard && navigator.clipboard.writeText){
    await navigator.clipboard.writeText(value);
  } else {
    text.focus();
    text.select();
    document.execCommand("copy");
  }
  status("Prompt-Kontext wurde kopiert.");
}
function collectDraftGuidanceOverrides(){
  const overrides = {};
  document.querySelectorAll("[data-guidance-field]").forEach(input => {
    const key = (input.dataset.guidanceField || "").trim();
    if(key) overrides[key] = input.value || "";
  });
  return overrides;
}
function applyDraftGuidanceOverrides(overrides){
  if(!overrides || typeof overrides !== "object") return;
  Object.entries(overrides).forEach(([field, value]) => {
    const input = document.querySelector(`[data-guidance-field="${CSS.escape(field)}"]`);
    if(input) input.value = value == null ? "" : String(value);
  });
}
function renderDraftTable(csvText){
  const target = document.getElementById("draftTable");
  const rows = parseCsv(csvText || "");
  const headers = rows[0] || [];
  const values = rows[1] || [];
  const {acfMap} = fieldMaps();
  if(!headers.length){
    target.className = "draft-table-wrap summary";
    target.textContent = "Erstelle einen Entwurf, um die Felder hier zu prüfen.";
    return;
  }
  target.className = "draft-table-wrap";
  target.innerHTML = "";
  const maxUserFieldLength = headers.reduce((max, header) => Math.max(max, displayDraftFieldName(header).length), 0);
  const maxAcfLength = headers.reduce((max, header) => {
    const acfName = valueByDraftField(acfMap, header) || "-";
    return Math.max(max, String(acfName).length);
  }, 0);
  const col1Width = Math.min(34, Math.max(12, maxUserFieldLength + 2));
  const col2Width = Math.min(40, Math.max(12, maxAcfLength + 2));
  target.style.setProperty("--draft-col1-width", `${col1Width}ch`);
  target.style.setProperty("--draft-col2-width", `${col2Width}ch`);
  const technicalRows = [];
  const acfRows = [];
  headers.forEach((header, index) => {
    const acfName = valueByDraftField(acfMap, header) || "-";
    const entry = {header, index, acfName};
    if(acfName && acfName !== "-") acfRows.push(entry);
    else technicalRows.push(entry);
  });

  const buildSection = (title, sectionRows, open=true) => {
    const details = document.createElement("details");
    details.className = "summary";
    details.open = open;
    const summary = document.createElement("summary");
    summary.textContent = `${title} (${sectionRows.length})`;
    details.appendChild(summary);
    const wrap = document.createElement("div");
    wrap.className = "draft-table-wrap";
    wrap.style.marginTop = "8px";
    wrap.innerHTML = `<table class="draft-table"><thead><tr><th>User Input Field</th><th>Linked ACF Field</th><th>Value</th></tr></thead><tbody></tbody></table>`;
    const body = wrap.querySelector("tbody");
    sectionRows.forEach(({header, index, acfName}) => {
      const tr = document.createElement("tr");
      const name = document.createElement("td");
      const acf = document.createElement("td");
      const value = document.createElement("td");
      const input = document.createElement("textarea");
      const label = document.createElement("div");
      label.className = "draft-field-label";
      const labelText = document.createElement("span");
      labelText.textContent = displayDraftFieldName(header);
      label.appendChild(labelText);
      if(traceForDraftField(header)){
        const traceButton = document.createElement("button");
        traceButton.type = "button";
        traceButton.className = "prompt-trace-button";
        traceButton.title = "Prompt-Regeln für dieses Feld anzeigen";
        traceButton.setAttribute("aria-label", `Prompt-Regeln für ${header} anzeigen`);
        traceButton.textContent = "?";
        traceButton.addEventListener("click", () => openPromptTrace(header));
        label.appendChild(traceButton);
      }
      name.appendChild(label);
      acf.textContent = acfName;
      input.dataset.csvField = header;
      input.className = "draft-value-input";
      input.value = values[index] || "";
      input.addEventListener("input", () => { syncDraftCsvFromTable(); updateButtons(); scheduleUiCacheSave(); });
      value.appendChild(input);
      tr.appendChild(name);
      tr.appendChild(acf);
      tr.appendChild(value);
      body.appendChild(tr);
    });
    details.appendChild(wrap);
    return details;
  };

  target.appendChild(buildSection("Technische Felder", technicalRows, true));
  target.appendChild(buildSection("ACF Felder", acfRows, true));
}
function renderKnowledgeStatus(data){
  const target = document.getElementById("knowledgeSummary");
  const actions = document.getElementById("knowledgeActions");
  actions.innerHTML = "";
  if(!data){
    target.textContent = "Der Status der Database Datei erscheint hier.";
    return;
  }
  sessionStorage.setItem("flairlab_knowledge_status", JSON.stringify(data));
  const preview = (data.guidance_preview || [])
    .map(item => `${item.user_field_name || ""} -> ${item.acf_field_name || ""}: ${item.ai_guidance || ""}`)
    .join("\\n");

  // Build ACF fields collapsible list
  const acfList = (data.acf_guidance_list || [])
    .map(item => `
    <div class="acf-field-item">
      <strong>${esc(item.acf_field || "")}</strong> <span class="field-label">${esc(item.user_field || "")}</span>
      ${item.source_sheets && item.source_sheets.length ? `<div class="field-source">Tabs: ${esc(item.source_sheets.join(", "))}</div>` : ""}
      ${item.guidance ? `<pre class="guidance-text">${esc(item.guidance)}</pre>` : `<em>keine Guidance</em>`}
    </div>`)
    .join("");

  target.innerHTML =
    `<strong>Datei:</strong> ${esc(data.filename || "")}<br>` +
    `<strong>Pfad:</strong> ${esc(data.path || "")}<br>` +
    `${data.fallback_used ? `<strong>Konfigurierter Pfad:</strong> ${esc(data.configured_path || "")}<br><strong>Hinweis:</strong> Standarddatei nicht gefunden, verwende vorhandene Workbook-Datei aus dem Knowledge-Ordner.<br>` : ""}` +
    `${data.post_type ? `<strong>Post Type:</strong> ${esc(data.post_type)}<br>` : ""}` +
    `<strong>Vorhanden:</strong> ${data.exists ? "ja" : "nein"}<br>` +
    `<strong>Tabellen:</strong> ${esc((data.loaded_sheets || []).join(", ") || data.configured_sheet || "")}<br>` +
    `<strong>Guidance-Zeilen:</strong> ${esc(data.guidance_items == null ? 0 : data.guidance_items)}` +
    `${data.error ? `<br><strong>Fehler:</strong> ${esc(data.error)}` : ""}` +
    `${acfList ? `<details class="summary"><summary>ACF Felder & Guidance (${data.acf_guidance_list ? data.acf_guidance_list.length : 0})</summary><div class="acf-fields-container">${acfList}</div></details>` : ""}` +
    `${preview ? `<details class="summary"><summary>Geladene Guidance prüfen</summary><pre class="status">${esc(preview)}</pre></details>` : ""}`;
  if(data.exists){
    const downloadUrl = data.download_url || "/app/knowledge/workbook";
    const downloadName = data.download_filename || data.filename || "database-datei.xlsm";
    const apiKeyValue = key();
    if(apiKeyValue){
      const href = `${downloadUrl}?api_key=${encodeURIComponent(apiKeyValue)}`;
      actions.innerHTML =
        `<a class="secondary" href="${href}" download="${esc(downloadName)}">Database Datei herunterladen</a>` +
        `<button class="secondary" type="button" onclick="run(downloadKnowledgeWorkbook)">Als Datei speichern</button>`;
    } else {
      actions.innerHTML =
        `<button class="secondary" type="button" onclick="run(downloadKnowledgeWorkbook)">Database Datei herunterladen</button>`;
    }
  }
}
async function loadKnowledgeStatus(){
  const postTypeElement = document.getElementById("postType");
  const postType = (postTypeElement && postTypeElement.value) || "";
  const cached = JSON.parse(sessionStorage.getItem("flairlab_knowledge_status") || "null");
  if(cached && (cached.post_type || "") === postType) renderKnowledgeStatus(cached);
  if(!key()) return;
  const suffix = postType ? `?post_type=${encodeURIComponent(postType)}` : "";
  const data = await api(`/app/knowledge/status${suffix}`, {headers:headers(false)});
  renderKnowledgeStatus(data);
}
async function uploadKnowledgeWorkbook(){
  const file = document.getElementById("knowledgeWorkbook").files[0];
  if(!file) throw new Error("Bitte zuerst eine Database Datei auswählen.");
  const postTypeElement = document.getElementById("postType");
  const postType = (postTypeElement && postTypeElement.value) || "";
  const form = new FormData();
  form.append("workbook", file);
  if(postType) form.append("post_type", postType);
  const data = await api("/app/knowledge/workbook", {method:"POST", headers:{"X-API-Key":key()}, body:form});
  sessionStorage.removeItem("flairlab_knowledge_status");
  renderKnowledgeStatus(data);
  await loadKnowledgeStatus();
  status(data);
}
async function downloadKnowledgeWorkbook(){
  if(!key()) throw new Error("Bitte zuerst API-Schlüssel eingeben.");
  const response = await fetch("/app/knowledge/workbook", {headers:headers(false)});
  if(!response.ok){
    const text = await response.text();
    throw new Error(text || "Database Datei konnte nicht geladen werden.");
  }
  const blob = await response.blob();
  const disposition = response.headers.get("Content-Disposition") || "";
  const match = disposition.match(/filename="?([^"]+)"?/i);
  const filename = (match && match[1]) || "database-datei.xlsm";
  const url = URL.createObjectURL(blob);
  const link = document.createElement("a");
  link.href = url;
  link.download = filename;
  document.body.appendChild(link);
  link.click();
  link.remove();
  setTimeout(() => URL.revokeObjectURL(url), 1500);
  status("Database Datei wurde heruntergeladen.");
}
function revokeImagePreviewUrls(){
  imagePreviewUrls.forEach(url => setTimeout(() => URL.revokeObjectURL(url), 1500));
  imagePreviewUrls = [];
}
function revokeImageCompareUrls(){
  imageCompareUrls.forEach(url => setTimeout(() => URL.revokeObjectURL(url), 1500));
  imageCompareUrls = [];
}
function trackImageCompareUrl(url){
  if(url) imageCompareUrls.push(url);
  return url;
}
function revokeVideoPreviewUrls(){
  videoPreviewUrls.forEach(url => setTimeout(() => URL.revokeObjectURL(url), 1500));
  videoPreviewUrls = [];
}
function syncVideoCardState(){
  document.querySelectorAll("#videoPreviews .image-preview").forEach((card, index) => {
    card.classList.toggle("is-selected", index === 0);
  });
}
function syncFeaturedCardState(){
  document.querySelectorAll(".image-preview").forEach(card => {
    const input = card.querySelector("input[name='featured']");
    card.classList.toggle("is-featured", !!(input && input.checked));
  });
}
function removeLocalFile(inputId, filename){
  const input = document.getElementById(inputId);
  if(typeof DataTransfer === "undefined"){
    input.value = "";
    return;
  }
  const transfer = new DataTransfer();
  [...input.files].forEach(file => {
    if(file.name !== filename) transfer.items.add(file);
  });
  input.files = transfer.files;
}
async function removeUploadedMedia(kind, filename){
  if(!sessionId) return;
  const data = await api(`/app/sessions/${sessionId}/media/${kind}/${encodeURIComponent(filename)}`, {
    method:"DELETE",
    headers:headers(false)
  });
  await renderFreshSession(data);
  status(`${kind === "images" ? "Bild" : kind === "videos" ? "Video" : "Sprachnachricht"} wurde entfernt.`);
}
async function setFeaturedImage(filename){
  if(!sessionId || !filename) return;
  const data = await api(`/app/sessions/${sessionId}/featured-image`, {
    method:"PUT",
    headers:headers(),
    body:JSON.stringify({filename})
  });
  await renderFreshSession(data);
  status("Titelbild wurde gespeichert.");
}
function renderMediaPreviewItems(targetId, items, options={}){
  const wrap = document.getElementById(targetId);
  wrap.innerHTML = "";
  items.forEach((item, index) => {
    const card = document.createElement("div");
    card.className = "image-preview";
    card.dataset.mediaValue = String(item.value || item.name || "");
    card.dataset.mediaKind = String(options.kind || "");
    const controls = document.createElement("div");
    controls.className = "image-card-controls";
    let hasControls = false;
    if(item.mediaType === "video"){
      const video = document.createElement("video");
      video.src = item.url;
      video.controls = true;
      video.muted = true;
      card.appendChild(video);
    } else {
      const image = document.createElement("img");
      image.src = item.url;
      image.alt = item.name || "Hochgeladenes Bild";
      card.appendChild(image);
    }
    const selected = document.createElement("div");
    selected.className = "media-select-label";
    selected.textContent = item.mediaType === "video" ? "Ausgewählt" : "Titelbild";
    card.appendChild(selected);
    const label = document.createElement("span");
    label.className = "image-preview-label";
    label.textContent = item.name || "";
    label.title = item.name || "";
    card.appendChild(label);
    if(options.selectable){
      const featured = document.createElement("label");
      featured.className = "image-preview-featured";
      const input = document.createElement("input");
      input.type = "radio";
      input.name = "featured";
      input.value = item.value || item.name || "";
      input.dataset.mediaId = item.media_id || "";
      input.checked = options.selectedMediaId
        ? input.dataset.mediaId === options.selectedMediaId
        : (options.selectedValue ? input.value === options.selectedValue : index === 0);
      input.addEventListener("change", () => {
        syncFeaturedCardState();
        scheduleUiCacheSave();
        if(item.persisted) setFeaturedImage(input.value).catch(error => {
          showErrorModal(error);
          status("Fehler: " + readableError(error));
        });
      });
      featured.appendChild(input);
      featured.appendChild(document.createTextNode("Titelbild"));
      featured.title = "Set this image as featured";
      controls.appendChild(featured);
      hasControls = true;
    }
    if(options.compareable && item.persisted && item.mediaType !== "video"){
      const pillowButton = document.createElement("button");
      pillowButton.type = "button";
      pillowButton.className = "pillow-status-btn";
      const ready = !!item.hasOriginal || !!item.isProcessed;
      pillowButton.title = ready ? "Open before/after preview" : "Pillow is still processing";
      if(ready){
        pillowButton.textContent = "Edits anzeigen";
      } else {
        pillowButton.classList.add("loading");
        pillowButton.disabled = true;
        const spinner = document.createElement("span");
        spinner.className = "pillow-spinner";
        spinner.setAttribute("aria-hidden", "true");
        const text = document.createElement("span");
        text.textContent = "Wird optimiert";
        pillowButton.appendChild(spinner);
        pillowButton.appendChild(text);
      }
      pillowButton.addEventListener("click", async event => {
        event.stopPropagation();
        if(pillowButton.disabled) return;
        if(options.onCompare){
          try {
            await options.onCompare(item.value, item.url);
          } catch (error) {
            showErrorModal(error);
            status("Fehler: " + readableError(error));
          }
        }
      });
      controls.appendChild(pillowButton);
      hasControls = true;
    }
    if(hasControls) card.appendChild(controls);
    if(options.selected) card.classList.toggle("is-selected", index === 0);
    if(options.removable){
      const removeButton = document.createElement("button");
      removeButton.type = "button";
      removeButton.className = "media-remove";
      removeButton.setAttribute("aria-label", `${item.name || "Datei"} entfernen`);
      removeButton.textContent = "x";
      removeButton.addEventListener("click", event => {
        event.stopPropagation();
        if(item.persisted){
          removeUploadedMedia(options.kind, item.value).catch(error => {
            showErrorModal(error);
            status("Fehler: " + readableError(error));
          });
          return;
        }
        if(options.inputId) removeLocalFile(options.inputId, item.name);
        if(options.kind === "videos") renderVideoChoices();
        else renderImageChoices();
      });
      card.appendChild(removeButton);
    }
    wrap.appendChild(card);
  });
  syncFeaturedCardState();
  syncVideoCardState();
}
function renderImagePreviewItems(items, options={}){
  renderMediaPreviewItems("imagePreviews", items, options);
}

function refreshImagePreviewProcessingState(){
  const images = (((currentSessionData || {}).files || {}).images || []);
  const imageByName = new Map();
  images.forEach(item => {
    if(item && item.filename) imageByName.set(String(item.filename), item);
  });

  document.querySelectorAll("#imagePreviews .image-preview").forEach(card => {
    if((card.dataset.mediaKind || "") !== "images") return;
    const filename = String(card.dataset.mediaValue || "");
    if(!filename) return;
    const imageItem = imageByName.get(filename);
    if(!imageItem) return;
    const compareButton = card.querySelector(".pillow-status-btn");
    if(!compareButton) return;
    const ready = !!imageItem.original_path || !!imageItem.processed_at;
    compareButton.title = ready ? "Open before/after preview" : "Pillow is still processing";
    if(ready){
      compareButton.disabled = false;
      compareButton.classList.remove("loading");
      compareButton.textContent = "Edits anzeigen";
      return;
    }
    compareButton.disabled = true;
    compareButton.classList.add("loading");
    compareButton.innerHTML = '<span class="pillow-spinner" aria-hidden="true"></span><span>Wird optimiert</span>';
  });
}

function findSessionImageItem(filename){
  const images = (((currentSessionData || {}).files || {}).images || []);
  return images.find(img => img && (img.filename === filename || img.original_filename === filename)) || null;
}

function currentCompareImageMetadata(){
  const imageItem = findSessionImageItem(currentCompareImageFilename);
  return (imageItem && imageItem.wp_metadata) || {};
}

function populateImageCompareMetadataForm(filename){
  const imageItem = findSessionImageItem(filename);
  const metadata = (imageItem && imageItem.wp_metadata) || {};
  const featuredFilename = ((((currentSessionData || {}).files || {}).featured_image_filename) || "");
  const images = (((currentSessionData || {}).files || {}).images || []);
  const galleryImages = images.filter(img => img && img.filename !== featuredFilename);
  const galleryIndex = Math.max(0, galleryImages.findIndex(img => img && img.filename === filename)) + 1;
  const role = filename === featuredFilename ? "Featured Image" : `Galeriebild ${galleryIndex || 1}`;
  document.getElementById("imageCompareRole").textContent = role;
  document.getElementById("imageCompareAlt").value = metadata.alt_text || "";
  document.getElementById("imageCompareMetaTitle").value = metadata.title || "";
  document.getElementById("imageCompareCaption").value = metadata.caption || "";
  document.getElementById("imageCompareDescription").value = metadata.description || "";
  renderImageCompareVisionFeedback(filename);
}

function renderImageCompareVisionFeedback(filename){
  const target = document.getElementById("imageCompareVisionFeedback");
  if(!target) return;
  const imageItem = findSessionImageItem(filename);
  const analysis = (imageItem && imageItem.vision_analysis) || null;
  if(!analysis){
    target.style.display = "none";
    target.innerHTML = "";
    return;
  }
  const issues = Array.isArray(analysis.issues) ? analysis.issues : [];
  const advice = Array.isArray(analysis.advice) ? analysis.advice : [];
  const cropFocus = analysis.crop_focus && typeof analysis.crop_focus === "object" ? analysis.crop_focus : null;
  let html = `<strong>AI Vision Hinweise</strong>`;
  html += `<div>Modell: ${(analysis.model || "gpt-4o-mini")}</div>`;
  if(cropFocus && typeof cropFocus.x === "number" && typeof cropFocus.y === "number"){
    html += `<div>Pillow-Fokus: x=${cropFocus.x.toFixed(2)}, y=${cropFocus.y.toFixed(2)}</div>`;
    html += `<div>Der Fokus wird beim nächsten Pillow-Lauf als Crop-Hinweis verwendet.</div>`;
  }
  if(issues.length){
    html += `<details><summary>Erkannte Probleme (${issues.length})</summary><div style="margin-left:1em;">`;
    issues.forEach(item => { html += `<div>• ${item}</div>`; });
    html += `</div></details>`;
  } else {
    html += `<div>Keine sichtbaren Probleme erkannt.</div>`;
  }
  if(advice.length){
    html += `<details><summary>Verbesserungshinweise (${advice.length})</summary><div style="margin-left:1em;">`;
    advice.forEach(item => { html += `<div>• ${item}</div>`; });
    html += `</div></details>`;
  }
  target.innerHTML = html;
  target.style.display = "block";
}

async function persistImageMetadata(useSuggestions=false){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  if(!currentCompareImageFilename) throw new Error("Kein Bild im Vergleichsfenster ausgewählt.");
  const payload = useSuggestions ? {
    use_suggestions:true
  } : {
    alt_text: document.getElementById("imageCompareAlt").value.trim(),
    title: document.getElementById("imageCompareMetaTitle").value.trim(),
    caption: document.getElementById("imageCompareCaption").value.trim(),
    description: document.getElementById("imageCompareDescription").value.trim(),
    use_suggestions:false,
  };
  const data = await api(`/app/sessions/${sessionId}/images/${encodeURIComponent(currentCompareImageFilename)}/metadata`, {
    method:"PUT",
    headers:headers(),
    body:JSON.stringify(payload)
  });
  currentSessionData = mergeSessionData(currentSessionData, data);
  const draft = (currentSessionData || {}).draft || {};
  if(Object.prototype.hasOwnProperty.call(draft, "csv_text")) {
    document.getElementById("draftCsv").value = draft.csv_text || "";
    renderDraftTable(draft.csv_text || "");
  }
  populateImageCompareMetadataForm(currentCompareImageFilename);
  updateButtons();
  scheduleUiCacheSave();
  status(useSuggestions ? "Bildmetadaten ohne Vision vorgeschlagen." : "Bildmetadaten gespeichert.");
}

async function saveImageMetadata(){
  await persistImageMetadata(false);
}

async function autofillImageMetadataWithoutVision(){
  await persistImageMetadata(true);
}

async function sendImageOptimizationPrompt(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  if(!currentCompareImageFilename) throw new Error("Kein Bild im Vergleichsfenster ausgewählt.");
  const promptInput = document.getElementById("imageOptimizePrompt");
  const prompt = (promptInput && promptInput.value ? promptInput.value : "").trim();
  if(!prompt) throw new Error("Bitte einen Prompt für die Bildoptimierung eingeben.");

  setImageCompareOptimizeLoading(true, "OpenAI Bildoptimierung läuft...");
  status("OpenAI Bildoptimierung läuft...");
  try {
    const data = await api(
      `/app/sessions/${sessionId}/images/${encodeURIComponent(currentCompareImageFilename)}/optimize`,
      {method:"POST", headers:headers(), body:JSON.stringify({prompt})}
    );
    currentSessionData = mergeSessionData(currentSessionData, data);
    await renderFreshSession(currentSessionData);
    renderAiUsage((currentSessionData && currentSessionData.ai_usage) || data.ai_usage || {});

    const refreshedImage = await fetch(
      `/app/sessions/${sessionId}/images/${encodeURIComponent(currentCompareImageFilename)}`,
      {headers:headers(false)}
    );
    if(refreshedImage.ok){
      const blob = await refreshedImage.blob();
      const afterUrl = URL.createObjectURL(blob);
      const after = document.getElementById("imageCompareAfter");
      if(after) after.src = trackImageCompareUrl(afterUrl);
    }

    populateImageCompareMetadataForm(currentCompareImageFilename);
    closeImageOptimizationPrompt();
    updateButtons();
    scheduleUiCacheSave();
    status("OpenAI Bildoptimierung abgeschlossen. Ergebnis ersetzt das Pillow-Bild.");
  } finally {
    setImageCompareOptimizeLoading(false);
  }
}

async function restoreComparedImageToOriginal(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  if(!currentCompareImageFilename) throw new Error("Kein Bild im Vergleichsfenster ausgewählt.");

  setImageCompareOptimizeLoading(true, "Originalbild wird wiederhergestellt...");
  status("Originalbild wird wiederhergestellt...");
  try {
    const data = await api(
      `/app/sessions/${sessionId}/images/${encodeURIComponent(currentCompareImageFilename)}/restore-original`,
      {method:"POST", headers:headers(false)}
    );
    currentSessionData = mergeSessionData(currentSessionData, data);
    await renderFreshSession(currentSessionData);

    const refreshedImage = await fetch(
      `/app/sessions/${sessionId}/images/${encodeURIComponent(currentCompareImageFilename)}`,
      {headers:headers(false)}
    );
    if(refreshedImage.ok){
      const blob = await refreshedImage.blob();
      const afterUrl = URL.createObjectURL(blob);
      const after = document.getElementById("imageCompareAfter");
      if(after) after.src = trackImageCompareUrl(afterUrl);
    }

    populateImageCompareMetadataForm(currentCompareImageFilename);
    updateButtons();
    scheduleUiCacheSave();
    status("Bearbeitetes Bild wurde verworfen. Das Original ist wiederhergestellt.");
  } finally {
    setImageCompareOptimizeLoading(false);
  }
}

async function applyVisionMetadataFromCompare(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  if(!currentCompareImageFilename) throw new Error("Kein Bild im Vergleichsfenster ausgewählt.");
  setImageCompareOptimizeLoading(true, "AI Vision analysiert Bild...");
  status("AI Vision analysiert Bild und erzeugt Metadaten...");
  try {
    const data = await api(
      `/app/sessions/${sessionId}/images/${encodeURIComponent(currentCompareImageFilename)}/metadata/vision`,
      {method:"POST", headers:headers(false)}
    );
    currentSessionData = mergeSessionData(currentSessionData, data);
    const draft = (currentSessionData || {}).draft || {};
    if(Object.prototype.hasOwnProperty.call(draft, "csv_text")) {
      document.getElementById("draftCsv").value = draft.csv_text || "";
      renderDraftTable(draft.csv_text || "");
    }
    renderAiUsage((currentSessionData && currentSessionData.ai_usage) || data.ai_usage || {});
    populateImageCompareMetadataForm(currentCompareImageFilename);
    const compareItem = findSessionImageItem(currentCompareImageFilename);
    const compareAnalysis = (compareItem && compareItem.vision_analysis && typeof compareItem.vision_analysis === "object") ? compareItem.vision_analysis : {};
    const hasCropFocus = !!(compareAnalysis.crop_focus && typeof compareAnalysis.crop_focus === "object");
    const hasCropIssue = Array.isArray(compareAnalysis.issues)
      ? compareAnalysis.issues.some(issue => /crop|ausschnitt|zuschnitt|abgeschnitten|angeschnitten/i.test(String(issue || "")))
      : false;
    if(hasCropFocus || hasCropIssue){
      status("Pillow wird mit Vision-Fokus neu berechnet...");
      await api(`/app/sessions/${sessionId}/images/process`, {method:"POST", headers:headers(false)});
      const refreshed = await api(`/app/sessions/${sessionId}`, {headers:headers(false)});
      currentSessionData = refreshed;
      await renderFreshSession(refreshed);
      populateImageCompareMetadataForm(currentCompareImageFilename);
    }
    updateButtons();
    scheduleUiCacheSave();
    status("AI Vision Metadaten und Hinweise wurden übernommen.");
  } finally {
    setImageCompareOptimizeLoading(false);
  }
}

async function renderUploadedImagePreviews(images, featuredValue="", featuredMediaId=""){
  document.getElementById("imagePreviews").innerHTML = "";
  if(!sessionId || !images || !images.length || !key()) return;
  revokeImagePreviewUrls();
  const items = [];
  for(const image of images){
    try {
      const response = await fetch(`/app/sessions/${sessionId}/images/${encodeURIComponent(image.filename)}`, {headers:headers(false)});
      if(!response.ok) continue;
      const blob = await response.blob();
      const url = URL.createObjectURL(blob);
      imagePreviewUrls.push(url);
      items.push({
        url,
        name:image.original_filename || image.filename,
        value:image.filename,
        media_id:image.media_id || "",
        persisted:true,
        hasOriginal: !!image.original_path,
        isProcessed: !!image.processed_at,
      });
    } catch (error) {
      console.warn("Image preview failed", error);
    }
  }
  if(items.length) renderImagePreviewItems(items, {
    selectable:true,
    removable:true,
    kind:"images",
    selectedValue:featuredValue,
    selectedMediaId:featuredMediaId,
    compareable:true,
    onCompare: async (filename, processedUrl) => {
      if(!sessionId) return;
      const beforeBlob = await fetchOriginalImageBlobWithRetry(filename);
      const beforeUrl = URL.createObjectURL(beforeBlob);
      const afterResponse = await fetch(`/app/sessions/${sessionId}/images/${encodeURIComponent(filename)}`, {headers:headers(false)});
      if(!afterResponse.ok) throw new Error("Bearbeitetes Bild konnte nicht geladen werden.");
      const afterBlob = await afterResponse.blob();
      const afterUrl = URL.createObjectURL(afterBlob);

      // Find applied operations from session data
      const images = (((currentSessionData || {}).files || {}).images || []);
      const imageItem = images.find(img => img.filename === filename);
      const appliedOps = (imageItem || {}).applied_operations || [];

      openImageCompareModal(filename, beforeUrl, afterUrl, appliedOps, imageItem);
    }
  });
}
function openImageCompareModal(filename, beforeUrl, afterUrl, appliedOperations=[], imageItem=null){
  currentCompareImageFilename = filename;
  closeImageOptimizationPrompt();
  setImageCompareOptimizeLoading(false);
  revokeImageCompareUrls();
  const summary = document.getElementById("imageCompareSummary");
  let html = `<strong>Datei:</strong> ${filename}`;

  if(appliedOperations && appliedOperations.length > 0) {
    html += `<br><details><summary><strong>Angewendete Operationen</strong></summary><div style="margin-left: 1em; font-size: 0.9em; color: #666;">`;
    appliedOperations.forEach(op => {
      html += `<div>• ${op}</div>`;
    });
    html += `</div></details>`;
  }

  summary.innerHTML = html;
  const before = document.getElementById("imageCompareBefore");
  const after = document.getElementById("imageCompareAfter");
  const modal = document.getElementById("imageCompareModal");
  before.src = trackImageCompareUrl(beforeUrl);
  after.src = trackImageCompareUrl(afterUrl);
  populateImageCompareMetadataForm(filename);
  modal.style.zIndex = "80";
  modal.classList.add("open");
}
async function renderUploadedVideoPreviews(videos){
  document.getElementById("videoPreviews").innerHTML = "";
  if(!sessionId || !videos || !videos.length || !key()) return;
  revokeVideoPreviewUrls();
  const items = [];
  for(const video of videos){
    try {
      const response = await fetch(`/app/sessions/${sessionId}/videos/${encodeURIComponent(video.filename)}`, {headers:headers(false)});
      if(!response.ok) continue;
      const blob = await response.blob();
      const url = URL.createObjectURL(blob);
      videoPreviewUrls.push(url);
      items.push({url, name:video.original_filename || video.filename, value:video.filename, persisted:true, mediaType:"video"});
    } catch (error) {
      console.warn("Video preview failed", error);
    }
  }
  if(items.length) renderMediaPreviewItems("videoPreviews", items, {removable:true, selected:true, kind:"videos"});
}
function renderVoiceList(uploadedVoices=[], localVoices=[]){
  const target = document.getElementById("voiceUploads");
  target.innerHTML = "";
  const addRow = (name, onRemove) => {
    const row = document.createElement("div");
    row.className = "image-choice";
    const label = document.createElement("span");
    label.textContent = name;
    const button = document.createElement("button");
    button.className = "secondary";
    button.type = "button";
    button.textContent = "Entfernen";
    button.addEventListener("click", onRemove);
    row.appendChild(label);
    row.appendChild(button);
    target.appendChild(row);
  };
  uploadedVoices.forEach(voice => {
    const name = voice.original_filename || voice.filename;
    addRow(name, () => run(() => removeUploadedMedia("voices", voice.filename)));
  });
  localVoices.forEach(voice => {
    addRow(voice.name, () => removeLocalVoice(voice.name));
  });
  if(!uploadedVoices.length && !localVoices.length) target.textContent = "Noch keine Sprachnachricht gespeichert.";
}
function renderVoiceChoices(){
  const files = [...document.getElementById("voice").files];
  const sessionFiles = (currentSessionData && currentSessionData.files) || {};
  const uploaded = sessionFiles.voices || (sessionFiles.voice ? [sessionFiles.voice] : []);
  renderVoiceList(uploaded, files);
  updateButtons();
}
function removeLocalVoice(filename){
  removeLocalFile("voice", filename);
  renderVoiceChoices();
}
async function startRecording(){
  if(!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia || typeof MediaRecorder === "undefined"){
    throw new Error("Dieser Browser unterstützt die Aufnahme in der App nicht.");
  }
  const stream = await navigator.mediaDevices.getUserMedia({audio:true});
  recordedChunks = [];
  mediaRecorder = new MediaRecorder(stream);
  mediaRecorder.addEventListener("dataavailable", event => {
    if(event.data.size) recordedChunks.push(event.data);
  });
  mediaRecorder.addEventListener("stop", () => {
    stream.getTracks().forEach(track => track.stop());
    recordedVoiceBlob = new Blob(recordedChunks, {type:mediaRecorder.mimeType || "audio/webm"});
    recordedVoiceName = `recording-${new Date().toISOString().replace(/[:.]/g, "-")}.webm`;
    const playback = document.getElementById("recordingPlayback");
    playback.src = URL.createObjectURL(recordedVoiceBlob);
    playback.style.display = "block";
    document.body.classList.remove("is-recording");
    document.getElementById("startRecordingButton").disabled = false;
    document.getElementById("stopRecordingButton").disabled = true;
    status(`Aufnahme bereit: ${recordedVoiceName}`);
    uploadRecordedVoiceAndRetranscribe().catch(error => {
      console.error(error);
      showErrorModal(error);
      status("Fehler: " + readableError(error));
      setBusy(false);
    });
  });
  mediaRecorder.start();
  document.body.classList.add("is-recording");
  document.getElementById("startRecordingButton").disabled = true;
  document.getElementById("stopRecordingButton").disabled = false;
}
function stopRecording(){
  if(mediaRecorder && mediaRecorder.state !== "inactive"){
    status("Aufnahme wird gespeichert und transkribiert...");
    mediaRecorder.stop();
  }
}
async function uploadRecordedVoiceAndRetranscribe(){
  if(!recordedVoiceBlob) return;
  if(!key()) {
    showApiKeyModal();
    throw new Error("Vor dem Upload der Aufnahme ist ein API-Schlüssel erforderlich.");
  }
  if(!sessionId) await createSession();
  const form = new FormData();
  form.append("voices", recordedVoiceBlob, recordedVoiceName || "recording.webm");
  const uploadData = await api(`/app/sessions/${sessionId}/uploads`, {method:"POST", headers:{"X-API-Key":key()}, body:form});
  recordedVoiceBlob = null;
  await renderFreshSession(uploadData);
  openPanel("panelUpload");
  status("Aufnahme hochgeladen. Alle Sprachnachrichten werden neu transkribiert...");
  const transcriptData = await api(`/app/sessions/${sessionId}/transcribe`, {method:"POST", headers:headers(false)});
  document.getElementById("transcript").value = (transcriptData.transcript && transcriptData.transcript.text) || "";
  await renderFreshSession(transcriptData);
  openPanel("panelUpload");
  status(transcriptData);
  setBusy(false);
}
let draftChatMediaRecorder = null;
let draftChatRecordedChunks = [];
let draftChatRecordedVoiceBlob = null;
let draftChatRecordedVoiceName = null;
async function startDraftChatRecording(){
  if(!navigator.mediaDevices || !navigator.mediaDevices.getUserMedia || typeof MediaRecorder === "undefined"){
    throw new Error("Dieser Browser unterstützt die Aufnahme in der App nicht.");
  }
  const stream = await navigator.mediaDevices.getUserMedia({audio:true});
  draftChatRecordedChunks = [];
  draftChatMediaRecorder = new MediaRecorder(stream);
  draftChatMediaRecorder.addEventListener("dataavailable", event => {
    if(event.data.size) draftChatRecordedChunks.push(event.data);
  });
  draftChatMediaRecorder.addEventListener("stop", () => {
    stream.getTracks().forEach(track => track.stop());
    draftChatRecordedVoiceBlob = new Blob(draftChatRecordedChunks, {type:draftChatMediaRecorder.mimeType || "audio/webm"});
    draftChatRecordedVoiceName = `draft-chat-recording-${new Date().toISOString().replace(/[:.]/g, "-")}.webm`;
    const playback = document.getElementById("draftChatRecordingPlayback");
    playback.src = URL.createObjectURL(draftChatRecordedVoiceBlob);
    playback.style.display = "block";
    document.body.classList.remove("is-draft-chat-recording");
    document.getElementById("startDraftChatRecordingButton").disabled = false;
    document.getElementById("stopDraftChatRecordingButton").disabled = true;
    document.getElementById("draftChatRecordingIndicator").style.display = "none";
    status(`Aufnahme bereit: ${draftChatRecordedVoiceName}`);
    transcribeDraftChatVoice().catch(error => {
      console.error(error);
      showErrorModal(error);
      status("Fehler: " + readableError(error));
      setBusy(false);
    });
  });
  draftChatMediaRecorder.start();
  document.body.classList.add("is-draft-chat-recording");
  document.getElementById("startDraftChatRecordingButton").disabled = true;
  document.getElementById("stopDraftChatRecordingButton").disabled = false;
  document.getElementById("draftChatRecordingIndicator").style.display = "inline";
  status("Aufnahme läuft...");
}
function stopDraftChatRecording(){
  if(draftChatMediaRecorder && draftChatMediaRecorder.state !== "inactive"){
    status("Aufnahme wird gespeichert und transkribiert...");
    draftChatMediaRecorder.stop();
  }
}
async function transcribeDraftChatVoice(){
  if(!draftChatRecordedVoiceBlob) return;
  if(!key()) {
    showApiKeyModal();
    throw new Error("Vor der Transkription ist ein API-Schlüssel erforderlich.");
  }
  if(!sessionId) await createSession();
  status("Sprachnachricht wird transkribiert...");
  const form = new FormData();
  form.append("voice", draftChatRecordedVoiceBlob, draftChatRecordedVoiceName || "recording.webm");
  try {
    const transcriptData = await api(`/app/sessions/${sessionId}/draft/chat/transcribe`, {
      method:"POST",
      headers:{"X-API-Key":key()},
      body:form
    });
    draftChatRecordedVoiceBlob = null;
    const transcribedText = (transcriptData && transcriptData.text) || "";
    if(transcribedText) {
      const inputField = document.getElementById("draftChatInput");
      if(inputField.value.trim()) {
        inputField.value += " " + transcribedText;
      } else {
        inputField.value = transcribedText;
      }
      updateButtons();
      scheduleUiCacheSave();
      status("Sprachnachricht transkribiert und zur Nachricht hinzugefügt.");
    } else {
      throw new Error("Transkription war leer.");
    }
  } catch(err) {
    throw new Error(`Transkription fehlgeschlagen: ${err.message}`);
  }
}
function syncDraftCsvFromTable(){
  const fields = [...document.querySelectorAll("[data-csv-field]")];
  if(!fields.length) return document.getElementById("draftCsv").value;
  const headers = fields.map(field => field.dataset.csvField);
  const values = fields.map(field => field.value);
  const csvText = csvFromRows([headers, values]);
  document.getElementById("draftCsv").value = csvText;
  return csvText;
}
function renderDraftChat(chat){
  const target = document.getElementById("draftChatLog");
  if(!chat || !chat.length){
    target.className = "chat-log summary";
    target.textContent = "Erstelle zuerst einen Entwurf. Danach kannst du hier Änderungen anfragen.";
    return;
  }
  target.className = "chat-log";
  target.innerHTML = "";
  chat.forEach(item => {
    const div = document.createElement("div");
    div.className = `chat-message ${item.role === "user" ? "user" : "assistant"}`;
    div.textContent = item.content || "";
    target.appendChild(div);
  });
  target.scrollTop = target.scrollHeight;
}
function logsUrl(){
  return sessionId ? `/app/sessions/${sessionId}/logs` : "";
}
function renderSessionLogsWindow(targetSessionId, data){
  const logWindow = window.open("", "_blank");
  const html = `<!doctype html><html><head><meta charset="utf-8"><title>Session-Logs ${esc(targetSessionId || sessionId || "")}</title>` +
    `<style>body{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;margin:24px;line-height:1.45;color:#1f2933;background:#fff}` +
    `pre{white-space:pre-wrap;overflow-wrap:anywhere;background:#f5f7fa;border:1px solid #d8e0e8;border-radius:8px;padding:14px}</style></head>` +
    `<body><h1>Session-Logs</h1><pre>${esc(JSON.stringify(data, null, 2))}</pre></body></html>`;
  if(logWindow){
    logWindow.document.open();
    logWindow.document.write(html);
    logWindow.document.close();
  } else {
    status(data);
  }
}
function cachedSessionLogsData(){
  const cachedPost = (currentSessionData && currentSessionData.wordpress_post) || {};
  const importLogs = cachedPost.import_logs || {};
  if(!Object.keys(importLogs).length) return null;
  return {
    session_id: sessionId || (currentSessionData && currentSessionData.session_id) || "",
    source: "client_cache",
    note: "Server-Session wurde nicht gefunden. Zeige gecachte Import-Logs aus der letzten WordPress-Antwort.",
    state: currentSessionData || {},
    ui_cache: (currentSessionData && currentSessionData.ui_cache) || {},
    cache_files: {},
    draft_csv: ((currentSessionData && currentSessionData.draft) || {}).csv_text || "",
    wordpress_import_output_dir: cachedPost.output_dir || null,
    wordpress_import_logs: importLogs,
  };
}
async function openSessionLogs(){
  if(!sessionId && !cachedSessionLogsData()) throw new Error("Bitte zuerst eine Session erstellen.");
  let data;
  if(sessionId){
    try {
      data = await api(logsUrl(), {headers:headers(false)});
    } catch (error) {
      if(!isMissingSessionError(error)) throw error;
      data = cachedSessionLogsData();
      if(!data) throw error;
    }
  } else {
    data = cachedSessionLogsData();
  }
  renderSessionLogsWindow(sessionId, data);
}
async function openSessionLogsById(targetSessionId){
  const id = String(targetSessionId || "").trim();
  if(!id) throw new Error("Session-ID fehlt.");
  const data = await api(`/app/sessions/${encodeURIComponent(id)}/logs`, {headers:headers(false)});
  renderSessionLogsWindow(id, data);
}
async function loadSessionById(targetSessionId){
  const id = String(targetSessionId || "").trim();
  if(!id) throw new Error("Session-ID fehlt.");
  sessionId = id;
  sessionStorage.setItem("flairlab_session_id", sessionId);
  await restoreSession();
  closePanel("panelSession");
}
function recentSessionsQueryString(){
  const params = new URLSearchParams();
  const limit = Math.max(1, Math.min(200, Number(document.getElementById("recentLimit").value || 20)));
  params.set("limit", String(limit));
  const clientId = (document.getElementById("recentClientId").value || "").trim();
  const postType = (document.getElementById("recentPostType").value || "").trim();
  const statusValue = (document.getElementById("recentStatus").value || "").trim();
  if(clientId) params.set("client_id", clientId);
  if(postType) params.set("post_type", postType);
  if(statusValue) params.set("status", statusValue);
  return params.toString();
}
function formatArchiveDateTime(value){
  const text = String(value || "").trim();
  if(!text) return "-";
  const date = new Date(text);
  if(Number.isNaN(date.getTime())) return text;
  return date.toLocaleString("de-DE", {dateStyle:"medium", timeStyle:"short"});
}
function archiveBool(value){
  return value ? "true" : "false";
}
function updateRecentSessionsSelectAllState(){
  const checkbox = document.getElementById("recentSessionsSelectAll");
  if(!checkbox) return;
  if(!recentSessionIdsInView.length){
    checkbox.checked = false;
    checkbox.indeterminate = false;
    return;
  }
  const selectedCount = recentSessionIdsInView.filter(id => recentSelectedSessionIds.has(id)).length;
  checkbox.checked = selectedCount === recentSessionIdsInView.length;
  checkbox.indeterminate = selectedCount > 0 && selectedCount < recentSessionIdsInView.length;
}
function toggleRecentSessionSelection(sessionId, checked){
  const id = String(sessionId || "").trim();
  if(!id) return;
  if(checked) recentSelectedSessionIds.add(id);
  else recentSelectedSessionIds.delete(id);
  updateRecentSessionsSelectAllState();
}
function toggleAllRecentSessions(checked){
  recentSessionIdsInView.forEach(id => {
    if(checked) recentSelectedSessionIds.add(id);
    else recentSelectedSessionIds.delete(id);
  });
  document.querySelectorAll("input[data-recent-session-id]").forEach(input => {
    input.checked = !!checked;
  });
  updateRecentSessionsSelectAllState();
}
async function deleteSelectedRecentSessions(){
  if(!key()) throw new Error("Bitte zuerst API-Schlüssel eingeben.");
  const sessionIds = [...recentSelectedSessionIds];
  if(!sessionIds.length) throw new Error("Bitte zuerst mindestens eine Session auswählen.");
  const data = await api("/app/sessions/delete", {
    method:"POST",
    headers:headers(),
    body:JSON.stringify({session_ids: sessionIds}),
  });
  if(sessionIds.includes(sessionId)){
    sessionId = "";
    currentSessionData = null;
    sessionStorage.removeItem("flairlab_session_id");
    document.getElementById("sessionSummary").textContent = "Keine aktive Session.";
  }
  recentSelectedSessionIds = new Set();
  await loadRecentSessions();
  status(`${Number(data.deleted || 0)} Session(s) gelöscht.`);
}
function renderRecentSessions(payload){
  const target = document.getElementById("recentSessionsList");
  const sessions = (payload && payload.sessions) || [];
  recentSessionIdsInView = sessions
    .map(item => String((item && item.session_id) || "").trim())
    .filter(Boolean);
  recentSelectedSessionIds.forEach(id => {
    if(!recentSessionIdsInView.includes(id)) recentSelectedSessionIds.delete(id);
  });
  if(!sessions.length){
    target.className = "summary";
    target.textContent = "Keine Sessions für die aktuellen Filter gefunden.";
    updateRecentSessionsSelectAllState();
    return;
  }
  target.className = "recent-sessions-list";
  target.innerHTML = "";
  sessions.forEach(item => {
    const sessionItemId = String((item && item.session_id) || "").trim();
    const card = document.createElement("div");
    card.className = "recent-session-item";

    const head = document.createElement("div");
    head.className = "recent-session-head";
    const selector = document.createElement("input");
    selector.type = "checkbox";
    selector.dataset.recentSessionId = sessionItemId;
    selector.checked = recentSelectedSessionIds.has(sessionItemId);
    selector.addEventListener("change", () => toggleRecentSessionSelection(sessionItemId, selector.checked));
    head.appendChild(selector);

    const title = document.createElement("strong");
    title.textContent = `${item.session_id || "-"} (${item.status || "-"})`;
    head.appendChild(title);
    card.appendChild(head);

    const meta = document.createElement("div");
    meta.className = "recent-session-meta";
    meta.innerHTML =
      `Client: ${esc(item.client_id || "-")}<br>` +
      `Post Type: ${esc(item.post_type || "-")}<br>` +
      `Created: ${esc(formatArchiveDateTime(item.created_at))}<br>` +
      `Updated: ${esc(formatArchiveDateTime(item.updated_at || item.created_at))}<br>` +
      `Storage: ${esc(item.storage || "-")}<br>` +
      `Post ID: ${esc(item.wordpress_post_id == null ? "-" : String(item.wordpress_post_id))}<br>` +
      `Bilder hochgeladen: ${esc(archiveBool(!!item.has_images))}<br>` +
      `Transkript erstellt: ${esc(archiveBool(!!item.has_transcript))}<br>` +
      `Entwurf erstellt: ${esc(archiveBool(!!item.has_draft))}`;
    card.appendChild(meta);

    const missingTotal = Number(item.missing_media_total || 0);
    if(missingTotal > 0){
      const parts = [];
      if(Number(item.missing_media_images || 0)) parts.push(`${item.missing_media_images} Bilder`);
      if(Number(item.missing_media_videos || 0)) parts.push(`${item.missing_media_videos} Videos`);
      if(Number(item.missing_media_voices || 0)) parts.push(`${item.missing_media_voices} Sprachnachrichten`);
      const suffix = parts.length ? ` (${parts.join(", ")})` : "";
      const warning = document.createElement("div");
      warning.className = "recent-session-warning";
      warning.textContent = `${missingTotal} Medien-Datei(en) nicht wiederherstellbar${suffix}. Beim Laden ist ggf. erneuter Upload nötig.`;
      card.appendChild(warning);
    }

    const actions = document.createElement("div");
    actions.className = "recent-session-actions";

    const logsButton = document.createElement("button");
    logsButton.type = "button";
    logsButton.className = "secondary";
    logsButton.textContent = "Logs öffnen";
    logsButton.addEventListener("click", () => run(() => openSessionLogsById(item.session_id)));
    actions.appendChild(logsButton);

    const loadButton = document.createElement("button");
    loadButton.type = "button";
    loadButton.textContent = "Session laden";
    loadButton.addEventListener("click", () => run(() => loadSessionById(item.session_id)));
    actions.appendChild(loadButton);

    card.appendChild(actions);
    target.appendChild(card);
  });
  updateRecentSessionsSelectAllState();
}
async function loadRecentSessions(){
  if(!key()) throw new Error("Bitte zuerst API-Schlüssel eingeben.");
  const data = await api(`/app/sessions/recent?${recentSessionsQueryString()}`, {headers:headers(false)});
  renderRecentSessions(data);
}
function closeResultModal(){
  document.getElementById("resultModal").classList.remove("open");
}
function showApiKeyModal(){
  document.getElementById("apiKeyModalInput").value = key();
  document.getElementById("apiKeyModal").classList.add("open");
  setTimeout(() => document.getElementById("apiKeyModalInput").focus(), 0);
}
function closeApiKeyModal(){
  document.getElementById("apiKeyModal").classList.remove("open");
}
async function saveKeyFromModal(){
  const modalKey = document.getElementById("apiKeyModalInput").value.trim();
  if(!modalKey) throw new Error("API-Schlüssel ist erforderlich.");
  document.getElementById("apiKey").value = modalKey;
  saveKey();
  closeApiKeyModal();
  if(!sessionId) await createSession();
  else await restoreSession();
}
function showResultModal(post){
  if(!post || !post.post_id) return;
  document.getElementById("resultModalSummary").innerHTML =
    `<strong>Beitrags-ID:</strong> ${esc(post.post_id || "")}<br>` +
    `<strong>Status:</strong> ${esc(post.status || "")}<br>` +
    `<strong>Schreibmodus:</strong> ${esc(post.post_write_mode || "")}<br>` +
    `<strong>Ausgabeordner:</strong> ${esc(post.output_dir || "")}`;
  const viewLink = post.view_url
    ? `<a href="${esc(post.view_url)}" target="_blank" rel="noopener">Beitrag ansehen</a>`
    : "";
  const editLink = post.edit_url
    ? `<a href="${esc(post.edit_url)}" target="_blank" rel="noopener">Beitrag bearbeiten</a>`
    : "";
  document.getElementById("resultModalLinks").innerHTML =
    viewLink +
    editLink +
    `<a href="#" class="secondary" onclick="run(openSessionLogs); return false;">Session-Logs</a>` +
    `<button class="secondary" type="button" onclick="closeResultModal()">Schließen</button>`;
  document.getElementById("resultModal").classList.add("open");
}
function fileList(files){
  if(!files || files.length === 0) return "none";
  return files.map(file => esc(file.original_filename || file.filename)).join(", ");
}
function cacheWithValues(cache){
  return cache && Object.keys(cache).length ? cache : null;
}
function updateWordPressTargetInfo(post){
  const target = document.getElementById("updateTargetInfo");
  if(!target) return;
  if(post && post.post_id){
    const editLink = post.edit_url
      ? ` <a href="${esc(post.edit_url)}" target="_blank" rel="noopener">Edit öffnen</a>`
      : "";
    target.innerHTML = `<strong>Update-Ziel:</strong> Beitrag ${esc(post.post_id)}.${editLink}`;
    return;
  }
  target.textContent = "Update-Ziel: noch kein zuvor erstellter Beitrag in dieser Session.";
}
function renderSession(data, options={}){
  if(!data) return;
  currentSessionData = data;
  document.getElementById("clientId").value = data.client_id || "flairlab";
  document.getElementById("postType").value = data.post_type || "Event";
  const transcript = data.transcript || {};
  const draft = data.draft || {};
  document.getElementById("transcript").value = transcript.text || "";
  document.getElementById("draftCsv").value = draft.csv_text || "";
  renderDraftTable(draft.csv_text || "");
  renderDraftChat(draft.chat || []);
  if(draft.category) document.getElementById("category").value = draft.category;
  const files = data.files || {};
  const featured = files.featured_image_filename || "none";
  const featuredMediaId = files.featured_image_media_id || "";
  const visionSelectedCount = (files.vision_selected_filenames || []).length;
  const wpMediaCount = Number(((data.wordpress_media_library || {}).uploaded_count) || 0);
  const aiUsage = (data.ai_usage && typeof data.ai_usage === "object") ? data.ai_usage : {};
  const imageProcessing = (data.image_processing && typeof data.image_processing === "object") ? data.image_processing : {};
  const imageProcessingStatus = String(imageProcessing.status || "").trim() || "idle";
  const post = data.wordpress_post || {};
  updateWordPressTargetInfo(post);
  document.getElementById("postLinks").innerHTML = post.post_id
    ? `${post.view_url ? `<a href="${esc(post.view_url)}" target="_blank" rel="noopener">Beitrag ansehen</a>` : ""}${post.edit_url ? `<a href="${esc(post.edit_url)}" target="_blank" rel="noopener">Beitrag bearbeiten</a>` : ""}<a href="#" onclick="run(openSessionLogs); return false;">Session-Logs</a>`
    : "";
  document.getElementById("sessionSummary").innerHTML =
    `<strong>Session:</strong> ${esc(data.session_id)}<br>` +
    `<strong>Status:</strong> ${esc(data.status || "created")}<br>` +
    `<strong>Sprache:</strong> ${fileList(files.voices || (files.voice ? [files.voice] : []))}<br>` +
    `<strong>Bilder:</strong> ${fileList(files.images)}<br>` +
    `<strong>Titelbild:</strong> ${esc(featured)}<br>` +
    `<strong>Video:</strong> ${fileList(files.videos)}<br>` +
    `<strong>Pillow:</strong> ${esc(imageProcessingStatus)}<br>` +
    `<strong>Vision-Auswahl:</strong> ${esc(String(visionSelectedCount))}<br>` +
    `<strong>WP Mediathek Uploads:</strong> ${esc(String(wpMediaCount))}<br>` +
    `<strong>AI Nutzung:</strong> ${esc(aiUsageSummaryText(aiUsage))}`;
  renderAiUsage(aiUsage);
  renderMediaRecoveryNotice(data.media_recovery || {});
  renderVoiceList(files.voices || (files.voice ? [files.voice] : []), [...document.getElementById("voice").files]);
  renderUploadedImagePreviews(files.images || [], featured, featuredMediaId);
  renderUploadedVideoPreviews(files.videos || []);
  openWorkflowPanels(data);
  if(options.applyCache !== false){
    const localCache = uiCacheKey() ? JSON.parse(sessionStorage.getItem(uiCacheKey()) || "null") : null;
    applyUiCache(cacheWithValues(data.ui_cache) || localCache);
    openPanelsFromCurrentInputs();
  }
  if(imageProcessingStatus === "processing"){
    schedulePillowStatusPoll();
  }
  updateButtons();
}
async function renderFreshSession(data){
  renderSession(data, {applyCache:false});
  saveUiCache().catch(error => console.warn("UI cache save failed", error));
}
function mergeSessionData(baseData, nextData){
  if(!baseData) return nextData;
  if(!nextData) return baseData;
  return {
    ...baseData,
    ...nextData,
    files: {
      ...(baseData.files || {}),
      ...(nextData.files || {})
    },
    transcript: {
      ...(baseData.transcript || {}),
      ...(nextData.transcript || {})
    },
    draft: {
      ...(baseData.draft || {}),
      ...(nextData.draft || {})
    },
    wordpress_post: {
      ...(baseData.wordpress_post || {}),
      ...(nextData.wordpress_post || {})
    },
    ui_cache: {
      ...(baseData.ui_cache || {}),
      ...(nextData.ui_cache || {})
    },
    media_recovery: {
      ...(baseData.media_recovery || {}),
      ...(nextData.media_recovery || {})
    },
    vision: {
      ...(baseData.vision || {}),
      ...(nextData.vision || {})
    },
    wordpress_media_library: {
      ...(baseData.wordpress_media_library || {}),
      ...(nextData.wordpress_media_library || {})
    },
    image_processing: {
      ...(baseData.image_processing || {}),
      ...(nextData.image_processing || {})
    },
    ai_usage: {
      ...(baseData.ai_usage || {}),
      ...(nextData.ai_usage || {})
    }
  };
}
function updateButtons(){
  const hasSession = !!sessionId;
  const hasVoice = hasUploadedVoice();
  const hasTranscript = !!document.getElementById("transcript").value.trim();
  const hasDraft = !!document.getElementById("draftCsv").value.trim();
  const hasChatMessage = !!document.getElementById("draftChatInput").value.trim();
  const hasPendingUpload = !!recordedVoiceBlob || !!document.getElementById("voice").files.length || hasPendingMediaFiles();
  const sessionFiles = (currentSessionData && currentSessionData.files) || {};
  const uploadedImages = sessionFiles.images || [];
  const canTranscribe = hasSession && hasVoice && !hasPendingUpload;
  const canGenerateDraft = hasSession && hasTranscript && uploadedImages.length > 0 && !hasPendingUpload;
  document.getElementById("uploadButton").disabled = !hasSession || !hasPendingUpload;
  document.getElementById("uploadWpMediaButton").disabled = !hasSession || !uploadedImages.length;
  document.getElementById("transcribeButton").disabled = !canTranscribe;
  document.getElementById("saveTranscriptButton").disabled = !hasSession || !hasTranscript;
  document.getElementById("generateFactsButton").disabled = !hasSession || !hasTranscript;
  document.getElementById("generateDraftButton").disabled = !canGenerateDraft;
  document.getElementById("transcriptNextButton").disabled = !canGenerateDraft;
  document.getElementById("saveDraftButton").disabled = !hasSession || !hasDraft;
  document.getElementById("sendDraftChatButton").disabled = !hasSession || !hasDraft || !hasChatMessage;
  document.getElementById("createPostButton").disabled = !hasSession || !hasDraft;
  const hasExistingPost = !!(currentSessionData && currentSessionData.wordpress_post && currentSessionData.wordpress_post.post_id);
  document.getElementById("updatePostButton").disabled = !hasSession || !hasDraft || !hasExistingPost;
}
async function restoreSession(){
  if(!sessionId || !key()) {
    updateButtons();
    return;
  }
  try {
    const data = await api(`/app/sessions/${sessionId}`, {headers:headers(false)});
    clearSessionRecoveryBanner();
    renderSession(data);
    status(data);
  } catch (error) {
    console.warn(error);
    const staleCacheKey = uiCacheKey();
    const localCache = staleCacheKey ? JSON.parse(sessionStorage.getItem(staleCacheKey) || "null") : null;
    if(isMissingSessionError(error)){
      applyUiCache(localCache);
      openPanelsFromCurrentInputs();
      await recreateMissingSession();
      if(localCache){
        applyUiCache(localCache);
        scheduleUiCacheSave();
      }
      status("Server-Session wurde neu erstellt und lokale Änderungen wurden wiederhergestellt.");
      return;
    }
    applyUiCache(localCache);
    openPanelsFromCurrentInputs();
    status("Server-Session konnte nicht geladen werden. Lokale Änderungen im Browser wurden behalten.");
  } finally {
    closePanel("panelSession");
    updateButtons();
  }
}
async function initializeApp(){
  updateButtons();
  await loadKnowledgeStatus().catch(error => console.warn(error));
  if(!key()){
    showApiKeyModal();
    status("Bitte API-Schlüssel eingeben, um zu starten.");
    return;
  }
  if(sessionId){
    await restoreSession();
    loadRecentSessions().catch(error => console.warn(error));
    return;
  }
  await createSession();
}
async function run(fn){
  const fnName = (fn && fn.name) || "";
  const startedAt = Date.now();
  let failed = false;
  const nonBlockingActions = new Set([
    "uploadFiles",
    "uploadPendingMediaFiles",
    "uploadWordPressMediaLibrary",
    "transcribe",
    "sendImageOptimizationPrompt",
    "applyVisionMetadataFromCompare",
    "saveImageMetadata",
    "restoreComparedImageToOriginal",
    "openImageOptimizationPrompt",
  ]);
  const useOverlay = !nonBlockingActions.has(fnName);
  if(useOverlay){
    setBusy(true, actionLabel(fn));
  } else {
    document.getElementById("statusRailStateText").textContent = actionLabel(fn);
    document.body.classList.add("is-busy");
  }
  try {
    await fn();
  } catch (error) {
    failed = true;
    console.error(error);
    showErrorModal(error);
    status(appendElapsedDuration("Fehler: " + readableError(error), startedAt));
  } finally {
    if(useOverlay){
      setBusy(false);
    } else {
      document.getElementById("statusRailStateText").textContent = appendElapsedDuration("Fertig", startedAt);
      document.body.classList.remove("is-busy");
      setTimeout(() => {
        if(busyCount === 0) document.getElementById("statusRailStateText").textContent = "Bereit";
      }, 1200);
    }
    if(!failed){
      const currentStatus = document.getElementById("status").textContent || actionLabel(fn);
      status(appendElapsedDuration(currentStatus, startedAt));
      document.getElementById("statusRailStateText").textContent = appendElapsedDuration("Fertig", startedAt);
    }
    updateButtons();
  }
  }
function saveKey(){ sessionStorage.setItem("flairlab_api_key", key()); status("API-Schlüssel wurde in dieser Browser-Session gespeichert."); loadKnowledgeStatus().catch(error => console.warn(error)); loadRecentSessions().catch(error => console.warn(error)); }
async function saveKeyAndMaybeCreateSession(){
  saveKey();
  if(!sessionId && key()) await createSession();
}
async function api(path, options={}) {
  const res = await fetch(path, options);
  const text = await res.text();
  let data; try { data = JSON.parse(text); } catch { data = text; }
  if (!res.ok) throw new Error(formatApiError(data));
  return data;
}
function formatApiError(data){
  if(typeof data === "string") return data;
  const detail = data && (data.detail || data.message || data.error || data);
  if(detail && typeof detail === "object"){
    const title = detail.message || detail.error_code || "HTTP request failed";
    if(Array.isArray(detail.details) && detail.details.length){
      const rows = detail.details.map(item => {
        const location = [item.sheet, item.row ? `Zeile ${item.row}` : "", item.column ? `Spalte ${item.column}` : ""]
          .filter(Boolean).join(" / ");
        return location ? `${location}: ${item.message}` : item.message;
      });
      return `${title}\\n${rows.join("\\n")}`;
    }
    return JSON.stringify(detail);
  }
  return String(detail || "HTTP request failed");
}
async function createSession(){
  saveKey();
  const data = await api("/app/sessions", {
    method:"POST",
    headers:headers(),
    body:JSON.stringify({client_id:document.getElementById("clientId").value || "flairlab", post_type:document.getElementById("postType").value})
  });
  sessionId = data.session_id;
  sessionStorage.setItem("flairlab_session_id", sessionId);
  clearSessionRecoveryBanner();
  await renderFreshSession(data);
  loadRecentSessions().catch(error => console.warn(error));
  openPanel("panelUpload", true);
  status(data);
  updateButtons();
}
async function createNewPostSession(){
  sessionId = "";
  currentSessionData = null;
  clearSessionRecoveryBanner();
  clearMediaRecoveryNotice();
  document.getElementById("voice").value = "";
  document.getElementById("images").value = "";
  document.getElementById("videos").value = "";
  revokeImagePreviewUrls();
  revokeVideoPreviewUrls();
  document.getElementById("imagePreviews").innerHTML = "";
  document.getElementById("videoPreviews").innerHTML = "";
  document.getElementById("voiceUploads").textContent = "Noch keine Sprachnachricht gespeichert.";
  document.getElementById("sessionSummary").textContent = "Keine aktive Session.";
  sessionStorage.removeItem("flairlab_session_id");
  await createSession();
}
async function recreateMissingSession(){
  const previousSessionId = sessionId;
  const preservedSessionData = currentSessionData || {};
  const transcriptText = document.getElementById("transcript").value || "";
  saveKey();
  const created = await api("/app/sessions", {
    method:"POST",
    headers:headers(),
    body:JSON.stringify({client_id:document.getElementById("clientId").value || "flairlab", post_type:document.getElementById("postType").value})
  });
  sessionId = created.session_id;
  sessionStorage.setItem("flairlab_session_id", sessionId);
  if(transcriptText.trim()){
    await api(`/app/sessions/${sessionId}/transcript`, {
      method:"PUT",
      headers:headers(),
      body:JSON.stringify({text:transcriptText})
    });
  }
  currentSessionData = mergeSessionData(created, preservedSessionData);
  await renderFreshSession(currentSessionData);
  loadRecentSessions().catch(error => console.warn(error));
  showSessionRecoveryBanner(previousSessionId, sessionId);
  status("Server-Session wurde neu erstellt. Die Aktion wird erneut ausgeführt.");
}
function renderImageChoices(){
  const wrap = document.getElementById("featuredChoices");
  wrap.innerHTML = "";
  const files = [...document.getElementById("images").files];
  revokeImagePreviewUrls();
  renderImagePreviewItems(files.map(file => {
    const url = URL.createObjectURL(file);
    imagePreviewUrls.push(url);
    return {url, name:file.name, value:file.name};
  }), {selectable:true, removable:true, kind:"images", inputId:"images"});
  updateButtons();
  scheduleAutoImageUpload();
}
function scheduleAutoImageUpload(){
  clearTimeout(imageAutoUploadTimer);
  imageAutoUploadTimer = setTimeout(() => {
    autoUploadSelectedImages().catch(error => {
      console.error(error);
      showErrorModal(error);
      status("Fehler: " + readableError(error));
    });
  }, 250);
}
async function autoUploadSelectedImages(){
  if(imageAutoUploadInFlight){
    imageAutoUploadQueued = true;
    return;
  }
  const selectedImages = [...document.getElementById("images").files];
  if(!selectedImages.length) return;
  if(!key()){
    status("Bitte zuerst API-Schlüssel eingeben, damit Bilder hochgeladen und verarbeitet werden können.");
    return;
  }
  if(!sessionId){
    await createSession();
  }

  imageAutoUploadInFlight = true;
  try {
    const form = new FormData();
    selectedImages.forEach(file => form.append("images", file));
    const featured = document.querySelector("input[name='featured']:checked");
    if(featured) form.append("featured_image_filename", featured.value);
    status("Bilder werden hochgeladen und per Pillow verarbeitet...");
    const data = await api(`/app/sessions/${sessionId}/uploads`, {method:"POST", headers:{"X-API-Key":key()}, body:form});
    clearUploadInputs({images:true});
    await renderFreshSession(data);
    status("Bilder hochgeladen. Verarbeitete Bilder können sofort mit 'Edits anzeigen' geöffnet werden.");
    schedulePillowStatusPoll();
  } finally {
    imageAutoUploadInFlight = false;
    if(imageAutoUploadQueued){
      imageAutoUploadQueued = false;
      scheduleAutoImageUpload();
    }
  }
}
function schedulePillowStatusPoll(){
  clearTimeout(pillowStatusPollTimer);
  if(!sessionId || !key()) return;
  const startedAt = Date.now();
  const poll = async () => {
    if(!sessionId || !key()) return;
    try {
      const data = await api(`/app/sessions/${sessionId}`, {headers:headers(false)});
      const processing = (data.image_processing && typeof data.image_processing === "object") ? data.image_processing : {};
      const statusValue = String(processing.status || "").trim().toLowerCase();
      if(statusValue === "complete"){
        const completedAt = String(processing.updated_at || "").trim();
        lastPillowProgressKey = "";
        await renderFreshSession(data);
        pillowRecoveryTriggered = false;
        if(completedAt && completedAt !== lastPillowCompletedAt){
          lastPillowCompletedAt = completedAt;
          status("Pillow processing finished. Click 'Edits anzeigen' on a processed image.");
        }
        return;
      }
      if(statusValue === "processing"){
        const progressKey = `${Number(processing.processed_count || 0)}/${Number(processing.total_count || 0)}/${String(processing.updated_at || "")}`;
        if(progressKey !== lastPillowProgressKey){
          lastPillowProgressKey = progressKey;
          currentSessionData = mergeSessionData(currentSessionData, data);
          refreshImagePreviewProcessingState();
          updateButtons();
        }
      }
      if(statusValue === "processing" && Date.now() - startedAt > 12000 && !pillowRecoveryTriggered){
        pillowRecoveryTriggered = true;
        try {
          await api(`/app/sessions/${sessionId}/images/process`, {method:"POST", headers:headers(false)});
        } catch (error) {
          console.warn("Pillow recovery probe failed", error);
        }
      }
      if(statusValue === "processing" && Date.now() - startedAt < 180000){
        pillowStatusPollTimer = setTimeout(() => {
          poll().catch(error => console.warn("Pillow poll failed", error));
        }, 2200);
      }
    } catch (error) {
      console.warn("Pillow status poll failed", error);
    }
  };
  poll().catch(error => console.warn("Pillow poll failed", error));
}
function renderVideoChoices(){
  const files = [...document.getElementById("videos").files];
  revokeVideoPreviewUrls();
  const items = files.map(file => {
    const url = URL.createObjectURL(file);
    videoPreviewUrls.push(url);
    return {url, name:file.name, value:file.name, mediaType:"video"};
  });
  renderMediaPreviewItems("videoPreviews", items, {removable:true, selected:true, kind:"videos", inputId:"videos"});
  updateButtons();
}
function clearUploadInputs(options={}){
  if(options.voices) document.getElementById("voice").value = "";
  if(options.images) {
    document.getElementById("images").value = "";
    document.getElementById("featuredChoices").innerHTML = "";
  }
  if(options.videos) document.getElementById("videos").value = "";
}
function appendSelectedMediaToForm(form){
  [...document.getElementById("images").files].forEach(file => form.append("images", file));
  const video = document.getElementById("videos").files[0];
  if(video) form.append("videos", video);
  const featured = document.querySelector("input[name='featured']:checked");
  if(featured) form.append("featured_image_filename", featured.value);
}
function hasPendingMediaFiles(){
  return !!document.getElementById("images").files.length || !!document.getElementById("videos").files.length;
}
function hasUploadedVoice(){
  const files = (currentSessionData && currentSessionData.files) || {};
  return !!files.voice || !!(files.voices && files.voices.length);
}
async function uploadPendingMediaFiles(){
  if(!hasPendingMediaFiles()) return null;
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  const hadImages = !!document.getElementById("images").files.length;
  const form = new FormData();
  appendSelectedMediaToForm(form);
  status("Medien werden gespeichert...");
  const data = await api(`/app/sessions/${sessionId}/uploads`, {method:"POST", headers:{"X-API-Key":key()}, body:form});
  clearUploadInputs({images:true, videos:true});
  await renderFreshSession(data);
  if(hadImages) schedulePillowStatusPoll();
  return data;
}
async function uploadFiles(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  const form = new FormData();
  const voices = [...document.getElementById("voice").files];
  const hadImages = !!document.getElementById("images").files.length;
  const hasNewVoice = !!voices.length || !!recordedVoiceBlob;
  const hasMedia = hasPendingMediaFiles();
  if(!hasNewVoice && !hasMedia){
    status("Keine neue Auswahl zum Speichern.");
    return;
  }
  if(!hasNewVoice && hasMedia){
    await uploadPendingMediaFiles();
    openPanel("panelUpload");
    status("Medien wurden gespeichert. Bitte Titelbild auswählen und danach den Entwurf erstellen.");
    updateButtons();
    return;
  }
  if(recordedVoiceBlob) form.append("voices", recordedVoiceBlob, recordedVoiceName || "recording.webm");
  voices.forEach(file => form.append("voices", file));
  appendSelectedMediaToForm(form);
  const data = await api(`/app/sessions/${sessionId}/uploads`, {method:"POST", headers:{"X-API-Key":key()}, body:form});
  clearUploadInputs({voices:true, images:true, videos:true});
  recordedVoiceBlob = null;
  await renderFreshSession(data);
  if(hadImages) schedulePillowStatusPoll();
  openPanel("panelUpload");
  status("Dateien hochgeladen. Alle Sprachnachrichten werden transkribiert...");
  const transcriptData = await api(`/app/sessions/${sessionId}/transcribe`, {method:"POST", headers:headers(false)});
  document.getElementById("transcript").value = (transcriptData.transcript && transcriptData.transcript.text) || "";
  await renderFreshSession(transcriptData);
  openPanel("panelUpload");
  status(transcriptData);
  updateButtons();
}
async function uploadWordPressMediaLibrary(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  const data = await api(`/app/sessions/${sessionId}/wordpress/media-library-upload`, {method:"POST", headers:headers(false)});
  const payload = data.wordpress_media_library || {};
  const uploaded = Number(payload.uploaded_count || 0);
  currentSessionData = mergeSessionData(currentSessionData, {wordpress_media_library: payload});
  status(`WP Mediathek-Upload abgeschlossen: ${uploaded} Bild(er) hochgeladen.`);
  updateButtons();
}
async function transcribe(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  status("Transkription läuft...");
  const data = await api(`/app/sessions/${sessionId}/transcribe`, {method:"POST", headers:headers(false)});
  document.getElementById("transcript").value = (data.transcript && data.transcript.text) || "";
  await renderFreshSession(data);
  openPanel("panelTranscript", true);
  status(data);
  updateButtons();
}
async function saveTranscript(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  const data = await api(`/app/sessions/${sessionId}/transcript`, {
    method:"PUT",
    headers:headers(),
    body:JSON.stringify({text:document.getElementById("transcript").value})
  });
  currentSessionData = mergeSessionData(currentSessionData, data);
  await renderFreshSession(currentSessionData);
  openPanel("panelTranscript", true);
  status(data);
  updateButtons();
}
async function generateFactsFromTranscript(){
  await saveTranscript();
  status("Transkript gespeichert. Fakten können jetzt aus dem Transkript generiert werden.");
}
async function generateDraft(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  if(!document.getElementById("transcript").value.trim()) throw new Error("Bitte zuerst transkribieren oder Notizen eintragen.");
  if(hasPendingMediaFiles()) await uploadPendingMediaFiles();
  status("CSV-Entwurf wird erstellt...");
  const data = await api(`/app/sessions/${sessionId}/draft`, {
    method:"POST",
    headers:headers(),
    body:JSON.stringify({
      category:document.getElementById("category").value || "auto event post",
      force_regenerate:true
    })
  });
  currentSessionData = mergeSessionData(currentSessionData, data);
  await renderFreshSession(currentSessionData);
  openPanel("panelDraft", true);
  status(data);
  updateButtons();
}
async function saveDraft(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  const csvText = syncDraftCsvFromTable();
  let data;
  try {
    data = await api(`/app/sessions/${sessionId}/draft`, {
      method:"PUT",
      headers:headers(),
      body:JSON.stringify({csv_text:csvText})
    });
  } catch (error) {
    if(!isMissingSessionError(error)) throw error;
    await recreateMissingSession();
    data = await api(`/app/sessions/${sessionId}/draft`, {
      method:"PUT",
      headers:headers(),
      body:JSON.stringify({csv_text:csvText})
    });
  }
  currentSessionData = mergeSessionData(currentSessionData, data);
  await renderFreshSession(currentSessionData);
  openPanel("panelDraft", true);
  status(data);
  updateButtons();
}
async function sendDraftChat(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  const message = document.getElementById("draftChatInput").value.trim();
  if(!message) throw new Error("Bitte eine Nachricht an den Entwurfs-Agenten schreiben.");
  await saveDraft();
  status("Entwurf wird mit dem Agenten aktualisiert...");
  const data = await api(`/app/sessions/${sessionId}/draft/chat`, {
    method:"POST",
    headers:headers(),
    body:JSON.stringify({message})
  });
  document.getElementById("draftChatInput").value = "";
  currentSessionData = mergeSessionData(currentSessionData, data);
  await renderFreshSession(currentSessionData);
  openPanel("panelDraft", true);
  status(data);
  updateButtons();
}
async function goToWordPressStep(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  if(!document.getElementById("draftCsv").value.trim()) throw new Error("Bitte zuerst einen Entwurf erstellen.");
  await saveDraft();
  openPanel("panelWordPress", true);
  status("Entwurf gespeichert. Bitte Beitragsstatus wählen und den WordPress-Beitrag erstellen.");
}
async function createWordPressPost(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  syncDraftCsvFromTable();
  status("WordPress-Beitrag wird erstellt...");
  await saveDraft();
  const data = await api(`/app/sessions/${sessionId}/wordpress-post`, {
    method:"POST",
    headers:headers(),
    body:JSON.stringify({status:document.getElementById("postStatus").value, existing_post_mode:"update"})
  });
  currentSessionData = mergeSessionData(currentSessionData, data);
  showResultModal((currentSessionData && currentSessionData.wordpress_post) || data.wordpress_post);
  await renderFreshSession(currentSessionData);
  openPanel("panelWordPress", true);
  status(data);
  updateButtons();
}
async function updateExistingWordPressPost(){
  if(!sessionId) throw new Error("Bitte zuerst eine Session erstellen.");
  const post = (currentSessionData && currentSessionData.wordpress_post) || {};
  if(!post.post_id) throw new Error("Es gibt noch keinen zuvor erstellten WordPress-Beitrag in dieser Session.");
  syncDraftCsvFromTable();
  status("Vorhandener WordPress-Beitrag wird aktualisiert...");
  await saveDraft();
  const data = await api(`/app/sessions/${sessionId}/wordpress-post`, {
    method:"POST",
    headers:headers(),
    body:JSON.stringify({
      status:document.getElementById("postStatus").value,
      existing_post_mode:"update",
      existing_post_id:Number(post.post_id),
      update_existing_generated_post:true
    })
  });
  currentSessionData = mergeSessionData(currentSessionData, data);
  showResultModal((currentSessionData && currentSessionData.wordpress_post) || data.wordpress_post);
  await renderFreshSession(currentSessionData);
  openPanel("panelWordPress", true);
  status(data);
  updateButtons();
}
document.getElementById("resultModal").addEventListener("click", event => {
  if(event.target.id === "resultModal") closeResultModal();
});
document.getElementById("errorModal").addEventListener("click", event => {
  if(event.target.id === "errorModal") closeErrorModal();
});
document.getElementById("imageCompareModal").addEventListener("click", event => {
  if(event.target.id === "imageCompareModal") closeImageCompareModal();
});
document.getElementById("apiKeyModalInput").addEventListener("keydown", event => {
  if(event.key === "Enter") run(saveKeyFromModal);
});
["clientId", "postType", "category", "postStatus", "transcript", "draftChatInput"].forEach(id => {
  document.getElementById(id).addEventListener("input", () => { updateButtons(); scheduleUiCacheSave(); });
  document.getElementById(id).addEventListener("change", () => { updateButtons(); scheduleUiCacheSave(); });
});
document.getElementById("postType").addEventListener("change", () => {
  sessionStorage.removeItem("flairlab_knowledge_status");
  loadKnowledgeStatus().catch(error => console.warn(error));
});
window.addEventListener("error", e => {
  showErrorModal(e.error || e.message);
  status("Fehler: " + e.message);
});
window.addEventListener("unhandledrejection", e => {
  showErrorModal(e.reason);
  status("Fehler: " + readableError(e.reason));
});
if(window.FLAIRLAB_PIPELINE_VERSION !== "v2"){
  initializeApp().catch(error => {
    console.error(error);
    showErrorModal(error);
    status("Fehler: " + readableError(error));
  });
}
</script>
</body>
</html>
"""


def rendered_app_html() -> str:
    pipeline_script = (
        "<script>window.FLAIRLAB_PIPELINE_VERSION="
        + json.dumps(CONTENT_PIPELINE_VERSION)
        + ";</script>"
    )
    html = APP_HTML.replace("<body>", f"<body>{pipeline_script}", 1)
    if CONTENT_PIPELINE_VERSION == "v2":
        adapter = '<script src="/v2/legacy-ui-adapter.js"></script>'
        document, closing = html.rsplit("</body>", 1)
        html = f"{document}{adapter}</body>{closing}"
    return html


@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return rendered_app_html()


@app.get("/app", response_class=HTMLResponse)
def app_interface() -> str:
    return rendered_app_html()


@app.get("/v2/legacy-ui-adapter.js", include_in_schema=False)
def v2_legacy_ui_adapter() -> Response:
    return Response(
        (APP_ROOT / "app/v2/api/legacy_ui_adapter.js").read_text(encoding="utf-8"),
        media_type="application/javascript",
    )


@app.get("/favicon.ico", include_in_schema=False)
def favicon() -> Response:
    return Response(status_code=204)


@app.get("/.well-known/appspecific/com.chrome.devtools.json", include_in_schema=False)
def chrome_devtools_probe() -> dict[str, Any]:
    return {}


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/clients")
def clients(_: None = Depends(verify_api_key)) -> dict[str, Any]:
    flairlab = get_client_config("flairlab")
    return {
        "clients": [
            {
                "client_id": "flairlab",
                "wp_base_url": flairlab.wp_base_url,
                "status": "configured" if flairlab.wp_username and flairlab.wp_app_password else "missing_credentials",
            }
        ]
    }


@app.get("/app/wordpress/preflight")
def wordpress_preflight(
  client_id: str = Query("flairlab"),
  _: None = Depends(verify_api_key),
) -> dict[str, Any]:
  set_active_client(client_id)
  return preflight_wordpress_permissions(strict=False)


@app.get("/app/knowledge/status")
def get_knowledge_status(
  post_type: str | None = Query(None),
  _: None = Depends(verify_api_key),
) -> dict[str, Any]:
  return knowledge_status_payload(post_type=post_type)


@app.get("/app/knowledge/workbook")
def download_knowledge_workbook(_: None = Depends(verify_download_api_key)) -> FileResponse:
    path = active_knowledge_workbook_path()
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Knowledge workbook not found.")
    media_type = "application/vnd.ms-excel.sheet.macroEnabled.12" if path.suffix.lower() == ".xlsm" else None
    download_name = path.name if path.suffix.lower() == ".xlsm" else f"{path.stem}.xlsm"
    return FileResponse(
    path,
    media_type=media_type,
    filename=download_name,
    headers={"Cache-Control": "private, max-age=3600"},
    )


@app.post("/app/knowledge/workbook")
async def upload_knowledge_workbook(
    workbook: UploadFile = File(...),
  post_type: str | None = Form(None),
    _: None = Depends(verify_api_key),
) -> dict[str, Any]:
    filename = safe_upload_name(workbook.filename, "knowledge_workbook.xlsm")
    suffix = Path(filename).suffix.lower()
    if suffix not in WORKBOOK_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Upload an .xlsm or .xlsx workbook.")

    destination = configured_knowledge_workbook_path()
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f"{destination.stem}.uploading{suffix}")

    with temporary.open("wb") as output:
        while chunk := await workbook.read(1024 * 1024):
            output.write(chunk)

    guidance = load_workbook_guidance(
      temporary,
      post_type=None,
      preferred_sheet=KNOWLEDGE_WORKBOOK_SHEET,
    )
    has_legacy_guidance = bool(guidance.get("items"))
    has_post_types_schema = bool(guidance.get("schema_mode") == "post_types_v3")
    if guidance.get("error") or (not has_legacy_guidance and not has_post_types_schema):
        temporary.unlink(missing_ok=True)
        raise HTTPException(
            status_code=400,
            detail=guidance.get("error") or "Workbook did not expose any AI guidance rows.",
        )
    if CONTENT_PIPELINE_VERSION == "v2" or has_post_types_schema:
      try:
        WorkbookValidator().validate(WorkbookLoader().load(temporary))
      except V2Error as exc:
        temporary.unlink(missing_ok=True)
        raise HTTPException(
          status_code=400,
          detail={
            **exc.as_dict(),
            "message": f"Workbook validation failed: {exc.message}",
          },
        ) from exc
      except Exception as exc:
        temporary.unlink(missing_ok=True)
        raise HTTPException(
          status_code=400,
          detail={
            "error_code": "invalid_workbook",
            "message": f"Workbook validation failed: {exc}",
            "details": [],
          },
        ) from exc

    gcs_upload_info: dict[str, Any] = {}
    use_local_destination = True
    if configured_knowledge_workbook_gcs_uri() and knowledge_source_policy() != "local_only":
      try:
        gcs_upload_info = upload_knowledge_workbook_to_gcs(temporary)
        sync_knowledge_workbook_from_gcs(destination)
        temporary.unlink(missing_ok=True)
        use_local_destination = False
      except Exception as exc:
        if knowledge_source_policy() == "gcs_required" or os.getenv("K_SERVICE"):
          temporary.unlink(missing_ok=True)
          raise HTTPException(
            status_code=502,
            detail=(
              "Workbook validated, but upload to configured GCS workbook storage failed. "
              f"Set KNOWLEDGE_SOURCE_POLICY=local_only for local testing or configure Google credentials. {exc}"
            ),
          ) from exc
        gcs_upload_info = {
          "warning": (
            "GCS workbook upload failed; using the local workbook file for this development run."
          ),
          "error": str(exc),
        }
    if use_local_destination:
      if destination.exists():
        backup = destination.with_suffix(destination.suffix + ".bak")
        destination.replace(backup)
      temporary.replace(destination)
    if CONTENT_PIPELINE_VERSION == "v2":
      get_v2_service().knowledge.reload()
    storage_label = "GCS" if gcs_upload_info.get("gcs_uri") else "local"
    generation = gcs_upload_info.get("gcs_generation")
    message = f"Database Datei aktualisiert ({storage_label})."
    if generation:
      message = f"Database Datei aktualisiert (GCS generation {generation})."
    return {
        "success": True,
        "message": message,
      "storage": gcs_upload_info,
      **knowledge_status_payload(post_type=post_type),
    }


@app.post("/app/sessions", response_model=SessionCreateResponse)
def create_session(
    payload: SessionCreateRequest,
    _: None = Depends(verify_api_key),
) -> SessionCreateResponse:
    session_id = uuid.uuid4().hex
    state = {
        "session_id": session_id,
        "client_id": payload.client_id,
        "post_type": payload.post_type,
        "status": "created",
        "created_at": datetime.now(timezone.utc).isoformat(),
      "ai_usage": {
        "call_count": 0,
        "prompt_tokens": 0,
        "completion_tokens": 0,
        "total_tokens": 0,
        "estimated_cost_usd": 0.0,
        "unknown_usage_calls": 0,
        "services": {},
        "updated_at": datetime.now(timezone.utc).isoformat(),
      },
        "steps": {
            "upload_voice_and_media": "pending",
            "select_featured_image": "pending",
            "transcribe_voice": "pending",
            "refine_with_ai_chat": "pending",
            "approve_csv": "pending",
            "create_wordpress_post": "pending",
        },
    }
    write_session_state(session_id, state)
    return SessionCreateResponse(
        session_id=session_id,
        client_id=payload.client_id,
        post_type=payload.post_type,
        status="created",
        next_step="upload_voice_and_media",
    )


@app.get("/app/sessions/recent")
def list_recent_sessions(
    limit: int = Query(25, ge=1, le=200),
    client_id: str | None = Query(None),
    post_type: str | None = Query(None),
    session_status: str | None = Query(None, alias="status"),
    source: str = Query("auto", pattern="^(auto|gcs|local)$"),
    _: None = Depends(verify_api_key),
) -> dict[str, Any]:
    source_mode = source
    if source_mode == "auto":
        source_mode = "gcs" if configured_session_state_gcs_prefix() else "local"

    if source_mode == "gcs":
        rows = list_gcs_session_states()
    else:
        rows = list_local_session_states()

    wanted_client = (client_id or "").strip().lower()
    wanted_post_type = (post_type or "").strip().lower()
    wanted_status = (session_status or "").strip().lower()

    filtered: list[dict[str, Any]] = []
    for row in rows:
        state = row.get("state") or {}
        if wanted_client and str(state.get("client_id") or "").strip().lower() != wanted_client:
            continue
        if wanted_post_type and str(state.get("post_type") or "").strip().lower() != wanted_post_type:
            continue
        if wanted_status and str(state.get("status") or "").strip().lower() != wanted_status:
            continue
        filtered.append(row)

    filtered.sort(
        key=lambda item: session_sort_value(item.get("state") or {}, item.get("storage_updated_at")),
        reverse=True,
    )
    selected = filtered[:limit]

    sessions: list[dict[str, Any]] = []
    for row in selected:
        state = row.get("state") or {}
        wordpress_post = state.get("wordpress_post") or {}
        files = state.get("files") if isinstance(state.get("files"), dict) else {}
        transcript_payload = state.get("transcript") if isinstance(state.get("transcript"), dict) else {}
        draft_payload = state.get("draft") if isinstance(state.get("draft"), dict) else {}
        has_images = len([img for img in list(files.get("images") or []) if isinstance(img, dict)]) > 0
        has_transcript = bool(str(transcript_payload.get("text") or "").strip())
        has_draft = bool(str(draft_payload.get("csv_text") or "").strip())
        media_recovery = state.get("media_recovery") if isinstance(state.get("media_recovery"), dict) else {}
        missing_total = int(media_recovery.get("missing_total") or 0)
        sessions.append(
            {
                "session_id": state.get("session_id"),
                "client_id": state.get("client_id"),
                "post_type": state.get("post_type"),
                "status": state.get("status"),
                "created_at": state.get("created_at"),
                "updated_at": session_last_updated_at(state, row.get("storage_updated_at")),
                "storage": row.get("storage"),
                "has_wordpress_post": bool(wordpress_post.get("post_id")),
                "wordpress_post_id": wordpress_post.get("post_id"),
                "wordpress_edit_url": wordpress_post.get("edit_url"),
                "has_images": has_images,
                "has_transcript": has_transcript,
                "has_draft": has_draft,
                "missing_media_total": missing_total,
                "missing_media_images": int(media_recovery.get("missing_images") or 0),
                "missing_media_videos": int(media_recovery.get("missing_videos") or 0),
                "missing_media_voices": int(media_recovery.get("missing_voices") or 0),
            }
        )

    return {
        "source": source_mode,
        "filters": {
            "client_id": client_id,
            "post_type": post_type,
            "status": session_status,
        },
        "count": len(sessions),
        "sessions": sessions,
    }


@app.post("/app/sessions/delete")
def delete_sessions(
    payload: SessionsDeleteRequest,
    _: None = Depends(verify_api_key),
) -> dict[str, Any]:
    session_ids: list[str] = []
    for raw in list(payload.session_ids or []):
      cleaned = str(raw or "").strip()
      if cleaned and cleaned not in session_ids:
        session_ids.append(cleaned)

    deleted_ids: list[str] = []
    not_found_ids: list[str] = []
    errors: dict[str, str] = {}

    for session_id in session_ids:
      deleted_any = False
      local_error: str | None = None
      gcs_error: str | None = None
      try:
        if delete_local_session_state(session_id):
          deleted_any = True
      except Exception as exc:
        local_error = f"local delete failed: {exc}"

      try:
        if delete_gcs_session_state(session_id):
          deleted_any = True
      except Exception as exc:
        gcs_error = f"gcs delete failed: {exc}"

      if deleted_any:
        deleted_ids.append(session_id)
      else:
        not_found_ids.append(session_id)

      joined = "; ".join([part for part in (local_error, gcs_error) if part])
      if joined:
        errors[session_id] = joined

    return {
      "requested": len(session_ids),
      "deleted": len(deleted_ids),
      "deleted_ids": deleted_ids,
      "not_found_ids": not_found_ids,
      "errors": errors,
    }


@app.get("/app/sessions/{session_id}", response_model=SessionStateResponse)
def get_session(
    session_id: str,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    return SessionStateResponse(**read_session_state(session_id))


@app.get("/app/sessions/{session_id}/logs")
def get_session_logs(
    session_id: str,
    _: None = Depends(verify_api_key),
) -> dict[str, Any]:
    state = read_session_state(session_id)
    session_dir = APP_SESSION_ROOT / session_id
    wordpress_post = state.get("wordpress_post", {})
    output_dir_value = wordpress_post.get("output_dir")
    output_dir = Path(output_dir_value) if output_dir_value else None

    import_logs: dict[str, Any] = {}
    import_logs = collect_wordpress_import_logs(output_dir)
    if not import_logs:
      cached_logs = wordpress_post.get("import_logs")
      if isinstance(cached_logs, dict):
        import_logs = cached_logs
    cache_dir = session_dir / "cache"
    cache_files: dict[str, Any] = {}
    if cache_dir.exists():
        for path in sorted(cache_dir.iterdir()):
            if path.suffix == ".json":
                cache_files[path.name] = read_json_if_exists(path)
            elif path.is_file():
                cache_files[path.name] = path.read_text(encoding="utf-8")

    return {
        "session_id": session_id,
        "session_dir": str(session_dir),
        "state": state,
        "ui_cache": state.get("ui_cache", {}),
        "cache_files": cache_files,
        "draft_csv": state.get("draft", {}).get("csv_text"),
        "wordpress_import_output_dir": str(output_dir) if output_dir else None,
        "wordpress_import_logs": import_logs,
    }


@app.get("/app/sessions/{session_id}/images/{filename}")
def get_session_image(
    session_id: str,
    filename: str,
    background_tasks: BackgroundTasks,
    x_user_id: str | None = Header(default=None, alias="X-User-ID"),
    _: None = Depends(verify_api_key),
) -> FileResponse:
    if CONTENT_PIPELINE_VERSION == "v2":
        session = get_v2_service().require_owner(session_id, x_user_id)
        references = list(session.image_refs)
        processed = list(session.processed_images)
        source_uri = next(
            (
                str(item.get("path"))
                for item in processed
                if item.get("filename") == filename and item.get("path")
            ),
            "",
        )
        content_type = "application/octet-stream"
        if not source_uri:
            reference = next(
                (item for item in references if item.filename == filename),
                None,
            )
            if reference is None:
                raise HTTPException(status_code=404, detail="Image not found in this session.")
            source_uri = reference.storage_uri
            content_type = reference.content_type
        suffix = Path(filename).suffix or ".bin"
        fd, temporary_name = tempfile.mkstemp(prefix="v2-preview-", suffix=suffix)
        os.close(fd)
        temporary = Path(temporary_name)
        try:
            get_v2_service().object_storage.get(source_uri, temporary)
        except Exception:
            temporary.unlink(missing_ok=True)
            raise
        background_tasks.add_task(temporary.unlink, missing_ok=True)
        return FileResponse(temporary, media_type=content_type, filename=filename)
    state = read_session_state(session_id)
    images = state.get("files", {}).get("images", [])
    match = find_media_item_by_filename(images, filename)
    if not match:
        raise HTTPException(status_code=404, detail="Image not found in this session.")
    path_value = match.get("path")
    if not path_value:
        raise HTTPException(status_code=410, detail="Image metadata is incomplete for this session.")
    path = Path(path_value)
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Image file is missing.")
    return FileResponse(path)


@app.get("/app/sessions/{session_id}/images/{filename}/original")
def get_session_original_image(
    session_id: str,
    filename: str,
    background_tasks: BackgroundTasks,
    x_user_id: str | None = Header(default=None, alias="X-User-ID"),
    _: None = Depends(verify_api_key),
 ) -> FileResponse:
    if CONTENT_PIPELINE_VERSION == "v2":
        session = get_v2_service().require_owner(session_id, x_user_id)
        processed = next(
            (item for item in session.processed_images if item.get("filename") == filename),
            None,
        )
        media_id = processed.get("media_id") if processed else None
        reference = next(
            (
                item for item in session.image_refs
                if item.filename == filename or item.media_id == media_id
            ),
            None,
        )
        if reference is None:
            raise HTTPException(status_code=404, detail="Original image not found.")
        suffix = Path(reference.filename).suffix or ".bin"
        fd, temporary_name = tempfile.mkstemp(prefix="v2-original-", suffix=suffix)
        os.close(fd)
        temporary = Path(temporary_name)
        try:
            get_v2_service().object_storage.get(reference.storage_uri, temporary)
        except Exception:
            temporary.unlink(missing_ok=True)
            raise
        background_tasks.add_task(temporary.unlink, missing_ok=True)
        return FileResponse(
            temporary,
            media_type=reference.content_type,
            filename=reference.filename,
        )
    state = read_session_state(session_id)
    images = state.get("files", {}).get("images", [])
    match = find_media_item_by_filename(images, filename)
    if not match:
        raise HTTPException(status_code=404, detail="Image not found in this session.")
    original_path_value = str(match.get("original_path") or "").strip()
    if not original_path_value:
        processing = state.get("image_processing") if isinstance(state.get("image_processing"), dict) else {}
        if str(processing.get("status") or "").strip().lower() == "processing":
            process_session_images_with_pillow(session_id)
            refreshed = read_session_state(session_id)
            images = refreshed.get("files", {}).get("images", [])
            match = find_media_item_by_filename(images, filename)
            original_path_value = str((match or {}).get("original_path") or "").strip()
    if original_path_value:
      path = Path(original_path_value)
      if path.exists() and path.is_file():
        return FileResponse(path)

    # Fallback: if the original cannot be recovered, return current image so compare modal still opens.
    current_path_value = str((match or {}).get("path") or "").strip()
    if current_path_value:
      current_path = Path(current_path_value)
      if current_path.exists() and current_path.is_file():
        return FileResponse(current_path)

    raise HTTPException(status_code=404, detail="Original image file is missing.")


@app.get("/app/sessions/{session_id}/videos/{filename}")
def get_session_video(
    session_id: str,
    filename: str,
    _: None = Depends(verify_api_key),
) -> FileResponse:
    state = read_session_state(session_id)
    videos = state.get("files", {}).get("videos", [])
    match = next((item for item in videos if item.get("filename") == filename), None)
    if not match:
        raise HTTPException(status_code=404, detail="Video not found in this session.")
    path_value = match.get("path")
    if not path_value:
        raise HTTPException(status_code=410, detail="Video metadata is incomplete for this session.")
    path = Path(path_value)
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Video file is missing.")
    return FileResponse(path)


@app.put("/app/sessions/{session_id}/featured-image", response_model=SessionStateResponse)
def update_session_featured_image(
    session_id: str,
    payload: FeaturedImageRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    images = state.get("files", {}).get("images", [])
    if not any(item.get("filename") == payload.filename for item in images):
        raise HTTPException(status_code=400, detail="Featured image must match one uploaded image filename.")
    state.setdefault("files", {})["featured_image_filename"] = payload.filename
    state["steps"]["select_featured_image"] = "complete"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.put("/app/sessions/{session_id}/images/{filename}/metadata", response_model=SessionStateResponse)
def update_session_image_metadata(
    session_id: str,
    filename: str,
    payload: ImageMetadataUpdateRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    files = state.setdefault("files", {})
    images = [item for item in list(files.get("images") or []) if isinstance(item, dict)]
    match = find_media_item_by_filename(images, filename)
    if not match:
        raise HTTPException(status_code=404, detail="Image not found in this session.")

    target_filename = str(match.get("filename") or filename)
    if payload.use_suggestions:
        metadata = suggest_image_metadata_for_state(state, target_filename)
    else:
        metadata = {
            "alt_text": payload.alt_text.strip(),
            "title": payload.title.strip(),
            "caption": payload.caption.strip(),
            "description": payload.description.strip(),
        }
    match["wp_metadata"] = metadata

    state["status"] = "image_metadata_updated"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/images/{filename}/metadata/vision", response_model=SessionStateResponse)
def update_session_image_metadata_with_vision(
    session_id: str,
    filename: str,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    files = state.setdefault("files", {})
    images = [item for item in list(files.get("images") or []) if isinstance(item, dict)]
    match = find_media_item_by_filename(images, filename)
    if not match:
        raise HTTPException(status_code=404, detail="Image not found in this session.")

    try:
        vision_result = analyze_image_metadata_with_vision(state, filename)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not analyze image with Vision: {exc}") from exc

    match["wp_metadata"] = dict(vision_result.get("metadata") or {})
    match["vision_analysis"] = {
        "issues": list(vision_result.get("issues") or []),
        "advice": list(vision_result.get("advice") or []),
      "crop_focus": vision_result.get("crop_focus"),
        "model": str(vision_result.get("model") or "gpt-4o-mini"),
        "updated_at": vision_result.get("updated_at") or datetime.now(timezone.utc).isoformat(),
    }

    state["status"] = "image_metadata_vision_updated"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/images/{filename}/optimize", response_model=SessionStateResponse)
def optimize_session_image_with_openai(
    session_id: str,
    filename: str,
    payload: ImageOptimizationRequest,
    _: None = Depends(verify_api_key),
  ) -> SessionStateResponse:
    if not OPENAI_API_KEY or OpenAI is None:
      raise HTTPException(status_code=503, detail="OpenAI image optimization is unavailable because OPENAI_API_KEY or openai package is missing.")

    state = read_session_state(session_id)
    files = state.setdefault("files", {})
    images = [item for item in list(files.get("images") or []) if isinstance(item, dict)]
    match = find_media_item_by_filename(images, filename)
    if not match:
      raise HTTPException(status_code=404, detail="Image not found in this session.")

    current_path_value = str(match.get("path") or "").strip()
    if not current_path_value:
      raise HTTPException(status_code=410, detail="Image path is missing for this session item.")
    current_path = Path(current_path_value)
    if not current_path.exists() or not current_path.is_file():
      raise HTTPException(status_code=404, detail="Image file is missing.")

    original_path_value = str(match.get("original_path") or "").strip()
    original_path = Path(original_path_value) if original_path_value else None
    input_source_path = original_path if original_path and original_path.exists() and original_path.is_file() else current_path

    prompt = str(payload.prompt or "").strip()
    if not prompt:
      raise HTTPException(status_code=400, detail="Optimization prompt is required.")

    edit_input_path = input_source_path
    temp_edit_input_path: Path | None = None
    if Image is not None:
      try:
        source_image = Image.open(input_source_path)
        if source_image.mode not in {"RGB", "L"}:
          source_image = source_image.convert("RGB")
        elif source_image.mode == "L":
          source_image = source_image.convert("RGB")
        with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_file:
          temp_edit_input_path = Path(temp_file.name)
        source_image.save(temp_edit_input_path, format="JPEG", quality=95)
        edit_input_path = temp_edit_input_path
      except Exception:
        edit_input_path = input_source_path

    try:
      client = OpenAI(api_key=OPENAI_API_KEY)
      with edit_input_path.open("rb") as image_file:
        response = client.images.edit(
          model="gpt-image-1",
          image=image_file,
          prompt=prompt,
        )
    except Exception as exc:
      raise HTTPException(status_code=500, detail=f"OpenAI image optimization failed: {exc}") from exc
    finally:
      if temp_edit_input_path and temp_edit_input_path.exists():
        try:
          temp_edit_input_path.unlink()
        except Exception:
          pass

    track_ai_usage_from_response(
      state,
      call_name="image_optimize",
      model="gpt-image-1",
      service="openai_images",
      response=response,
    )

    data_items = getattr(response, "data", None) or []
    if not data_items:
      raise HTTPException(status_code=500, detail="OpenAI image optimization returned no image data.")

    first_item = data_items[0]
    b64_data = getattr(first_item, "b64_json", None)
    if not b64_data and isinstance(first_item, dict):
      b64_data = first_item.get("b64_json")
    image_url = getattr(first_item, "url", None)
    if not image_url and isinstance(first_item, dict):
      image_url = first_item.get("url")
    if not b64_data and not image_url:
      raise HTTPException(status_code=500, detail="OpenAI image optimization response did not include image content.")

    try:
      if b64_data:
        optimized_bytes = base64.b64decode(b64_data)
      else:
        with urlopen(str(image_url), timeout=60) as response_stream:
          optimized_bytes = response_stream.read()
    except Exception as exc:
      raise HTTPException(status_code=500, detail=f"Could not decode optimized image output: {exc}") from exc

    try:
      if Image is None:
        current_path.write_bytes(optimized_bytes)
      else:
        optimized_image = Image.open(BytesIO(optimized_bytes))
        if optimized_image.mode not in {"RGB", "L"}:
          optimized_image = optimized_image.convert("RGB")
        elif optimized_image.mode == "L":
          optimized_image = optimized_image.convert("RGB")

        suffix = current_path.suffix.lower()
        if suffix == ".webp":
          optimized_image.save(current_path, format="WEBP", quality=90, method=6)
        elif suffix in {".jpg", ".jpeg"}:
          optimized_image.save(current_path, format="JPEG", quality=92, optimize=True, progressive=True)
        elif suffix == ".png":
          optimized_image.save(current_path, format="PNG", optimize=True)
        else:
          optimized_image.save(current_path, format="WEBP", quality=90, method=6)
    except Exception as exc:
      raise HTTPException(status_code=500, detail=f"Could not write optimized image output: {exc}") from exc

    match["size"] = current_path.stat().st_size
    match["content_type"] = detect_image_mime(current_path)
    match["processed_at"] = datetime.now(timezone.utc).isoformat()
    operations = [str(op) for op in list(match.get("applied_operations") or []) if str(op).strip()]
    operations.append("openai_image_optimize")
    match["applied_operations"] = operations[-20:]
    match["image_optimization"] = {
      "prompt": prompt,
      "model": "gpt-image-1",
      "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    state["status"] = "image_optimized_with_openai"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/images/{filename}/restore-original", response_model=SessionStateResponse)
def restore_session_image_original(
    session_id: str,
    filename: str,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    files = state.setdefault("files", {})
    images = [item for item in list(files.get("images") or []) if isinstance(item, dict)]
    match = find_media_item_by_filename(images, filename)
    if not match:
      raise HTTPException(status_code=404, detail="Image not found in this session.")

    current_path_value = str(match.get("path") or "").strip()
    original_path_value = str(match.get("original_path") or "").strip()
    if not current_path_value:
      raise HTTPException(status_code=410, detail="Image path is missing for this session item.")
    if not original_path_value:
      raise HTTPException(status_code=404, detail="Original image is not available for this file.")

    current_path = Path(current_path_value)
    original_path = Path(original_path_value)
    if not current_path.exists() or not current_path.is_file():
      raise HTTPException(status_code=404, detail="Current image file is missing.")
    if not original_path.exists() or not original_path.is_file():
      raise HTTPException(status_code=404, detail="Original image file is missing.")

    try:
      if Image is None:
        current_path.write_bytes(original_path.read_bytes())
      else:
        restored = Image.open(original_path)
        if restored.mode not in {"RGB", "L"}:
          restored = restored.convert("RGB")
        elif restored.mode == "L":
          restored = restored.convert("RGB")

        suffix = current_path.suffix.lower()
        if suffix == ".webp":
          restored.save(current_path, format="WEBP", quality=90, method=6)
        elif suffix in {".jpg", ".jpeg"}:
          restored.save(current_path, format="JPEG", quality=92, optimize=True, progressive=True)
        elif suffix == ".png":
          restored.save(current_path, format="PNG", optimize=True)
        else:
          restored.save(current_path, format="WEBP", quality=90, method=6)
    except Exception as exc:
      raise HTTPException(status_code=500, detail=f"Could not restore original image: {exc}") from exc

    match["size"] = current_path.stat().st_size
    match["content_type"] = detect_image_mime(current_path)
    match["processed_at"] = datetime.now(timezone.utc).isoformat()
    operations = [str(op) for op in list(match.get("applied_operations") or []) if str(op).strip()]
    operations.append("restore_original")
    match["applied_operations"] = operations[-20:]
    match["image_optimization"] = {
      "restored_from_original": True,
      "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    state["status"] = "image_restored_to_original"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.put("/app/sessions/{session_id}/vision-selection", response_model=SessionStateResponse)
def update_session_vision_selection(
    session_id: str,
    payload: VisionSelectionRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    files = state.setdefault("files", {})
    images = [item for item in list(files.get("images") or []) if isinstance(item, dict)]
    available_names = {str(item.get("filename") or "").strip() for item in images}
    selected = []
    for name in payload.filenames:
      cleaned = str(name or "").strip()
      if cleaned and cleaned in available_names and cleaned not in selected:
        selected.append(cleaned)
    files["vision_selected_filenames"] = selected
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/vision/analyze")
def analyze_session_vision(
    session_id: str,
    _: None = Depends(verify_api_key),
) -> dict[str, Any]:
    state = read_session_state(session_id)
    vision = analyze_selected_images_with_vision(session_id)
    return {
      "session_id": session_id,
      "vision": vision,
      "selected": selected_vision_filenames(state),
    }


@app.post("/app/sessions/{session_id}/images/process")
def trigger_session_image_processing(
    session_id: str,
    _: None = Depends(verify_api_key),
) -> dict[str, Any]:
    process_session_images_with_pillow(session_id)
    state = read_session_state(session_id)
    return {
      "session_id": session_id,
      "image_processing": state.get("image_processing", {}),
      "images": state.get("files", {}).get("images", []),
    }


@app.post("/app/sessions/{session_id}/wordpress/media-library-upload")
def upload_session_images_media_library(
    session_id: str,
    _: None = Depends(verify_api_key),
) -> dict[str, Any]:
    result = upload_session_images_to_wordpress_media_library(session_id)
    return {
      "session_id": session_id,
      "status": "wordpress_media_uploaded",
      "wordpress_media_library": result,
    }


@app.delete("/app/sessions/{session_id}/media/{media_type}/{filename}", response_model=SessionStateResponse)
def delete_session_media(
    session_id: str,
    media_type: str,
    filename: str,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    if media_type not in {"images", "videos", "voices"}:
        raise HTTPException(status_code=400, detail="Unsupported media type.")
    state = read_session_state(session_id)
    files = state.setdefault("files", {})
    items = list(files.get(media_type, []))
    match = next((item for item in items if item.get("filename") == filename), None)
    if not match:
        raise HTTPException(status_code=404, detail="Media file not found in this session.")

    path = Path(match["path"])
    if path.exists() and path.is_file():
        path.unlink()
    remaining = [item for item in items if item.get("filename") != filename]
    files[media_type] = remaining

    if media_type == "voices":
        files["voice"] = remaining[0] if remaining else None
        state["steps"]["upload_voice_and_media"] = "complete" if remaining else "pending"
        state["steps"]["transcribe_voice"] = "pending"
    elif media_type == "images":
        current_featured = files.get("featured_image_filename")
        if current_featured == filename or not any(item.get("filename") == current_featured for item in remaining):
            files["featured_image_filename"] = remaining[0]["filename"] if remaining else None
        state["steps"]["select_featured_image"] = "complete" if files.get("featured_image_filename") else "pending"
    elif media_type == "videos":
        files["selected_video_filename"] = remaining[0]["filename"] if remaining else None

    state["status"] = "files_uploaded"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.put("/app/sessions/{session_id}/ui-cache", response_model=SessionStateResponse)
def update_session_ui_cache(
    session_id: str,
    payload: UiCacheUpdateRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    state["ui_cache"] = {
        **payload.cache,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/uploads", response_model=SessionStateResponse)
async def upload_session_files(
  background_tasks: BackgroundTasks,
    session_id: str,
    voices: list[UploadFile] | None = File(default=None),
    images: list[UploadFile] | None = File(default=None),
    videos: list[UploadFile] | None = File(default=None),
    featured_image_filename: str | None = Form(default=None),
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    session_dir = APP_SESSION_ROOT / session_id
    existing_files = state.get("files", {})

    voice_infos = list(existing_files.get("voices") or ([existing_files["voice"]] if existing_files.get("voice") else []))
    for index, voice in enumerate(voices or [], start=1):
        voice_info = save_upload_file(voice, session_dir / "voice", f"voice_{len(voice_infos) + index}")
        validate_extension(voice_info, AUDIO_EXTENSIONS, "voice")
        voice_infos.append(voice_info)

    image_infos = list(existing_files.get("images", []))
    for index, image in enumerate(images or [], start=1):
        image_info = save_upload_file(image, session_dir / "images", f"image_{len(image_infos) + index}.jpg")
        validate_extension(image_info, IMAGE_EXTENSIONS, "image")
        image_infos.append(image_info)

    video_infos = []
    for index, video in enumerate(videos or [], start=1):
        video_info = save_upload_file(video, session_dir / "videos", f"video_{index}.mp4")
        validate_extension(video_info, VIDEO_EXTENSIONS, "video")
        video_infos.append(video_info)
    if not video_infos:
        video_infos = list(existing_files.get("videos", []))

    if len(video_infos) > 1:
        raise HTTPException(status_code=400, detail="Only one video per post is supported for now.")

    if not voice_infos and not image_infos and not video_infos:
        raise HTTPException(status_code=400, detail="Upload at least one media file.")

    if featured_image_filename:
        featured_match = next(
            (
                item
                for item in image_infos
                if item["original_filename"] == featured_image_filename
                or item["filename"] == safe_upload_name(featured_image_filename, "featured_image")
            ),
            None,
        )
        if not featured_match:
            raise HTTPException(status_code=400, detail="Featured image must match one uploaded image filename.")
        featured_image_filename = featured_match["filename"]

    effective_featured_image = featured_image_filename or existing_files.get("featured_image_filename")
    previous_vision_selection = [
      str(name).strip()
      for name in list(existing_files.get("vision_selected_filenames") or [])
      if str(name).strip()
    ]
    next_image_names = {str(item.get("filename") or "").strip() for item in image_infos}
    state["files"] = {
        "voice": voice_infos[0] if voice_infos else None,
        "voices": voice_infos,
        "images": image_infos,
        "videos": video_infos,
        "featured_image_filename": effective_featured_image,
        "selected_video_filename": video_infos[0]["filename"] if video_infos else None,
      "vision_selected_filenames": [name for name in previous_vision_selection if name in next_image_names],
    }
    state["status"] = "files_uploaded"
    if images:
      state["image_processing"] = {
        "status": "processing",
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "message": "Pillow processing in progress.",
      }
    state["steps"]["upload_voice_and_media"] = "complete"
    state["steps"]["select_featured_image"] = "complete" if effective_featured_image else "pending"
    state["steps"]["transcribe_voice"] = "pending"
    upload_session_media_to_gcs(session_id, state)
    write_session_state(session_id, state)
    if images:
      background_tasks.add_task(process_session_images_with_pillow, session_id)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/transcribe", response_model=SessionStateResponse)
def transcribe_session_voice(
    session_id: str,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    files = state.get("files", {})
    voices = files.get("voices") or ([files["voice"]] if files.get("voice") else [])
    if not voices:
        raise HTTPException(status_code=400, detail="Upload a voice file before transcription.")

    transcript_parts = []
    transcript_items = []
    for index, voice in enumerate(voices, start=1):
        raw_text = transcribe_audio_file(voice["path"], model=DEFAULT_TRANSCRIPTION_MODEL)
        cleaned_text = sanitize_transcript_text(raw_text)
        track_ai_usage(
            state,
            call_name="voice_transcription",
            model=DEFAULT_TRANSCRIPTION_MODEL,
          service="openai_transcription",
            unknown_usage=True,
        )
        transcript_parts.append(cleaned_text.strip())
        transcript_items.append({
            "index": index,
            "filename": voice.get("original_filename") or voice.get("filename"),
            "text": cleaned_text,
            "raw_text": raw_text,
        })
    transcript_text = "\n\n".join(part for part in transcript_parts if part)
    state["transcript"] = {
        "text": transcript_text,
        "source": "openai_transcription_rule_cleaned",
        "model": DEFAULT_TRANSCRIPTION_MODEL,
        "items": transcript_items,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    state["status"] = "transcribed"
    state["steps"]["transcribe_voice"] = "complete"
    state["steps"]["refine_with_ai_chat"] = "pending"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/draft/chat/transcribe", response_model=DraftChatTranscriptionResponse)
async def transcribe_draft_chat_voice(
    session_id: str,
    voice: UploadFile = File(...),
    _: None = Depends(verify_api_key),
) -> DraftChatTranscriptionResponse:
    read_session_state(session_id)
    suffix = Path(voice.filename or "draft-chat.webm").suffix.lower() or ".webm"
    if suffix not in AUDIO_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported audio file extension: {voice.filename or 'draft-chat.webm'}")

    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
            temp_path = Path(temp_file.name)
            while chunk := await voice.read(1024 * 1024):
                temp_file.write(chunk)

        raw_text = await run_in_threadpool(transcribe_audio_file, temp_path, DEFAULT_TRANSCRIPTION_MODEL)
        cleaned_text = sanitize_transcript_text(raw_text)
        state = read_session_state(session_id)
        track_ai_usage(
          state,
          call_name="draft_chat_transcription",
          model=DEFAULT_TRANSCRIPTION_MODEL,
          service="openai_transcription",
          unknown_usage=True,
        )
        write_session_state(session_id, state)
        return DraftChatTranscriptionResponse(
            text=cleaned_text,
            raw_text=raw_text,
            source="draft_chat_audio_transcription",
            model=DEFAULT_TRANSCRIPTION_MODEL,
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not transcribe draft chat audio: {exc}") from exc
    finally:
        await voice.close()
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)


@app.put("/app/sessions/{session_id}/transcript", response_model=SessionStateResponse)
def update_session_transcript(
    session_id: str,
    payload: TranscriptUpdateRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    state["transcript"] = {
        "text": payload.text,
        "source": "manual_update",
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    state["status"] = "transcript_updated"
    state["steps"]["transcribe_voice"] = "complete"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/draft", response_model=SessionStateResponse)
async def generate_session_draft(
    session_id: str,
    payload: DraftGenerateRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    session_dir = APP_SESSION_ROOT / session_id
    if not state.get("files", {}).get("images"):
        raise HTTPException(status_code=400, detail="Upload at least one image before generating a draft.")

    try:
        draft_state = state_with_vision_context(state)
        specs = await run_in_threadpool(load_workbook_specs, state.get("post_type"))
        workbook_path = await run_in_threadpool(active_knowledge_workbook_path)
        guidance_data = get_session_guidance_cached(
            state,
            workbook_path,
            state.get("post_type"),
            KNOWLEDGE_WORKBOOK_SHEET,
        )
        internal_links_context = get_session_internal_links_cached(state, workbook_path)
        draft = await run_in_threadpool(
          create_session_draft,
          session_dir,
          draft_state,
          specs,
          payload.category,
          payload.force_regenerate,
          guidance_data,
          internal_links_context,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not generate draft: {exc}") from exc

    draft_usage_events = draft.pop("_usage_events", []) if isinstance(draft, dict) else []
    track_ai_usage_events(state, draft_usage_events, default_service="openai_text")
    state["draft"] = draft
    state["status"] = "draft_generated"
    state["steps"]["refine_with_ai_chat"] = "complete"
    state["steps"]["approve_csv"] = "pending"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.put("/app/sessions/{session_id}/draft", response_model=SessionStateResponse)
def update_session_draft(
    session_id: str,
    payload: DraftUpdateRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    if not state.get("draft"):
        raise HTTPException(status_code=400, detail="Generate a draft before saving draft edits.")

    session_dir = APP_SESSION_ROOT / session_id
    zip_path = rebuild_session_package(session_dir, payload.csv_text, state)
    state["draft"]["csv_text"] = payload.csv_text
    state["draft"]["zip_path"] = str(zip_path)
    state["draft"]["updated_at"] = datetime.now(timezone.utc).isoformat()
    state["status"] = "draft_updated"
    state["steps"]["approve_csv"] = "pending"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/draft/chat", response_model=SessionStateResponse)
async def chat_with_draft_agent(
    session_id: str,
    payload: DraftChatRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    message = payload.message.strip()
    if not message:
        raise HTTPException(status_code=400, detail="Chat message is required.")

    state = read_session_state(session_id)
    if not state.get("draft"):
        raise HTTPException(status_code=400, detail="Generate a draft before chatting with the draft agent.")

    session_dir = APP_SESSION_ROOT / session_id
    try:
        draft_state = state_with_vision_context(state)
        specs = await run_in_threadpool(load_workbook_specs, state.get("post_type"))
        workbook_path = await run_in_threadpool(active_knowledge_workbook_path)
        guidance_data = get_session_guidance_cached(
            state,
            workbook_path,
            state.get("post_type"),
            KNOWLEDGE_WORKBOOK_SHEET,
        )
        internal_links_context = get_session_internal_links_cached(state, workbook_path)
        draft = await run_in_threadpool(
          revise_session_draft,
          session_dir,
          draft_state,
          specs,
          message,
          guidance_data,
          internal_links_context,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not revise draft: {exc}") from exc

    draft_usage_events = draft.pop("_usage_events", []) if isinstance(draft, dict) else []
    track_ai_usage_events(state, draft_usage_events, default_service="openai_text")
    state["draft"] = draft
    state["status"] = "draft_revised"
    state["steps"]["refine_with_ai_chat"] = "complete"
    state["steps"]["approve_csv"] = "pending"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/draft/refine", response_model=SessionStateResponse)
async def refine_draft_with_answers(
    session_id: str,
    payload: DraftRefinementRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    """Apply user answers to refinement questions and update draft."""
    state = read_session_state(session_id)
    if not state.get("draft"):
        raise HTTPException(status_code=400, detail="Generate a draft before refining.")

    if not payload.answers:
        raise HTTPException(status_code=400, detail="Provide at least one answer.")

    session_dir = APP_SESSION_ROOT / session_id
    transcript = state.get("transcript", {}).get("text", "").strip()

    try:
        specs = await run_in_threadpool(load_workbook_specs, state.get("post_type"))
        draft_state = state_with_vision_context(state)
        draft = await run_in_threadpool(
            apply_refinement_answers,
            session_dir,
            draft_state,
            specs,
            payload.answers,
            transcript,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not refine draft: {exc}") from exc

    state["draft"] = draft
    state["status"] = "draft_refined"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/app/sessions/{session_id}/wordpress-post", response_model=SessionStateResponse)
async def create_session_wordpress_post(
    session_id: str,
    payload: CreateWordPressPostRequest,
    _: None = Depends(verify_api_key),
) -> SessionStateResponse:
    state = read_session_state(session_id)
    draft = state.get("draft", {})
    csv_text = draft.get("csv_text")
    if not csv_text:
        raise HTTPException(status_code=400, detail="Generate and approve a draft before creating the post.")

    session_dir = APP_SESSION_ROOT / session_id
    zip_path = rebuild_session_package(session_dir, csv_text, state)
    allowed_statuses = {"draft", "publish", "pending", "private"}
    if payload.status not in allowed_statuses:
        raise HTTPException(status_code=400, detail=f"Unsupported post status: {payload.status}")

    args = build_import_args(
        zip_path=zip_path,
        event_name=session_id,
        status=payload.status,
        row=0,
        required_category="auto event post",
        existing_post_mode=payload.existing_post_mode,
        existing_post_id=None,
        client_id=state.get("client_id", "flairlab"),
    )
    if payload.update_existing_generated_post:
        previous_post_id = state.get("wordpress_post", {}).get("post_id")
        if not previous_post_id:
            raise HTTPException(status_code=400, detail="No previously generated post found in this session.")
        args.existing_post_id = int(previous_post_id)
    elif payload.existing_post_id:
        args.existing_post_id = int(payload.existing_post_id)
    args.output_root = APP_SESSION_ROOT / session_id / "wordpress_imports"

    try:
        output_dir = await run_in_threadpool(run_import, args)
        post_response = await run_in_threadpool(post_response_from_output, output_dir, state.get("client_id", "flairlab"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not create WordPress post: {exc}") from exc

    import_logs = collect_wordpress_import_logs(output_dir)
    metadata_injection = await run_in_threadpool(
      apply_session_image_metadata_to_wordpress_assets,
      state,
      post_response,
      import_logs,
    )
    post_response["media_metadata_injection"] = metadata_injection
    post_response["import_logs"] = import_logs

    state["draft"]["zip_path"] = str(zip_path)
    state["wordpress_post"] = post_response
    state["status"] = "wordpress_post_created"
    state["steps"]["approve_csv"] = "complete"
    state["steps"]["create_wordpress_post"] = "complete"
    write_session_state(session_id, state)
    return SessionStateResponse(**state)


@app.post("/event-posts/from-zip")
async def action_event_post_from_zip(
    payload: EventPostActionRequest,
    _: None = Depends(verify_api_key),
):
    return await import_event_post_from_zip(payload)


app.mount("/action", action_app)
