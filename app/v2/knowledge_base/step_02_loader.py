from __future__ import annotations

import hashlib
from datetime import datetime, timezone
from pathlib import Path
from threading import RLock
from typing import Any, TypeVar

from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError

from app.v2.errors import ErrorDetail, InvalidWorkbookError
from app.v2.knowledge_base.step_01_models import (
    ACFFieldSchema,
    AgentInstruction,
    ApplicationState,
    BlueprintRow,
    ContextManifestRow,
    ImageAnalysisRule,
    ImageMetadataField,
    ImageMetadataRule,
    InternalLinkRecord,
    InternalLinkRule,
    OutputSpecification,
    PillowRule,
    PostExample,
    PostTypeConfig,
    SEORule,
    SharedFieldSchema,
    StoryPattern,
    StyleRule,
    ValidationListValue,
    WorkbookSnapshot,
    WorkbookVersion,
    WorkflowStep,
)

ModelT = TypeVar("ModelT", bound=BaseModel)

REQUIRED_SHEETS = {
    "agent_workflow",
    "post_types",
    "shared_fields_schema",
    "seo_rules",
    "ACF_fields_schema",
    "post_blueprint",
    "story_patterns",
    "style_rules",
    "image_rules_pillow",
    "image_analysis_rules",
    "image_metadata_schema",
    "image_metadata_rules",
    "internal_links_database",
    "internal_link_rules",
    "output_specification",
    "validation_lists",
    "application_state",
    "post_examples",
    "agent_instructions",
    "context_building",
}

LIST_COLUMNS = {
    "source_fact_keys",
    "context_tags",
    "trigger_tags",
    "required_facts",
    "required_content_signals",
    "excluded_tags",
    "semantic_prompt_hints",
    "anchor_variants",
    "trigger_value",
    "target_field_keys",
    "allowed_next_states",
    "required_data",
}

BOOLEAN_COLUMNS = {
    "enabled",
    "approved",
    "required",
    "active",
    "template_ready",
    "generation_enabled",
    "user_selectable",
    "required_for_output",
    "required_for_analysis",
    "include_in_ai_schema",
    "include_in_payload",
    "include_in_image_metadata_context",
    "terminal",
}


def _split_list(value: Any) -> tuple[str, ...]:
    if value is None:
        return ()
    return tuple(part.strip() for part in str(value).split(";") if part.strip())


def _parse_boolean(value: Any, *, sheet: str, row: int, column: str) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().upper()
    if normalized == "TRUE":
        return True
    if normalized == "FALSE":
        return False
    raise InvalidWorkbookError(
        "Workbook contains an invalid boolean.",
        details=[
            ErrorDetail(
                sheet=sheet,
                row=row,
                column=column,
                error_code="invalid_boolean",
                message=f"Expected TRUE/FALSE or an Excel boolean, received {value!r}.",
            )
        ],
    )


class WorkbookLoader:
    """Load an immutable, typed workbook snapshot and cache it by SHA-256."""

    def __init__(self) -> None:
        self._cache: dict[str, WorkbookSnapshot] = {}
        self._lock = RLock()

    def load(self, path: str | Path) -> WorkbookSnapshot:
        workbook_path = Path(path).expanduser().resolve()
        if not workbook_path.is_file():
            raise InvalidWorkbookError(
                f"V2 workbook not found: {workbook_path}",
                details=[
                    ErrorDetail(
                        sheet=None,
                        row=None,
                        column=None,
                        error_code="workbook_not_found",
                        message=str(workbook_path),
                    )
                ],
            )
        digest = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
        with self._lock:
            cached = self._cache.get(digest)
            if cached is not None:
                return cached

        workbook = load_workbook(workbook_path, read_only=False, data_only=True, keep_vba=True)
        missing = sorted(REQUIRED_SHEETS.difference(workbook.sheetnames))
        if missing:
            raise InvalidWorkbookError(
                "Workbook is missing required sheets.",
                details=[
                    ErrorDetail(
                        sheet=name,
                        row=None,
                        column=None,
                        error_code="missing_required_sheet",
                        message=f"Required sheet {name!r} is missing.",
                    )
                    for name in missing
                ],
            )

        snapshot = WorkbookSnapshot(
            version=WorkbookVersion(
                filename=workbook_path.name,
                sha256=digest,
                loaded_at=datetime.now(timezone.utc),
                schema_version=self._schema_version(workbook),
            ),
            post_types=self._models(workbook["post_types"], PostTypeConfig),
            shared_fields=self._models(workbook["shared_fields_schema"], SharedFieldSchema),
            acf_fields=self._models(workbook["ACF_fields_schema"], ACFFieldSchema),
            blueprint=self._models(workbook["post_blueprint"], BlueprintRow),
            seo_rules=self._models(workbook["seo_rules"], SEORule),
            style_rules=self._models(workbook["style_rules"], StyleRule),
            story_patterns=self._models(workbook["story_patterns"], StoryPattern),
            image_analysis_rules=self._models(workbook["image_analysis_rules"], ImageAnalysisRule),
            pillow_rules=self._models(workbook["image_rules_pillow"], PillowRule),
            image_metadata_fields=self._models(workbook["image_metadata_schema"], ImageMetadataField),
            image_metadata_rules=self._models(workbook["image_metadata_rules"], ImageMetadataRule),
            internal_links=self._models(workbook["internal_links_database"], InternalLinkRecord),
            internal_link_rules=self._models(workbook["internal_link_rules"], InternalLinkRule),
            workflow_steps=self._models(workbook["agent_workflow"], WorkflowStep),
            application_states=self._models(workbook["application_state"], ApplicationState),
            agent_instructions=self._models(workbook["agent_instructions"], AgentInstruction),
            context_manifest=self._models(workbook["context_building"], ContextManifestRow),
            validation_values=self._models(workbook["validation_lists"], ValidationListValue),
            post_examples=self._models(workbook["post_examples"], PostExample),
            output_specifications=self._models(workbook["output_specification"], OutputSpecification),
        )
        with self._lock:
            self._cache[digest] = snapshot
        return snapshot

    @staticmethod
    def _schema_version(workbook: Any) -> str | None:
        if "README" not in workbook.sheetnames:
            return None
        for row in workbook["README"].iter_rows(values_only=True):
            if row and str(row[0] or "").strip().lower() in {"schema_version", "schema version"}:
                return str(row[1] or "").strip() or None
        return None

    def _models(self, worksheet: Any, model: type[ModelT]) -> tuple[ModelT, ...]:
        headers = [str(cell.value).strip() if cell.value is not None else None for cell in worksheet[1]]
        required_headers = set(model.model_fields).difference({"sheet_row"})
        missing_headers = sorted(required_headers.difference(header for header in headers if header))
        if missing_headers:
            raise InvalidWorkbookError(
                f"Sheet {worksheet.title!r} is missing required columns.",
                details=[
                    ErrorDetail(
                        sheet=worksheet.title,
                        row=1,
                        column=column,
                        error_code="missing_required_column",
                        message=f"Required column {column!r} is missing.",
                    )
                    for column in missing_headers
                ],
            )

        parsed: list[ModelT] = []
        for row_number in range(2, worksheet.max_row + 1):
            raw = {
                header: worksheet.cell(row_number, column_number).value
                for column_number, header in enumerate(headers, 1)
                if header
            }
            if not any(value is not None for value in raw.values()):
                continue
            normalized = self._normalize(raw, worksheet.title, row_number)
            normalized["sheet_row"] = row_number
            try:
                parsed.append(model.model_validate(normalized))
            except ValidationError as exc:
                raise InvalidWorkbookError(
                    f"Sheet {worksheet.title!r} contains invalid typed data.",
                    details=[
                        ErrorDetail(
                            sheet=worksheet.title,
                            row=row_number,
                            column=".".join(str(part) for part in issue["loc"]),
                            error_code="invalid_typed_value",
                            message=issue["msg"],
                            context={"input": issue.get("input")},
                        )
                        for issue in exc.errors()
                    ],
                ) from exc
        return tuple(parsed)

    @staticmethod
    def _normalize(raw: dict[str, Any], sheet: str, row: int) -> dict[str, Any]:
        normalized = dict(raw)
        for column in LIST_COLUMNS.intersection(normalized):
            normalized[column] = _split_list(normalized[column])
        for column in BOOLEAN_COLUMNS.intersection(normalized):
            if normalized[column] is not None:
                normalized[column] = _parse_boolean(
                    normalized[column],
                    sheet=sheet,
                    row=row,
                    column=column,
                )
        if "example_id" in normalized and normalized["example_id"] is not None:
            normalized["example_id"] = str(normalized["example_id"]).removesuffix(".0")
        return normalized
