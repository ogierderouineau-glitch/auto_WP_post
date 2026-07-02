from __future__ import annotations

from datetime import datetime, timezone
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from app.v2.errors import InvalidWorkbookError
from app.v2.knowledge_base.step_01_models import WorkbookSnapshot, WorkbookVersion
from app.v2.knowledge_base.step_02_loader import WorkbookLoader
from app.v2.knowledge_base.step_03_validator import WorkbookValidator
from app.v2.knowledge_base.step_04_service import KnowledgeBaseService
from app.v2.workflow.step_03_registry import WORKFLOW_HANDLER_METHODS

WORKBOOK = Path(
    os.getenv(
        "V2_TEST_WORKBOOK",
        "/home/ogier-derouineau/Documents/FLAIRLAB_Knowledge_Base_Revised_V6.xlsm",
    )
)


class KnowledgeBaseServiceTests(unittest.TestCase):
    def _snapshot(self, sha256: str) -> WorkbookSnapshot:
        return WorkbookSnapshot(
            version=WorkbookVersion(
                filename="test.xlsm",
                sha256=sha256,
                loaded_at=datetime.now(timezone.utc),
                schema_version=None,
            ),
            post_types=(),
            shared_fields=(),
            acf_fields=(),
            blueprint=(),
            seo_rules=(),
            style_rules=(),
            story_patterns=(),
            image_analysis_rules=(),
            pillow_rules=(),
            image_metadata_fields=(),
            image_metadata_rules=(),
            internal_links=(),
            internal_link_rules=(),
            workflow_steps=(),
            application_states=(),
            agent_instructions=(),
            context_manifest=(),
            validation_values=(),
            post_examples=(),
            output_specifications=(),
        )

    def test_missing_workbook_hash_falls_back_to_current_snapshot(self) -> None:
        service = KnowledgeBaseService("unused.xlsm")
        current = self._snapshot("current-hash")
        service._snapshots[current.version.sha256] = current
        service._current_hash = current.version.sha256

        self.assertIs(service.by_hash("missing-hash"), current)

    def test_missing_workbook_hash_fallback_can_be_disabled(self) -> None:
        service = KnowledgeBaseService("unused.xlsm")
        current = self._snapshot("current-hash")
        service._snapshots[current.version.sha256] = current
        service._current_hash = current.version.sha256

        with patch.dict(os.environ, {"V2_ALLOW_CURRENT_WORKBOOK_FALLBACK": "0"}):
            with self.assertRaises(KeyError):
                service.by_hash("missing-hash")


@unittest.skipUnless(WORKBOOK.is_file(), f"V2 test workbook not found: {WORKBOOK}")
class WorkbookTests(unittest.TestCase):
    def setUp(self) -> None:
        self.loader = WorkbookLoader()
        self.validator = WorkbookValidator()

    def test_v6_loads_and_all_exact_joins_validate(self) -> None:
        snapshot = self.validator.validate(self.loader.load(WORKBOOK))
        self.assertEqual(snapshot.version.sha256, "dd6b6c339e9b95e691a2405dcbe502d439396b357a38f08199e16b64da5adc4a")
        self.assertEqual(len(snapshot.workflow_steps), 19)
        self.assertEqual(len(snapshot.blueprint), 14)
        self.assertEqual(len(snapshot.image_metadata_rules), 2)
        self.assertEqual(
            {row.step_key for row in snapshot.workflow_steps},
            set(WORKFLOW_HANDLER_METHODS),
        )
        event_story = next(row for row in snapshot.acf_fields if row.field_key == "event_story")
        self.assertEqual((event_story.min_words, event_story.max_words), (80, 100))

    def test_configured_fields_are_excluded_from_ai_schema(self) -> None:
        snapshot = self.validator.validate(self.loader.load(WORKBOOK))
        configured = [row for row in snapshot.shared_fields if row.source_mode == "configured"]
        self.assertTrue(configured)
        self.assertTrue(all(not row.include_in_ai_schema for row in configured))

    def test_duplicate_active_url_fails(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            copy = Path(temporary) / "duplicate.xlsm"
            copy.write_bytes(WORKBOOK.read_bytes())
            workbook = load_workbook(copy, keep_vba=True)
            sheet = workbook["internal_links_database"]
            target_url_column = next(
                cell.column for cell in sheet[1] if cell.value == "target_url"
            )
            active_column = next(cell.column for cell in sheet[1] if cell.value == "active")
            sheet.cell(3, target_url_column).value = sheet.cell(2, target_url_column).value
            sheet.cell(3, active_column).value = True
            workbook.save(copy)
            with self.assertRaises(InvalidWorkbookError) as raised:
                self.validator.validate(self.loader.load(copy))
            codes = {detail.error_code for detail in raised.exception.details}
            self.assertIn("duplicate_active_url", codes)

    def test_unknown_source_fact_key_fails(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            copy = Path(temporary) / "bad-source.xlsm"
            copy.write_bytes(WORKBOOK.read_bytes())
            workbook = load_workbook(copy, keep_vba=True)
            sheet = workbook["ACF_fields_schema"]
            field_key_column = next(cell.column for cell in sheet[1] if cell.value == "field_key")
            source_column = next(cell.column for cell in sheet[1] if cell.value == "source_fact_keys")
            row = next(
                index
                for index in range(2, sheet.max_row + 1)
                if sheet.cell(index, field_key_column).value == "fact_event"
            )
            sheet.cell(row, source_column).value = "not_a_fact"
            workbook.save(copy)
            with self.assertRaises(InvalidWorkbookError) as raised:
                self.validator.validate(self.loader.load(copy))
            codes = {detail.error_code for detail in raised.exception.details}
            self.assertIn("unknown_source_fact_key", codes)
